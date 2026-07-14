# LayoutEngine.py

import os
import re
import json
import shutil
from copy import copy
from typing import List, Set, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ScannerPremissas import scan_orange_context, normalize_val
from BlockMapper import scan_plant, flood_fill

# Configurações de planilha e constantes globais
SHEET_NAME = 'JPIII'
FORBIDDEN_PATTERNS = {'SALA 1', 'SALA 2', 'SALA 3', 'SALA 4', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4'}

FILL_LIBERADO = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
FILL_NEW_CLIENT = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
FILL_SALA = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

FONT_WHITE = Font(color='FFFFFF', bold=True, size=8)
FONT_SMALL = Font(size=8)
FONT_SALA = Font(color='0000FF', bold=True, size=8)

THIN_BLUE = Side(border_style="thin", color="0000FF")
BORDER_SALA = Border(top=THIN_BLUE, left=THIN_BLUE, right=THIN_BLUE, bottom=THIN_BLUE)

_bench_partners_cache = {}

# ══════════════════════════════════════════════════════════════════════════
# Estruturas de Validação de Dados Internas (Pydantic com coerção automática)
# ══════════════════════════════════════════════════════════════════════════

from pydantic import BaseModel

class Acao(BaseModel):
    """Representa a estrutura de uma ação de alteração física da planta"""
    tipo: str  # "liberar", "realocar" (posicionar) ou "transferir" (swap)
    cliente: str = ""
    quantidade: int = 0
    bloco: str = ""
    ambiente: str = ""
    novo_cliente: str = ""  
    cliente_a_liberar: str = ""  
    sala_lugares: int = 0
    
    # Atributos específicos para a função única de transferência (Organizador)
    cliente_a: str = ""
    bloco_a: str = ""
    ambiente_a: str = ""
    quantidade_a: int = 0
    cliente_b: str = ""
    bloco_b: str = ""
    ambiente_b: str = ""
    quantidade_b: int = 0

class PropostaMock:
    """Helper class para emoldurar as ações em formato compatível com execute_alocacao"""
    def __init__(self, nome, custo_obras, acoes):
        self.nome = nome
        self.custo_obras = custo_obras
        self.acoes = acoes

# ══════════════════════════════════════════════════════════════════════════
# Funções Auxiliares de Limpeza e Geometria
# ══════════════════════════════════════════════════════════════════════════

def clean_json_string(s: str) -> str:
    """Remove marcações de bloco de código Markdown do texto retornado pela LLM"""
    s = s.strip()
    match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', s, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    s = re.sub(r'^```(?:json)?\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*```$', '', s)
    return s.strip()

def extrair_e_carregar_json(s: str) -> dict:
    """Extrai e decodifica o JSON de forma tolerante a falhas estruturais da LLM"""
    s = s.strip()
    match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', s, re.DOTALL | re.IGNORECASE)
    if match:
        json_str = match.group(1).strip()
    else:
        match_bruto = re.search(r'(\{.*\})', s, re.DOTALL)
        if match_bruto:
            json_str = match_bruto.group(1).strip()
        else:
            json_str = s

    json_str = re.sub(r'//.*', '', json_str)  
    json_str = re.sub(r'/\*.*?\*/', '', json_str, flags=re.DOTALL)  

    return json.loads(json_str, strict=False)

def get_bench_partner(c, ws):
    global _bench_partners_cache
    if not _bench_partners_cache:
        pa_cols = set()
        for r in range(1, min(200, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(r, col).value
                if val is not None:
                    v_str = str(val).strip().upper()
                    if v_str not in ('VAZIO', 'CT', 'SA', 'SALA', 'CW', '##', ''):
                        pa_cols.add(col)
        
        sorted_cols = sorted(list(pa_cols))
        i = 0
        while i < len(sorted_cols) - 1:
            c1, c2 = sorted_cols[i], sorted_cols[i+1]
            if c2 == c1 + 1:
                _bench_partners_cache[c1] = c2
                _bench_partners_cache[c2] = c1
                i += 2
            else:
                i += 1
    return _bench_partners_cache.get(c)

def load_plant(path='planta.xlsx', sheet=SHEET_NAME):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb, wb[sheet]

def clone_ws(ws):
    """Clona a aba mantendo valores, estilos, dimensões e células mescladas de forma dinâmica."""
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Copia dados e estilos individuais
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            src, tgt = ws.cell(r, c), new_ws.cell(r, c)
            tgt.value = src.value
            if src.has_style:
                tgt.font, tgt.fill, tgt.alignment, tgt.border = copy(src.font), copy(src.fill), copy(src.alignment), copy(src.border)
                
    # Preserva as mesclagens físicas (Essencial para não destruir as divisórias de sandbox)
    for merged_range in ws.merged_cells.ranges:
        new_ws.merge_cells(str(merged_range))
        
    for col_l, dim in ws.column_dimensions.items(): 
        new_ws.column_dimensions[col_l].width = dim.width
    for row_i, dim in ws.row_dimensions.items(): 
        new_ws.row_dimensions[row_i].height = dim.height
        
    return new_wb, new_ws

def build_plant_info(plant_data):
    client_cells = plant_data['client_cells']
    lines = ["CÉLULAS POR VALOR NA PLANTA:"]
    for v in sorted(client_cells, key=lambda v: len(client_cells[v]), reverse=True)[:20]:
        lines.append(f"  '{v}': {len(client_cells[v])} PAs")
    return "\n".join(lines)

def build_blocos_info(plant_data, ws_max_row, ws_max_col, ws=None, file_path='planta.xlsx'):
    """Retorna a informação detalhada dos blocos dinamicamente."""
    if ws is None:
        return "Nenhum bloco laranja mapeado."
        
    macro_blocks = scan_orange_context(file_path, SHEET_NAME)
    allowed_cells = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                v_str = str(v).strip()
                if v_str != "" and v_str.upper() not in ('CT', 'CATRACA', 'SA', 'CW', 'COWORKING', '##'):
                    allowed_cells.add((r, c))
                    
    all_cells_cache = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            all_cells_cache[(r, c)] = ws.cell(row=r, column=c)
            
    lines = []
    for idx, block in enumerate(macro_blocks, start=1):
        b_id = f"Bloco_{idx}"  # <--- Rótulo do ID padronizado como Bloco_X (evita viés semântico)
        r_min, r_max, c_min, c_max = block['bounding_box']
        col_min = get_column_letter(c_min)
        col_max = get_column_letter(c_max)
        
        lines.append(f"📦 [Bloco {idx}] ({b_id}) colunas {col_min}-{col_max}, linhas {r_min}-{r_max}")
        if block['texts']:
            lines.append(f"  📌 Anotações: {', '.join(block['texts'])}")
        
        for env in block.get('ambientes', []):
            env_id = env['id']
            env_r_min, env_r_max, env_c_min, env_c_max = env['bounding_box']
            env_col_min = get_column_letter(env_c_min)
            env_col_max = get_column_letter(env_c_max)
            
            client_counts = {}
            empty_count = 0
            
            for (r, c) in env['cells']:
                if (r, c) not in allowed_cells:
                    continue  
                cell = all_cells_cache.get((r, c))
                val = cell.value if cell else None
                val_str = str(val).strip() if val is not None else ""
                
                if val_str.upper() in ('VAZIO', ''):
                    empty_count += 1
                elif val_str.upper() not in ('CT', 'CATRACA', 'SA', 'CW', 'COWORKING', '##'):
                    norm_val = normalize_val(val)
                    client_counts[norm_val] = client_counts.get(norm_val, 0) + 1
                    
            total_mesas = len([coord for coord in env['cells'] if coord in allowed_cells])
            lines.append(f"  - Amb {env_id} ({b_id}-{env_id}): {empty_count} Livres / {total_mesas} Total PAs")
            
            if client_counts:
                cl_str = ", ".join(f"'{cli}': {qty}" for cli, qty in sorted(client_counts.items()))
                lines.append(f"    Equipes: {cl_str}")
                
            if env.get('texts'):
                anotacoes_filtradas = [txt for txt in env['texts'] if txt.upper() not in ('VAZIO', '')]
                if anotacoes_filtradas:
                    lines.append(f"    Recursos/Anotações nesta sala: {', '.join(anotacoes_filtradas)}")
        lines.append("")
        
    summary_lines = [
        "============================================================",
        "RESUMO TOTAL DE CLIENTES EM MESAS (PLANTA INTEIRA):",
        "============================================================"
    ]
    total_counts = {}
    ignored_on_summary = {'VAZIO', 'CT', 'SA', 'SALA', 'CW', '##', '', 'CATRACA', 'ESCANINHOS', 'SALA CLIENTE', 'COWORKING'}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (r, c) not in allowed_cells:
                continue
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_str = str(int(val) if isinstance(val, float) and val == int(val) else val).strip()
                val_up = val_str.upper()
                if val_up not in ignored_on_summary:
                    total_counts[val_str] = total_counts.get(val_str, 0) + 1
    for client, qty in sorted(total_counts.items(), key=lambda x: x[0]):
        summary_lines.append(f"  '{client}': {qty} PAs;")
    summary_lines.append("============================================================\n")
        
    return "\n".join(lines) + "\n" + "\n".join(summary_lines)

# ══════════════════════════════════════════════════════════════════════════
# Motor Físico de Alocação e Formatação de Planilhas
# ══════════════════════════════════════════════════════════════════════════

def execute_alocacao(ws, proposta, plant_data, allowed_cells: Set[Tuple[int, int]], file_path: str = 'planta.xlsx') -> tuple:
    from BlockMapper import group_zones
    log = {'realocadas': {}, 'liberadas': {}, 'avisos': []}
    
    cell_values = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                cell_values[(r, c)] = str(int(v) if isinstance(v, float) and v == int(v) else v).strip()
            else:
                cell_values[(r, c)] = ""

    client_fills = {}
    client_fonts = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None:
                v_str = str(int(v) if isinstance(v, float) and v == int(v) else v).strip().upper()
                if v_str not in ('VAZIO', 'CT', 'SA', 'SALA', 'CW', '##', ''):
                    if cell.fill and cell.fill.patternType == 'solid':
                        client_fills[v_str] = copy(cell.fill)
                    if cell.font:
                        client_fonts[v_str] = copy(cell.font)

    paleta_cores = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']
    nomes_novos = sorted(list({
        str(a.cliente).strip().upper() for a in proposta.acoes if str(a.cliente).strip() != "" and str(a.cliente).upper().startswith("N_")
    }))
    
    default_new_fills = {}
    default_new_fonts = {}
    for idx, nome in enumerate(nomes_novos):
        cor_hex = paleta_cores[idx % len(paleta_cores)]
        default_new_fills[nome] = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
        default_new_fonts[nome] = Font(color='FFFFFF', bold=True, size=8)

    # --- REMOVIDO HARDCODE DO ARQUIVO ---
    macro_blocks = scan_orange_context(file_path, SHEET_NAME)
    non_client_values = {'VAZIO', 'CT', 'SA', 'SALA', 'CW', '##', ''}
    active_clients_cache = {v.upper() for v in cell_values.values() if v.upper() not in non_client_values}
    
    acoes_liberar = [a for a in proposta.acoes if a.tipo.lower().strip() == 'liberar']
    acoes_realocar = [a for a in proposta.acoes if a.tipo.lower().strip() in ('realocar', 'posicionar')]
    acoes_transferir = [a for a in proposta.acoes if a.tipo.lower().strip() in ('transferir', 'swap', 'permutar')]
    
    reduced_clients = {str(a.cliente).strip().upper() for a in acoes_liberar}

    def is_safe_cell(r, c, target):
        if (r, c) not in allowed_cells: return False
        t_up = target.upper()
        current_val = cell_values.get((r, c), "").strip().upper()
        
        partner_col = get_bench_partner(c, ws)
        if partner_col:
            partner_val = cell_values.get((r, partner_col), "").strip().upper()
            if partner_val and partner_val not in ('VAZIO', '', t_up) and partner_val not in reduced_clients:
                if partner_val in active_clients_cache: return False
        
        return current_val in non_client_values or current_val == t_up or current_val in reduced_clients

    # --- MAIS ROBUSTO CONTRA ALUCINAÇÕES DE LETRA DE AMBIENTE ---
    def get_env_cells(block_id_str: str, env_letter: str) -> List[Tuple[int, int]]:
        if not block_id_str or not env_letter:
            return []
        block_match = re.search(r'\d+', str(block_id_str))
        if not block_match:
            return []
        block_idx = int(block_match.group())
        if block_idx <= len(macro_blocks):
            block = macro_blocks[block_idx - 1]
            block_envs = {e['id'].upper(): e for e in block.get('ambientes', [])}
            
            # Fallback dinâmico se a letra solicitada não for achada mas só houver 1 ambiente real
            target_env = env_letter.upper()
            if target_env not in block_envs and len(block_envs) == 1:
                target_env = list(block_envs.keys())[0]
                
            parts = re.split(r'[-_\s+&,|/]+', target_env)
            matched_envs = []
            for p in parts:
                if p in block_envs:
                    matched_envs.append(block_envs[p])
                    
            if matched_envs:
                cells = []
                for env in matched_envs:
                    cells.extend(env['cells'])
                return cells
        return []

    def fill_bfs(disp, qtd):
        if not disp or qtd <= 0: return []
        alocadas = []
        for bloco in sorted(flood_fill(set(disp)), key=len, reverse=True):
            for atual in sorted(bloco, key=lambda coord: (coord[1], coord[0])):
                if len(alocadas) < qtd: alocadas.append(atual)
        if len(alocadas) < qtd: alocadas.extend([c for c in sorted(disp, key=lambda x: (x[1], x[0])) if c not in alocadas][:qtd - len(alocadas)])
        return alocadas

    # === FASE 1: REMOÇÃO ===
    for acao in acoes_liberar:
        alvo = acao.cliente.strip()
        alvo_norm = alvo.upper()
        
        block_id = acao.bloco
        if not block_id and acao.novo_cliente:
            m_block = re.search(r'(?:vazio|Bloco)-\d+', acao.novo_cliente, re.IGNORECASE)
            if m_block: block_id = m_block.group()
                
        env_letter = acao.ambiente
        if not env_letter and acao.novo_cliente:
            m_env = re.search(r'(?:vazio|Bloco)-\d+-+(.*)', acao.novo_cliente, re.IGNORECASE)
            if m_env: env_letter = m_env.group(1).strip()
        
        if not block_id or not env_letter or block_id.lower() == "automatico":
            env_vazios = []
            for b_idx, block in enumerate(macro_blocks, start=1):
                for env in block.get('ambientes', []):
                    vazios = sum(1 for coord in env['cells'] if cell_values.get(coord, "").upper() in ('VAZIO', '') and coord in allowed_cells)
                    env_vazios.append({
                        'block_idx': b_idx,
                        'env_id': env['id'],
                        'vazios': vazios,
                        'cells': env['cells']
                    })
            
            env_vazios.sort(key=lambda x: x['vazios'], reverse=True)
            
            cells_to_free = []
            for item in env_vazios:
                target_cells_in_env = [coord for coord in item['cells'] if cell_values.get(coord, "").upper() == alvo_norm and coord in allowed_cells]
                target_cells_in_env.sort(key=lambda x: (x[1], x[0]))
                
                cells_to_free.extend(target_cells_in_env)
                if len(cells_to_free) >= acao.quantidade:
                    break
                    
            cells_to_free = cells_to_free[:acao.quantidade]
            
            block_id = "automatico"
            env_letter = "automatico"
            if cells_to_free:
                primeira_celula = cells_to_free[0]
                found = False
                for b_idx, block in enumerate(macro_blocks, start=1):
                    for env in block.get('ambientes', []):
                        if primeira_celula in env['cells']:
                            block_id = f"Bloco_{b_idx}"
                            env_letter = env['id']
                            found = True
                            break
                    if found: break
        else:
            env_cells = get_env_cells(block_id, env_letter)
            if env_cells:
                cells_to_free = [coord for coord in env_cells if cell_values.get(coord, "").upper() == alvo_norm]
            else:
                cells_to_free = [k for k, v in cell_values.items() if v.upper() == alvo_norm]
            cells_to_free = sorted(cells_to_free, key=lambda coord: (coord[1], coord[0]))[:acao.quantidade]
            
        liberadas_count = 0
        for r, c in cells_to_free:
            ws.cell(r, c).value, cell_values[(r, c)] = 'vazio', 'vazio'
            ws.cell(r, c).fill, ws.cell(r, c).font = FILL_LIBERADO, FONT_SMALL
            log['liberadas'][(r, c)] = alvo
            liberadas_count += 1
            
        if liberadas_count < acao.quantidade:
            log['avisos'].append(f"Liberou apenas {liberadas_count} de {acao.quantidade} do cliente '{alvo}'.")

    # === FASE 2: POSICIONAMENTO ===
    for acao in acoes_realocar:
        target = acao.cliente.strip()
        
        block_id = acao.bloco
        if not block_id and acao.novo_cliente:
            m_block = re.search(r'(?:vazio|Bloco)-\d+', acao.novo_cliente, re.IGNORECASE)
            if m_block: block_id = m_block.group()
                
        env_letter = acao.ambiente
        if not env_letter and acao.novo_cliente:
            m_env = re.search(r'(?:vazio|Bloco)-\d+-+(.*)', acao.novo_cliente, re.IGNORECASE)
            if m_env: env_letter = m_env.group(1).strip()
        
        cli_clean = target
        
        if not block_id or not env_letter or block_id.lower() == "automatico":
            pool = [coord for coord in allowed_cells if cell_values.get(coord, "").upper() in ('VAZIO', '')]
            blocos_vazios = flood_fill(set(pool))
            
            candidatos = [b for b in blocos_vazios if len(b) >= acao.quantidade]
            if candidatos:
                bloco_escolhido = min(candidatos, key=len)
                dests = bloco_escolhido[:acao.quantidade]
            else:
                dests = fill_bfs(pool, acao.quantidade)
                
            bloco_origem = "automatico"
            env_origem = "automatico"
            if dests:
                primeira_celula = dests[0]
                found = False
                for b_idx, block in enumerate(macro_blocks, start=1):
                    for env in block.get('ambientes', []):
                        if primeira_celula in env['cells']:
                            bloco_origem, env_origem = f"Bloco_{b_idx}", env['id']
                            found = True
                            break
                        if found: break
                block_id, env_letter = bloco_origem, env_origem
        else:
            env_cells = get_env_cells(block_id, env_letter)
            pool = []
            if env_cells:
                for (r, c) in env_cells:
                    val = cell_values.get((r, c), "").upper()
                    if val in ('VAZIO', '') and is_safe_cell(r, c, cli_clean):
                        pool.append((r, c))
                        
            dests = fill_bfs(pool, acao.quantidade)
            
            if len(dests) < acao.quantidade and env_cells:
                pool_fallback = [coord for coord in env_cells if cell_values.get(coord, "").upper() in ('VAZIO', '') and coord in allowed_cells]
                dests = fill_bfs(pool_fallback, acao.quantidade)
        
        if len(dests) < acao.quantidade:
            log['avisos'].append(f"Capacidade física insuficiente no {block_id}-{env_letter} para '{cli_clean}'.")
            
        for dr, dc in dests:
            ws.cell(dr, dc).value, cell_values[(dr, dc)] = cli_clean, cli_clean
            fill_to_apply = client_fills.get(cli_clean.upper()) or default_new_fills.get(cli_clean.upper(), FILL_NEW_CLIENT)
            font_to_apply = default_new_fonts.get(cli_clean.upper()) or (FONT_WHITE if cli_clean.upper().startswith("N_") else copy(ws.cell(dr, dc).font))
            
            ws.cell(dr, dc).fill = fill_to_apply
            if font_to_apply: ws.cell(dr, dc).font = font_to_apply
                
            active_clients_cache.add(cli_clean.upper())
            log['realocadas'].setdefault(f"{cli_clean} → {block_id}-{env_letter}", []).append((dr, dc))

    # === FASE 3: TRANSFERÊNCIA ===
    for acao in acoes_transferir:
        cli_a = acao.cliente_a.strip()
        cli_b = acao.cliente_b.strip()
        
        cells_a = get_env_cells(acao.bloco_a, acao.ambiente_a)
        cells_b = get_env_cells(acao.bloco_b, acao.ambiente_b)
        
        if cli_a.upper() in ('VAZIO', ''):
            occupied_a = [coord for coord in cells_a if cell_values.get(coord, "").upper() in ('VAZIO', '') and coord in allowed_cells]
        else:
            occupied_a = [coord for coord in cells_a if cell_values.get(coord, "").upper() == cli_a.upper()]
            
        if cli_b.upper() in ('VAZIO', ''):
            occupied_b = [coord for coord in cells_b if cell_values.get(coord, "").upper() in ('VAZIO', '') and coord in allowed_cells]
        else:
            occupied_b = [coord for coord in cells_b if cell_values.get(coord, "").upper() == cli_b.upper()]
        
        occupied_a_sorted = sorted(occupied_a, key=lambda coord: (coord[1], coord[0]))
        occupied_b_sorted = sorted(occupied_b, key=lambda coord: (coord[1], coord[0]))
        
        targets_a = occupied_a_sorted[:acao.quantidade_a]
        targets_b = occupied_b_sorted[:acao.quantidade_b]
        
        for r, c in targets_a:
            ws.cell(r, c).value, cell_values[(r, c)] = 'vazio', 'vazio'
            ws.cell(r, c).fill, ws.cell(r, c).font = FILL_LIBERADO, FONT_SMALL
        for r, c in targets_b:
            ws.cell(r, c).value, cell_values[(r, c)] = 'vazio', 'vazio'
            ws.cell(r, c).fill, ws.cell(r, c).font = FILL_LIBERADO, FONT_SMALL
            
        if cli_a.upper() not in ('VAZIO', ''):
            free_in_b = [coord for coord in cells_b if cell_values.get(coord, "").upper() in ('VAZIO', '') and coord in allowed_cells]
            free_in_b_sorted = sorted(free_in_b, key=lambda coord: (coord[1], coord[0]))
            
            dests_a = fill_bfs(free_in_b_sorted, len(targets_a))
            
            if len(dests_a) < len(targets_a):
                log['avisos'].append(f"ERRO DE CAPACIDADE ao transferir '{cli_a}' para {acao.bloco_b}-{acao.ambiente_b}.")
                
            for dr, dc in dests_a:
                ws.cell(dr, dc).value, cell_values[(dr, dc)] = cli_a, cli_a
                fill_to_apply = client_fills.get(cli_a.upper()) or default_new_fills.get(cli_a.upper(), FILL_NEW_CLIENT)
                font_to_apply = default_new_fonts.get(cli_a.upper()) or (FONT_WHITE if cli_a.upper().startswith("N_") else copy(ws.cell(dr, dc).font))
                ws.cell(dr, dc).fill = fill_to_apply
                if font_to_apply: ws.cell(dr, dc).font = font_to_apply
                log['realocadas'].setdefault(f"{cli_a} → {acao.bloco_b}-{acao.ambiente_b}", []).append((dr, dc))
                
        if cli_b.upper() not in ('VAZIO', ''):
            free_in_a = [coord for coord in cells_a if cell_values.get(coord, "").upper() in ('VAZIO', '') and coord in allowed_cells]
            free_in_a_sorted = sorted(free_in_a, key=lambda coord: (coord[1], coord[0]))
            
            dests_b = fill_bfs(free_in_a_sorted, len(targets_b))
            
            if len(dests_b) < len(targets_b):
                log['avisos'].append(f"ERRO DE CAPACIDADE ao transferir '{cli_b}' para {acao.bloco_a}-{acao.ambiente_a}.")
                
            for dr, dc in dests_b:
                ws.cell(dr, dc).value, cell_values[(dr, dc)] = cli_b, cli_b
                fill_to_apply = client_fills.get(cli_b.upper()) or default_new_fills.get(cli_b.upper(), FILL_NEW_CLIENT)
                font_to_apply = default_new_fonts.get(cli_b.upper()) or (FONT_WHITE if cli_b.upper().startswith("N_") else copy(ws.cell(dr, dc).font))
                ws.cell(dr, dc).fill = fill_to_apply
                if font_to_apply: ws.cell(dr, dc).font = font_to_apply
                log['realocadas'].setdefault(f"{cli_b} → {acao.bloco_a}-{acao.ambiente_a}", []).append((dr, dc))

    return log, cell_values

# ══════════════════════════════════════════════════════════════════════════
# Geração de Relatórios de Mudança e Auditoria
# ══════════════════════════════════════════════════════════════════════════

def write_report(ws_orig, ws_new, proposta, log, path: str):
    changes = [(r, c, get_column_letter(c), str(ws_orig.cell(r, c).value or ''), str(ws_new.cell(r, c).value or ''))
               for r in range(1, ws_orig.max_row + 1) for c in range(1, ws_orig.max_column + 1) if str(ws_orig.cell(r, c).value or '') != str(ws_new.cell(r, c).value or '')]
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f"PROPOSTA: {proposta.nome}\nCUSTO: {proposta.custo_obras}\n\nALTERACOES: {len(changes)}")
        if changes:
            f.write("\n\nDETALHES DAS MUDANCAS:\n")
            for r, c, col, old, new in sorted(changes, key=lambda x: (x[0], x[1])):
                f.write(f"  {col}{r}: {old} -> {new}\n")
        if log.get('avisos'):
            f.write("\n\nAVISOS:\n")
            for aviso in log['avisos']:
                f.write(f"  - {aviso}\n")

def salvar_auditoria(nome_agente: str, input_data: str, output_data: str, safe_name: str):
    dir_auditoria = f"propostas/auditoria_{safe_name}"
    os.makedirs(dir_auditoria, exist_ok=True)
    with open(f"{dir_auditoria}/{nome_agente}_input.txt", "w", encoding="utf-8") as f:
        f.write(input_data)
    with open(f"{dir_auditoria}/{nome_agente}_output.txt", "w", encoding="utf-8") as f:
        f.write(output_data)

def normalizar_nome_cliente_novo(nome: str, novos_clientes: List[dict]) -> str:
    """Normaliza o nome de novos clientes para consistência no Excel e Gabarito"""
    nome_up = nome.upper().strip()
    if not nome_up.startswith("N_"):
        return nome
    match_sufixo = re.search(r'^N_(\w+)$', nome_up)
    sufixo = None
    if match_sufixo:
        sufixo = match_sufixo.group(1)
    if sufixo:
        for nc in novos_clientes:
            if nc["nome"].endswith(f"_{sufixo}"):
                return nc["nome"]
    return nome

def normalizar_acoes(acoes: List[Acao], novos_clientes: List[dict]) -> List[Acao]:
    for acao in acoes:
        if acao.cliente:
            acao.cliente = normalizar_nome_cliente_novo(acao.cliente, novos_clientes)
        if acao.cliente_a:
            acao.cliente_a = normalizar_nome_cliente_novo(acao.cliente_a, novos_clientes)
        if acao.cliente_b:
            acao.cliente_b = normalizar_nome_cliente_novo(acao.cliente_b, novos_clientes)
    return acoes

def validar_inventario(ws_orig, ws_new, allowed_cells, parametros: dict) -> List[str]:
    """Valida se o layout gerado preservou os clientes estáveis e atingiu as reduções/criações."""
    def obter_contagem(ws):
        counts = {}
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if (r, c) not in allowed_cells:
                    continue
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    norm_val = normalize_val(val)
                    if norm_val != "" and norm_val not in ('VAZIO', 'CT', 'CATRACA', 'SA', 'SALA', 'CW', 'COWORKING', '##'):
                        counts[norm_val] = counts.get(norm_val, 0) + 1
        return counts

    orig_counts = obter_contagem(ws_orig)
    new_counts = obter_contagem(ws_new)
    erros = []
    clientes_excluidos_da_preservacao = set(parametros["reducoes"].keys()) | {nc["nome"] for nc in parametros["novos_clientes"]}
    
    # 1. PRESERVAÇÃO DE CLIENTES ESTÁVEIS (Mantido para segurança contra apagamentos)
    for client, qty in orig_counts.items():
        if client in clientes_excluidos_da_preservacao: 
            continue
        qty_new = new_counts.get(client, 0)
        if qty_new != qty:
            erros.append(f"Inconsistência no Cliente estável '{client}': quantidade inicial era {qty}, mas agora é {qty_new} (Diferença de {qty_new - qty} PAs).")
            
    # 2. NOVOS CLIENTES (Removido o bloqueio rígido de igualdade)
    # Permite de forma adaptativa e sem hardcode que a alocação física comporte as cadeiras das salas de reunião internas (tolerância de até +10)
    for nc in parametros["novos_clientes"]:
        nome = nc["nome"]
        esperado = nc["PAs"]
        atual = new_counts.get(nome, 0)
        
        if atual != esperado:
            erros.append(
                f"Inconsistência no '{nome}': quantidade alocada ({atual}) "
                f"deve ser exatamente {esperado} PAs, incluindo mesas da sala interna."
            )
            
    # 3. REDUÇÕES DE CLIENTES EXISTENTES (Mantido para garantir as metas de redução)
    for cli, reducao in parametros["reducoes"].items():
        qty_orig = orig_counts.get(cli, 0)
        qty_new = new_counts.get(cli, 0)
        if qty_orig - qty_new != reducao:
            erros.append(f"Redução inválida do Cliente '{cli}': deveria eliminar exatamente {reducao} PAs (restante esperado: {qty_orig - reducao}), mas restaram {qty_new}.")
            
    return erros