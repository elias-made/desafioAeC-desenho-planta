"""
PlantGraph — Converte a planilha em estrutura de grafo espacial.

Estrutura:
  {
    "AQ13": {
      "cliente": "Novo A",
      "bench": 17,
      "row": 13,
      "col": 43
    },
    ...
  }

Responsável por:
  1. Mapear cada célula para cliente + bancada
  2. Identificar bancadas usando logica de parceiros (get_bench_partner)
  3. Gerar JSON serializável para envio à LLM
"""

from typing import Dict, Any, Optional, Tuple
from openpyxl.utils import get_column_letter, column_index_from_string


# ════════════════════════════════════════════════════════════════════════════
# Detecção de Bancadas
# ════════════════════════════════════════════════════════════════════════════

def identify_bench_partners(ws, max_scan_rows: int = 200) -> Dict[int, int]:
    """
    Identifica pares de colunas que formam bancadas.
    
    Uma bancada é formada por duas colunas consecutivas com dados operacionais.
    
    Retorna: {col: partner_col, ...}
    """
    pa_cols = set()
    
    # Escaneia as primeiras N linhas para identificar colunas com operadores
    for r in range(1, min(max_scan_rows, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is not None:
                v_str = str(val).strip().upper()
                # Identifica coluna com dados operacionais (não é vazio, sala, etc)
                if v_str not in ('VAZIO', 'CT', 'SA', 'SALA', 'CW', '##', '', 'CATRACA', 'SALA CLIENTE', 'COWORKING'): # '0' removido para ser tratado como operacional ativo
                    pa_cols.add(c)
    
    sorted_cols = sorted(list(pa_cols))
    bench_partners: Dict[int, int] = {}
    
    # Agrupa colunas consecutivas em bancadas (2 a 2)
    i = 0
    while i < len(sorted_cols) - 1:
        c1, c2 = sorted_cols[i], sorted_cols[i + 1]
        
        # Se são consecutivas, formam uma bancada
        if c2 == c1 + 1:
            bench_partners[c1] = c2
            bench_partners[c2] = c1
            i += 2
        else:
            i += 1
    
    return bench_partners


def assign_bench_id(col: int, bench_partners: Dict[int, int]) -> Optional[int]:
    """
    Atribui um ID de bancada a uma coluna.
    
    Colunas parceiras recebem o mesmo ID (menor das duas).
    Colunas isoladas recebem None.
    """
    if col in bench_partners:
        partner = bench_partners[col]
        return min(col, partner)
    
    return None


# ════════════════════════════════════════════════════════════════════════════
# Construção do Grafo
# ════════════════════════════════════════════════════════════════════════════

def build_plant_graph(ws) -> Dict[str, Dict[str, Any]]:
    """
    Constrói o grafo da planta a partir da worksheet.
    
    Retorna:
      {
        "AQ13": {
          "cliente": "Novo A",
          "bench": 17,
          "row": 13,
          "col": 43,
          "status": "ocupada" | "vazia" | "sala" | "infraestrutura"
        },
        ...
      }
    """
    bench_partners = identify_bench_partners(ws)
    graph: Dict[str, Dict[str, Any]] = {}
    
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            
            if v is None:
                continue
            
            # Normaliza valor
            v_str = str(int(v) if isinstance(v, float) and v == int(v) else v).strip()
            
            # Coordenada Excel
            cell_ref = f"{get_column_letter(c)}{r}"
            
            # Identifica tipo de célula
            if v_str in ('VAZIO', ''): # '0' removido para marcar status 'ocupada'
                status = 'vazia'
                cliente = None
            elif v_str in ('SALA', 'SALA CLIENTE', 'Sala'):
                status = 'sala'
                cliente = 'SALA'
            elif v_str in ('CT', 'CATRACA'):
                status = 'infraestrutura'
                cliente = 'CATRACA'
            elif v_str in ('SA', 'CW', 'COWORKING'):
                status = 'infraestrutura'
                cliente = v_str
            elif v_str == '##':
                status = 'infraestrutura'
                cliente = 'PAREDE'
            else:
                status = 'ocupada'
                cliente = v_str
            
            # Bancada
            bench_id = assign_bench_id(c, bench_partners)
            
            # Registro no grafo
            graph[cell_ref] = {
                'cliente': cliente,
                'bench': bench_id,
                'row': r,
                'col': c,
                'status': status,
                'col_letter': get_column_letter(c),
            }
    
    return graph


# ════════════════════════════════════════════════════════════════════════════
# Análise e Consultas
# ═══════════════════════════════════════════════════════════════════════════

def get_client_cells(graph: Dict[str, Dict[str, Any]], cliente: str) -> list:
    """Retorna todas as células de um cliente específico."""
    return [
        cell_ref
        for cell_ref, info in graph.items()
        if info.get('cliente') == cliente
    ]


def get_bench_cells(graph: Dict[str, Dict[str, Any]], bench_id: int) -> list:
    """Retorna todas as células de uma bancada específica."""
    return [
        cell_ref
        for cell_ref, info in graph.items()
        if info.get('bench') == bench_id
    ]


def get_clients_overview(graph: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Retorna resumo agregado por cliente.
    
    {
      "Novo A": {
        "count": 124,
        "benches": [17, 18],
        "status": "ocupada",
        "fragmented": False
      },
      ...
    }
    """
    overview: Dict[str, Dict[str, Any]] = {}
    
    for cell_ref, info in graph.items():
        cliente = info.get('cliente')
        
        if cliente is None:
            continue
        
        if cliente not in overview:
            overview[cliente] = {
                'count': 0,
                'benches': set(),
                'rows': set(),
                'cols': set(),
                'cells': []
            }
        
        overview[cliente]['count'] += 1
        
        if info.get('bench') is not None:
            overview[cliente]['benches'].add(info['bench'])
        
        overview[cliente]['rows'].add(info['row'])
        overview[cliente]['cols'].add(info['col'])
        overview[cliente]['cells'].append(cell_ref)
    
    # Converte sets para lists para JSON
    result = {}
    for cliente, data in overview.items():
        result[cliente] = {
            'count': data['count'],
            'benches': sorted(list(data['benches'])),
            'row_range': [min(data['rows']), max(data['rows'])],
            'col_range': [min(data['cols']), max(data['cols'])],
            'cells': sorted(data['cells']),
        }
    
    return result


def get_fragmentation_analysis(graph: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Analisa fragmentação de cada cliente.
    
    Retorna {cliente: {'fragments': N, 'largest_block': M, 'analysis': '...'}}
    """
    from collections import defaultdict, deque
    
    analysis: Dict[str, Dict[str, Any]] = {}
    
    # Agrupa células por cliente
    client_cells: Dict[str, set] = defaultdict(set)
    for cell_ref, info in graph.items():
        cliente = info.get('cliente')
        if cliente and info.get('status') == 'ocupada':
            client_cells[cliente].add((info['row'], info['col']))
    
    # Para cada cliente, encontra blocos contíguos
    for cliente, cells in client_cells.items():
        if not cells:
            continue
        
        visited = set()
        blocks = []
        
        for seed in sorted(cells):
            if seed in visited:
                continue
            
            # BFS para encontrar bloco contíguo
            block = []
            q = deque([seed])
            
            while q:
                r, c = q.popleft()
                if (r, c) in visited or (r, c) not in cells:
                    continue
                
                visited.add((r, c))
                block.append((r, c))
                
                for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                    neighbor = (r + dr, c + dc)
                    if neighbor not in visited and neighbor in cells:
                        q.append(neighbor)
            
            blocks.append(len(block))
        
        blocks_sorted = sorted(blocks, reverse=True)
        
        analysis[cliente] = {
            'fragments': len(blocks),
            'largest_block': blocks_sorted[0] if blocks_sorted else 0,
            'block_sizes': blocks_sorted,
            'fragmented': len(blocks) > 1,
        }
    
    return analysis


def serialize_plant_graph(graph: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Serializa o grafo para JSON (remove tipos não-serializáveis).
    
    Formato simplificado:
      {
        "AQ13": {
          "cliente": "Novo A",
          "bench": 17,
          "status": "ocupada"
        },
        ...
      }
    """
    serialized = {}
    
    for cell_ref, info in graph.items():
        serialized[cell_ref] = {
            'cliente': info.get('cliente'),
            'bench': info.get('bench'),
            'status': info.get('status'),
        }
    
    return serialized


# ════════════════════════════════════════════════════════════════════════════
# Interface Pública
# ════════════════════════════════════════════════════════════════════════════

def create_plant_graph(ws) -> Dict[str, Dict[str, Any]]:
    """
    Cria o grafo completo da planta a partir de uma worksheet.
    
    Este é o ponto de entrada principal.
    """
    return build_plant_graph(ws)


def get_plant_graph_summary(graph: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """
    Retorna resumo do grafo para fins informativos.
    """
    overview = get_clients_overview(graph)
    fragmentation = get_fragmentation_analysis(graph)
    
    return {
        'total_cells': len(graph),
        'clients': overview,
        'fragmentation': fragmentation,
    }