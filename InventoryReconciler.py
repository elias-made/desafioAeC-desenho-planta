"""Reconciliacao deterministica do inventario apos a construcao geometrica."""

import re
from copy import copy

from openpyxl.styles import Font, PatternFill

import ScannerPremissas
from LayoutEngine import FILL_LIBERADO, FILL_NEW_CLIENT, FONT_SMALL, FONT_WHITE
from ScannerPremissas import normalize_val, scan_orange_context

IGNORADOS = {'VAZIO', 'CT', 'CATRACA', 'SA', 'SALA', 'CW', 'COWORKING', '##'}
PALETA_NOVOS = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']


def _contar(sheet, cells):
    counts = {}
    for r, c in cells:
        valor = sheet.cell(row=r, column=c).value
        if valor is None:
            continue
        nome = normalize_val(valor)
        if nome and nome not in IGNORADOS:
            counts[nome] = counts.get(nome, 0) + 1
    return counts


def _estilos_clientes(ws, new_ws):
    fills, fonts = {}, {}
    for sheet in (ws, new_ws):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                nome = normalize_val(cell.value)
                if not nome or nome in IGNORADOS:
                    continue
                if cell.fill and cell.fill.patternType == 'solid':
                    fills[nome] = copy(cell.fill)
                if cell.font:
                    fonts[nome] = copy(cell.font)
    return fills, fonts


def _metas(ws, allowed_cells, parametros, ambientes, falhos):
    targets = _contar(ws, allowed_cells)
    for cliente, reducao in parametros['reducoes'].items():
        if cliente in targets:
            targets[cliente] -= reducao
    for novo in parametros['novos_clientes']:
        nome = novo['nome']
        if nome in falhos:
            continue
        sala = sum(int(a.get('sala_lugares', 0) or 0) for a in ambientes if a.get('cliente_destinado') == nome)
        targets[nome] = novo['PAs'] + sala
    return targets


def _ambientes_dedicados(dest_file, sheet_name, new_ws, ambientes):
    ScannerPremissas._orange_context_cache = {}
    blocos = scan_orange_context(dest_file, sheet_name)
    result = {}
    for ambiente in ambientes:
        bloco_match = re.search(r'\d+', str(ambiente.get('bloco')))
        cliente = ambiente.get('cliente_destinado')
        if not bloco_match:
            continue
        indice = int(bloco_match.group()) - 1
        if not 0 <= indice < len(blocos):
            continue
        for env in blocos[indice].get('ambientes', []):
            if any(new_ws.cell(r, c).value == cliente for r, c in env['cells']):
                result.setdefault(cliente, set()).update(env['cells'])
    return result


def reconciliar_inventario(
    ws, new_ws, allowed_cells, salas_internas_cells, parametros,
    ambientes, ambientes_falhos, dest_file, sheet_name,
):
    """Ajusta excessos/deficits e devolve celulas e parametros para validacao."""
    inventory_cells = set(allowed_cells) | set(salas_internas_cells)
    client_fills, client_fonts = _estilos_clientes(ws, new_ws)
    defaults_fill, defaults_font = {}, {}
    for idx, novo in enumerate(parametros['novos_clientes']):
        nome = novo['nome']
        if nome in ambientes_falhos:
            continue
        cor = PALETA_NOVOS[idx % len(PALETA_NOVOS)]
        defaults_fill[nome] = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
        defaults_font[nome] = Font(color='FFFFFF', bold=True, size=8)

    targets = _metas(ws, allowed_cells, parametros, ambientes, ambientes_falhos)
    dedicados = _ambientes_dedicados(dest_file, sheet_name, new_ws, ambientes)
    atuais = _contar(new_ws, inventory_cells)

    for cliente, target in targets.items():
        excesso = atuais.get(cliente, 0) - target
        if excesso <= 0:
            continue
        candidatos = []
        for r, c in allowed_cells:
            if normalize_val(new_ws.cell(r, c).value) == cliente and (r, c) not in dedicados.get(cliente, set()):
                candidatos.append((r, c))
        for r, c in sorted(candidatos, key=lambda x: (x[1], x[0]))[:excesso]:
            cell = new_ws.cell(r, c)
            cell.value = 'vazio'
            cell.fill = FILL_LIBERADO
            cell.font = FONT_SMALL

    atuais = _contar(new_ws, inventory_cells)
    for cliente, target in targets.items():
        deficit = target - atuais.get(cliente, 0)
        if deficit <= 0:
            continue
        vazios = []
        for r, c in allowed_cells:
            if (r, c) in salas_internas_cells:
                continue
            valor = new_ws.cell(r, c).value
            if valor is None or normalize_val(valor) in ('VAZIO', ''):
                vazios.append((r, c))
        vazios.sort(key=lambda x: (x[1], x[0]))
        fill = client_fills.get(cliente) or defaults_fill.get(cliente, FILL_NEW_CLIENT)
        font = client_fonts.get(cliente) or defaults_font.get(cliente) or (FONT_WHITE if cliente.startswith('N_') else FONT_SMALL)
        for r, c in vazios[:deficit]:
            cell = new_ws.cell(r, c)
            cell.value = cliente
            cell.fill = fill
            cell.font = font

    novos_validacao = []
    for novo in parametros['novos_clientes']:
        nome = novo['nome']
        if nome in ambientes_falhos:
            continue
        sala = sum(int(a.get('sala_lugares', 0) or 0) for a in ambientes if a.get('cliente_destinado') == nome)
        novos_validacao.append({'nome': nome, 'PAs': novo['PAs'] + sala})
    parametros_validacao = {'reducoes': parametros['reducoes'], 'novos_clientes': novos_validacao}
    return inventory_cells, parametros_validacao