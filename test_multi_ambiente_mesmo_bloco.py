"""
Teste: Criação de múltiplos ambientes no mesmo bloco (Bloco_1-A).

Simula o cenário do posicionador onde N_1 (124 PAs + sala 4 lugares) e N_3 (20 PAs)
são criados consecutivamente no Bloco_1, ambiente A.

Valida que:
1. Não há sobreposição de células entre N_1 e N_3
2. A quantidade de mesas alocadas está correta para cada ambiente
3. As divisórias são desenhadas corretamente para ambos
4. A restauração do snapshot não apaga o trabalho de iterações anteriores
"""

import os
import shutil
import openpyxl
from copy import copy
from typing import Set, Tuple

from ScannerPremissas import scan_orange_context, normalize_val, _orange_context_cache
import ScannerPremissas
from AmbienteBuilder import (
    separar_ambiente_e_desenhar_divisorias,
    _selecionar_mesas_contiguas,
    _gerar_layout_sala_estruturado,
    gerar_alternativas_mesas,
    _celulas_contorno_do_ambiente,
    get_env_cells as amb_get_env_cells
)

# ══════════════════════════════════════════════════════════════════════════
# Configuração
# ══════════════════════════════════════════════════════════════════════════

SHEET_NAME = 'JPIII'
INPUT_FILE = 'planta.xlsx'
OUTPUT_FILE = 'planta_teste_multi_ambiente.xlsx'

# Cenário do posicionador: dois ambientes no Bloco_1-A
CRIAR_AMBIENTES = [
    {
        "bloco": "Bloco_1",
        "ambiente": "A",
        "quantidade_mesas": 124,
        "cliente_destinado": "N_1",
        "sala_lugares": 4
    },
    {
        "bloco": "Bloco_1",
        "ambiente": "A",
        "quantidade_mesas": 20,
        "cliente_destinado": "N_3",
        "sala_lugares": 0
    },
]


def get_env_cells_local(block_id_str, env_letter, macro_blocks):
    """Replica a lógica corrigida do main.py: pega TODAS as células do bloco."""
    import re
    if not block_id_str or not env_letter:
        return []
    block_match = re.search(r'\d+', str(block_id_str))
    if not block_match:
        return []
    block_idx = int(block_match.group())
    if block_idx <= len(macro_blocks):
        block = macro_blocks[block_idx - 1]
        block_envs = {e['id'].upper(): e for e in block.get('ambientes', [])}
        if not block_envs:
            return []
        # Retorna TODAS as células de todos os ambientes do bloco
        cells = []
        for env in block_envs.values():
            cells.extend(env['cells'])
        return cells
    return []


def main():
    print("=" * 80)
    print("TESTE: Criação de Múltiplos Ambientes no Mesmo Bloco (Bloco_1-A)")
    print("=" * 80)

    if not os.path.exists(INPUT_FILE):
        print(f"ERRO: Arquivo '{INPUT_FILE}' não encontrado. Execute a partir da raiz do projeto.")
        return

    # Copia a planta original para não alterar o arquivo fonte
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_NAME]

    # Snapshot original para restauração
    original_snapshot = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            original_snapshot[(r, c)] = {
                'value': cell.value,
                'fill': copy(cell.fill) if cell.has_style else None,
                'font': copy(cell.font) if cell.has_style else None
            }

    # ══════════════════════════════════════════════════════════════════════════
    # Execução com acumulador (lógica corrigida do main.py)
    # ══════════════════════════════════════════════════════════════════════════

    celulas_ja_consumidas: Set[Tuple[int, int]] = set()
    celulas_consumidas_por_bloco: dict = {}
    resultados = {}

    paleta_cores = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']

    for i, amb in enumerate(CRIAR_AMBIENTES):
        bloco_id = amb["bloco"]
        ambiente_letra = amb["ambiente"]
        qtd_mesas = amb["quantidade_mesas"]
        cliente_dest = amb["cliente_destinado"]
        sala_lugares = amb.get("sala_lugares", 0)

        print(f"\n{'─' * 60}")
        print(f"Iteração {i+1}: Criando {cliente_dest} ({qtd_mesas} PAs"
              + (f" + sala {sala_lugares} lugares" if sala_lugares else "") + f") no {bloco_id}-{ambiente_letra}")
        print(f"{'─' * 60}")

        # Limpa cache e re-scana (simula o comportamento do main.py)
        ScannerPremissas._orange_context_cache = {}
        macro_blocks_atual = scan_orange_context(OUTPUT_FILE, SHEET_NAME)

        env_cells_raw = set(get_env_cells_local(bloco_id, ambiente_letra, macro_blocks_atual))

        # CORREÇÃO: Subtrai células já consumidas por iterações anteriores
        consumidas_neste_bloco = celulas_consumidas_por_bloco.get(bloco_id, set())
        env_cells = env_cells_raw - consumidas_neste_bloco

        print(f"   Células disponíveis no bloco (raw): {len(env_cells_raw)}")
        print(f"   Células já consumidas neste bloco:  {len(consumidas_neste_bloco)}")
        print(f"   Células disponíveis (efetivas):     {len(env_cells)}")

        if not env_cells:
            print(f"   ERRO: Nenhuma célula disponível para {cliente_dest}!")
            resultados[cliente_dest] = {"sucesso": False, "motivo": "sem células"}
            continue

        allocated_sala = set()
        room_cells_override = None

        if sala_lugares and sala_lugares > 0:
            print(f"   Planejando sala fechada de {sala_lugares} mesas...")
            allocated_sala, room_cells_override = _gerar_layout_sala_estruturado(ws, env_cells, sala_lugares)
            print(f"   Sala planejada: {len(allocated_sala)} mesas em {len(room_cells_override)} células de room")

        available_env_cells = env_cells - (room_cells_override if room_cells_override else set())
        
        # Diagnóstico: quantas células são mesas de fato?
        from AmbienteBuilder import _eh_celula_de_mesa_local
        mesas_disponiveis = set()
        for r, c in available_env_cells:
            cell = ws.cell(row=r, column=c)
            if _eh_celula_de_mesa_local(cell):
                mesas_disponiveis.add((r, c))
        print(f"   Células disponíveis para seleção: {len(available_env_cells)}")
        print(f"   Dessas, são mesas reconhecidas:   {len(mesas_disponiveis)}")
        if mesas_disponiveis:
            sample = list(mesas_disponiveis)[:5]
            for r, c in sample:
                print(f"      ({r},{c}) valor='{ws.cell(row=r, column=c).value}'")
        else:
            # Mostrar o que tem nas células disponíveis
            sample = list(available_env_cells)[:10]
            print(f"   Amostra de células disponíveis (não-mesa):")
            for r, c in sample:
                cell = ws.cell(row=r, column=c)
                print(f"      ({r},{c}) valor='{cell.value}' fill={cell.fill.patternType if cell.fill else None}")
        
        alternativas = gerar_alternativas_mesas(available_env_cells, ws, qtd_mesas)
        allocated_ambiente = alternativas[0] if alternativas else set()

        if not allocated_ambiente:
            print(f"   ERRO: Nenhuma mesa contígua encontrada para {cliente_dest}!")
            resultados[cliente_dest] = {"sucesso": False, "motivo": "sem mesas contíguas"}
            continue

        print(f"   Mesas alocadas para salão aberto: {len(allocated_ambiente)}")

        allocated_total = allocated_ambiente | (room_cells_override if room_cells_override else set())

        # Desenha divisórias do salão aberto
        # A nova lógica BFS para automaticamente ao encontrar mesas de outros clientes
        ambiente_fisico = {"valido": False, "motivo": "sem alternativa válida"}
        for alternativa in alternativas:
            candidato_total = alternativa | (
                room_cells_override if room_cells_override else set()
            )
            tentativa = separar_ambiente_e_desenhar_divisorias(
                ws=ws,
                env_cells=env_cells,
                allocated_cells=candidato_total,
                reconstruir_sala=False,
                room_cells_override=room_cells_override,
            )
            if tentativa.get("valido"):
                allocated_ambiente = alternativa
                allocated_total = candidato_total
                ambiente_fisico = tentativa
                break
            ambiente_fisico = tentativa
        if not ambiente_fisico.get("valido"):
            resultados[cliente_dest] = {
                "sucesso": False,
                "motivo": ambiente_fisico.get("motivo", "ambiente físico inválido")
            }
            continue

        # Desenha sala fechada
        if room_cells_override:
            separar_ambiente_e_desenhar_divisorias(
                ws=ws,
                env_cells=env_cells,
                allocated_cells=allocated_sala,
                reconstruir_sala=True,
                room_cells_override=room_cells_override
            )

        # Aplica cor/valor nas mesas do salão
        cor_hex = paleta_cores[i % len(paleta_cores)]
        from openpyxl.styles import PatternFill, Font
        fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
        font = Font(color='FFFFFF', bold=True, size=8)

        for r, c in allocated_ambiente:
            cell = ws.cell(row=r, column=c)
            if cell.value != "CT":
                cell.value = cliente_dest
                cell.fill = fill
                cell.font = font

        # Preenche mesas da sala de reunião com estilo cinza "vazio"

        # Preenche mesas da sala de reunião com estilo cinza "vazio"
        if room_cells_override:
            fill_sala = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
            font_sala = Font(name="Calibri", size=9)
            for r, c in allocated_sala:
                cell = ws.cell(row=r, column=c)
                if cell.value != "CT":
                    cell.value = "vazio"
                    cell.fill = fill_sala
                    cell.font = font_sala

        # Registra no acumulador apenas as mesas efetivamente alocadas
        all_allocated_cells = (
            allocated_ambiente
            | (allocated_sala if allocated_sala else set())
            | ambiente_fisico.get("relocated_cells", set())
        )
        celulas_ja_consumidas.update(all_allocated_cells)
        celulas_consumidas_por_bloco.setdefault(bloco_id, set()).update(
            ambiente_fisico["room_cells"]
            | ambiente_fisico.get("relocated_cells", set())
        )

        # Restaura células não consumidas (protegendo iterações anteriores)
        for r, c in env_cells:
            if (r, c) not in all_allocated_cells and (r, c) not in celulas_ja_consumidas:
                cell = ws.cell(row=r, column=c)
                if cell.value != "CT":
                    snap = original_snapshot.get((r, c))
                    if snap:
                        cell.value = snap['value']
                        if snap['fill']:
                            cell.fill = snap['fill']
                        if snap['font']:
                            cell.font = snap['font']

        # Salva após cada iteração
        wb.save(OUTPUT_FILE)

        resultados[cliente_dest] = {
            "sucesso": True,
            "mesas_alocadas": len(allocated_ambiente),
            "sala_mesas": len(allocated_sala),
            "room_cells": len(room_cells_override) if room_cells_override else 0,
            "celulas_total": len(all_allocated_cells),
            "coords_ambiente": allocated_ambiente,
            "coords_sala": allocated_sala,
            "room_cells_fisico": ambiente_fisico["room_cells"],
            "corridor_cells": ambiente_fisico["corridor_cells"],
            "ct_cell": ambiente_fisico["ct_cell"],
            "mesas_sem_saida": ambiente_fisico["mesas_sem_saida"],
        }

        print(f"   ✓ {cliente_dest} criado com sucesso!")

    # ══════════════════════════════════════════════════════════════════════════
    # Validações
    # ══════════════════════════════════════════════════════════════════════════

    print("\n" + "=" * 80)
    print("VALIDAÇÕES")
    print("=" * 80)

    falhas = 0

    # 1. Verificar sobreposição
    if "N_1" in resultados and "N_3" in resultados:
        if resultados["N_1"]["sucesso"] and resultados["N_3"]["sucesso"]:
            cells_n1 = resultados["N_1"]["coords_ambiente"] | resultados["N_1"]["coords_sala"]
            cells_n3 = resultados["N_3"]["coords_ambiente"] | resultados["N_3"]["coords_sala"]
            sobreposicao = cells_n1 & cells_n3

            if sobreposicao:
                print(f"\n❌ FALHA: Sobreposição detectada entre N_1 e N_3!")
                print(f"   Células sobrepostas: {len(sobreposicao)}")
                print(f"   Exemplos: {list(sobreposicao)[:5]}")
                falhas += 1
            else:
                print(f"\n✓ OK: Sem sobreposição entre N_1 e N_3")

    # 2. Verificar quantidades
    for cliente, res in resultados.items():
        if not res["sucesso"]:
            print(f"\n❌ FALHA: {cliente} não foi criado - {res['motivo']}")
            falhas += 1
            continue

        esperado = next(a["quantidade_mesas"] for a in CRIAR_AMBIENTES if a["cliente_destinado"] == cliente)
        real = res["mesas_alocadas"]
        if real < esperado:
            print(f"\n⚠️  AVISO: {cliente} alocou {real} mesas (esperado: {esperado}) - deficit de {esperado - real}")
            falhas += 1
        else:
            print(f"\n✓ OK: {cliente} alocou {real} mesas (esperado: {esperado})")

        if res["mesas_sem_saida"] or not res["ct_cell"]:
            print(f"\n❌ FALHA: {cliente} não possui circulação válida até uma CT")
            falhas += 1
        else:
            print(f"\n✓ OK: todas as mesas de {cliente} alcançam a CT {res['ct_cell']}")

    # 3. Verificar que N_1 não foi apagado pela iteração de N_3
    wb_check = openpyxl.load_workbook(OUTPUT_FILE)
    ws_check = wb_check[SHEET_NAME]

    count_n1 = 0
    count_n3 = 0
    for r in range(1, ws_check.max_row + 1):
        for c in range(1, ws_check.max_column + 1):
            val = ws_check.cell(row=r, column=c).value
            if val is not None:
                norm = normalize_val(val)
                if norm == "N_1":
                    count_n1 += 1
                elif norm == "N_3":
                    count_n3 += 1

    print(f"\n{'─' * 60}")
    print(f"Contagem final na planilha salva:")
    print(f"   N_1: {count_n1} células com valor 'N_1'")
    print(f"   N_3: {count_n3} células com valor 'N_3'")

    if count_n1 == 0 and resultados.get("N_1", {}).get("sucesso"):
        print(f"\n❌ FALHA CRÍTICA: N_1 foi apagado pela iteração de N_3! (restauração destrutiva)")
        falhas += 1
    elif count_n1 > 0:
        print(f"\n✓ OK: N_1 preservado na planilha final ({count_n1} células)")

    if count_n3 == 0 and resultados.get("N_3", {}).get("sucesso"):
        print(f"\n❌ FALHA: N_3 não aparece na planilha final")
        falhas += 1
    elif count_n3 > 0:
        print(f"\n✓ OK: N_3 presente na planilha final ({count_n3} células)")

    # Resumo final
    print(f"\n{'=' * 80}")
    if falhas == 0:
        print("RESULTADO: ✓ TODOS OS TESTES PASSARAM")
    else:
        print(f"RESULTADO: ❌ {falhas} FALHA(S) DETECTADA(S)")
    print(f"{'=' * 80}")
    print(f"\nPlanilha de saída: {OUTPUT_FILE}")


if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception as e:
        traceback.print_exc()
