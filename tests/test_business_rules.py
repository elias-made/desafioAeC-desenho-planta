import copy
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from CapacityValidator import avaliar_capacidade
from PremiseNormalizer import normalizar_nomes_novos_por_ordem, normalizar_quantidades_ambientes
from InventoryReconciler import reconciliar_inventario
from AmbienteBuilder import _corredores_alcancaveis_da_saida, _gerar_layout_sala_estruturado
from ScannerPremissas import invalidate_orange_cache, scan_orange_context
from Agents import OrganizadorDeps


class PremiseNormalizerTests(unittest.TestCase):
    def test_ordem_do_txt_prevalece_sobre_nomes_da_llm(self):
        premissas = """Criar espaÃƒÂ§o com 124 PAs + sala com 4 lugares
Criar espaÃƒÂ§o com 165 PAs + sala com 1 lugar
Criar espaÃƒÂ§o no primeiro bloco com 20 novos PAs
Criar espaÃƒÂ§o com 70 PAs no mesmo bloco do cliente com 124 PAs"""
        data = {
            "gabarito": {"novos_clientes": [
                {"nome": "N_1"}, {"nome": "N_2"}, {"nome": "N_3"}, {"nome": "N_4"}
            ]},
            "criar_ambientes": [
                {"quantidade_mesas": 20, "sala_lugares": 0, "cliente_destinado": "N_4"},
                {"quantidade_mesas": 124, "sala_lugares": 4, "cliente_destinado": "N_1"},
                {"quantidade_mesas": 70, "sala_lugares": 0, "cliente_destinado": "N_3"},
                {"quantidade_mesas": 165, "sala_lugares": 1, "cliente_destinado": "N_2"},
            ],
            "acoes_primarias": [
                {"cliente": "N_4"}, {"cliente": "N_3"}
            ],
        }
        result = normalizar_nomes_novos_por_ordem(copy.deepcopy(data), premissas)
        por_quantidade = {a["quantidade_mesas"]: a["cliente_destinado"] for a in result["criar_ambientes"]}
        self.assertEqual(por_quantidade, {20: "N_3", 124: "N_1", 70: "N_4", 165: "N_2"})
        self.assertEqual([a["cliente"] for a in result["acoes_primarias"]], ["N_3", "N_4"])

    def test_quantidade_explicita_nao_soma_sala_duas_vezes(self):
        ambientes = [{"cliente_destinado": "N_1", "quantidade_mesas": 128, "sala_lugares": 4}]
        clientes = [{"nome": "N_1", "PAs": 124}]
        normalizar_quantidades_ambientes(ambientes, clientes, {"N_1": 124})
        self.assertEqual(ambientes[0]["quantidade_mesas"], 124)
        self.assertEqual(ambientes[0]["quantidade_mesas"] + ambientes[0]["sala_lugares"], 128)

    def test_formato_antigo_e_convertido(self):
        ambientes = [{"cliente_destinado": "N_2", "quantidade_mesas": 166, "sala_lugares": 1}]
        clientes = [{"nome": "N_2", "PAs": 166}]
        normalizar_quantidades_ambientes(ambientes, clientes, {})
        self.assertEqual(clientes[0]["PAs"], 165)
        self.assertEqual(ambientes[0]["quantidade_mesas"], 165)


class AgentContractTests(unittest.TestCase):
    def test_organizador_aceita_contexto_dos_ambientes_criados(self):
        deps = OrganizadorDeps(
            plant_info="planta", blocos_info="blocos", premissas="premissas",
            rascunho_layout="rascunho", ambientes_criados="ambientes",
        )
        self.assertEqual(deps.ambientes_criados, "ambientes")

class ScannerCacheTests(unittest.TestCase):
    def test_scanner_em_memoria_nao_reabre_arquivo_e_respeita_invalidacao(self):
        ws = Workbook().active
        ws.cell(3, 3).value = ""
        invalidate_orange_cache(ws=ws)
        with patch("ScannerPremissas.openpyxl.load_workbook", side_effect=AssertionError("nao deve abrir arquivo")):
            primeiro = scan_orange_context("inexistente.xlsx", "Sheet", ws=ws)
            segundo = scan_orange_context("inexistente.xlsx", "Sheet", ws=ws)
        self.assertIs(primeiro, segundo)
        invalidate_orange_cache(ws=ws)
        terceiro = scan_orange_context("inexistente.xlsx", "Sheet", ws=ws)
        self.assertIsNot(primeiro, terceiro)

class GeometryUtilityTests(unittest.TestCase):
    def _area_livre(self, tamanho=12):
        ws = Workbook().active
        ws.cell(tamanho, tamanho)
        env = {(r, c) for r in range(1, tamanho + 1) for c in range(1, tamanho + 1)}
        return ws, env

    def test_sala_de_um_lugar_usa_apenas_mesa_e_corredor(self):
        ws, env = self._area_livre()
        mesas, sala = _gerar_layout_sala_estruturado(ws, env, 1)
        self.assertEqual(len(mesas), 1)
        self.assertEqual(len(sala), 2)
        self.assertTrue(mesas <= sala)

    def test_sala_de_quatro_lugares_usa_modulo_mesa_corredor_mesa(self):
        ws, env = self._area_livre()
        mesas, sala = _gerar_layout_sala_estruturado(ws, env, 4)
        corredores = sala - mesas
        self.assertEqual(len(mesas), 4)
        self.assertEqual(len(sala), 6)
        self.assertEqual(len(corredores), 2)
        for r, c in mesas:
            self.assertTrue(
                any((r + dr, c + dc) in corredores for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)))
            )

    def test_corredor_interno_da_sala_encontra_corredor_externo(self):
        ws, env = self._area_livre()
        mesas, sala = _gerar_layout_sala_estruturado(ws, env, 4)
        corredores_internos = sala - mesas
        self.assertTrue(any(
            (r + dr, c + dc) in env - sala
            for r, c in corredores_internos
            for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1))
        ))

    def test_corredores_internos_alcancam_a_saida(self):
        ws = Workbook().active
        ws.cell(6, 6).value = ""  # dimensao externa ao ambiente do teste
        room = {(r, c) for r in range(2, 5) for c in range(2, 5)}
        desks = {(3, 3)}
        corredores, saidas, alcancaveis = _corredores_alcancaveis_da_saida(ws, room, desks)
        self.assertEqual(corredores, room - desks)
        self.assertTrue(saidas)
        self.assertEqual(alcancaveis, corredores)

class InventoryReconcilerTests(unittest.TestCase):
    @patch("InventoryReconciler._ambientes_dedicados", return_value={"N_1": {(1, 1), (2, 1)}})
    def test_mesa_da_sala_fora_das_celulas_originais_entra_na_contagem(self, _):
        original = Workbook().active
        atual = Workbook().active
        original.cell(1, 1).value = "vazio"
        original.cell(1, 2).value = "vazio"
        atual.cell(1, 1).value = "N_1"
        atual.cell(1, 2).value = "vazio"
        atual.cell(2, 1).value = "N_1"  # lugar da sala fora de allowed_cells
        inventory, validacao = reconciliar_inventario(
            original, atual, {(1, 1), (1, 2)}, {(2, 1)},
            {"reducoes": {}, "novos_clientes": [{"nome": "N_1", "PAs": 1}]},
            [{"cliente_destinado": "N_1", "sala_lugares": 1}], set(), "ignorado.xlsx", "Planilha",
        )
        self.assertIn((2, 1), inventory)
        self.assertEqual(atual.cell(1, 2).value, "vazio")
        self.assertEqual(validacao["novos_clientes"], [{"nome": "N_1", "PAs": 2}])

class CapacityValidatorTests(unittest.TestCase):
    def setUp(self):
        self.ws = Workbook().active
        self.allowed = {(1, i) for i in range(1, 7)}
        for i in range(1, 4):
            self.ws.cell(1, i).value = "vazio"
        for i in range(4, 7):
            self.ws.cell(1, i).value = "1"

    def test_sala_e_contada_uma_vez(self):
        result = avaliar_capacidade(
            self.ws, self.allowed, {"1": 3}, [{"nome": "N_1", "PAs": 4}],
            [{"cliente_destinado": "N_1", "quantidade_mesas": 4, "sala_lugares": 2}],
            {"1": {(1, 4), (1, 5), (1, 6)}},
        )
        self.assertTrue(result.viavel)
        self.assertEqual(result.capacidade, 6)
        self.assertEqual(result.demanda, 6)

    def test_brancos_nao_sao_capacidade(self):
        self.ws.cell(2, 1).value = None
        result = avaliar_capacidade(self.ws, self.allowed | {(2, 1)}, {}, [{"nome": "N", "PAs": 4}], [], {})
        self.assertFalse(result.viavel)
        self.assertEqual(result.vazios, 3)
        self.assertEqual(result.deficit, 1)


if __name__ == "__main__":
    unittest.main()
