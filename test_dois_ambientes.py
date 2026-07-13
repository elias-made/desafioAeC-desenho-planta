"""
Teste: Simulação de criação de 2 ambientes (como no main.py)

Cenário: 
- N_1: 124 PAs + sala 4 lugares → Bloco_2-A
- N_2: 165 PAs + sala 1 lugar → Bloco_7-A

Este script simula exatamente o fluxo do main.py para identificar
a causa das paredes irregulares quando há 2 ambientes.
"""

import os
import shutil
import openpyxl
import re
from collections import defaultdict

import ScannerPremissas
from AmbienteBuilder import (
    separar_ambiente_e_desenhar_divisorias,
    _gerar_layout_sala_estruturado,
    _selecionar_mesas_contiguas,
    get_env_cells,
    _celulas_contorno_do_ambiente,
)
from openpyxl.styles import PatternFill, Font

SHEET_NAME = 'JPIII'
INPUT_FILE = 'planta.xlsx'
OUTPUT_FILE = 'planta_teste_dois_ambientes.xlsx'

# Cenário exato do posicionador
CRIAR_AMBIENTES = [
    {"bloco": "Bloco_2", "ambiente": "A", "quantidade_mesas": 124, "cliente_destinado": "N_1", "sala_lugares": 4},
    {"bloco": "Bloco_7", "ambiente": "A", "quantidade_mesas": 165, "cliente_destinado": "N_2", "sala_lugares": 1},
]


def main():
    print("=" * 80)
    print("TESTE: Criação de 2 Ambientes (N_1 no Bloco_2, N_2 no Bloco_7)")
    print("=" * 80)

    if not os.path.exists(INPUT_FILE):
        print(f"ERRO: {INPUT_FILE} não encontrado")
        return

    # Copia a planta original para o arquivo de teste
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_NAME]

    resultados = {}
    paleta_cores = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']

    # Processa cada ambiente sequencialmente
    for amb_idx, amb in enumerate(CRIAR_AMBIENTES):
        print(f"\n{'─' * 70}")
        print(f"Iteração {amb_idx + 1}: {amb['cliente_destinado']} em {amb['bloco']}-{amb['ambiente']}")

        # Re-escaneia a planilha - USA PLANTA ORIGINAL (como no main.py corrigido)
        ScannerPremissas._orange_context_cache = {}
        macro_blocks = ScannerPremissas.scan_orange_context("planta.xlsx", SHEET_NAME)

        bloco_id = amb["bloco"]
        ambiente_letra = amb["ambiente"]
        qtd_mesas = amb["quantidade_mesas"]
        cliente_dest = amb["cliente_destinado"]
        sala_lugares = amb.get("sala_lugares", 0)

        # CORREÇÃO AUTOMÁTICA DE LETRA (como no main.py)
        ambientes_anteriores_do_bloco = [
            a for a in CRIAR_AMBIENTES[:amb_idx]
            if a["bloco"] == bloco_id
        ]

        if ambientes_anteriores_do_bloco:
            letra_atual = ambiente_letra.upper()[0] if ambiente_letra else "A"
            letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            idx_letra = letras.index(letra_atual) if letra_atual in letras else 0
            nova_letra = letras[idx_letra + len(ambientes_anteriores_do_bloco) % len(letras)]
            print(f"   ⚠️ Ajustando '{bloco_id}-{ambiente_letra}' para '{bloco_id}-{nova_letra}'")
            ambiente_letra = nova_letra

        # Usa a letra original se o ambiente existir na planta
        # Se não existir (ambiente novo), pega o MAIOR ambiente disponível do bloco
        env_cells_raw_list = get_env_cells(bloco_id, ambiente_letra, macro_blocks)
        if not env_cells_raw_list:
            # Se a letra original não existir, pega o MAIOR ambiente disponível
            block_match = re.search(r'\d+', str(bloco_id))
            if block_match:
                block_idx = int(block_match.group())
                if block_idx <= len(macro_blocks):
                    block = macro_blocks[block_idx - 1]
                    envs = block.get('ambientes', [])
                    if envs:
                        sorted_envs = sorted(envs, key=lambda e: len(e['cells']), reverse=True)
                        env_cells_raw_list = sorted_envs[0]['cells']
                        print(f"   Usando maior ambiente: '{sorted_envs[0]['id']}' com {len(env_cells_raw_list)} células")
                    else:
                        env_cells_raw_list = []
                else:
                    env_cells_raw_list = []
            else:
                env_cells_raw_list = []

        print(f"   Células brutas encontradas: {len(env_cells_raw_list) if env_cells_raw_list else 0}")

        env_cells = set(env_cells_raw_list) if env_cells_raw_list else set()
        env_cells = env_cells | _celulas_contorno_do_ambiente(ws, env_cells)
        print(f"   Células totais (com contorno): {len(env_cells)}")

        if not env_cells:
            print(f"   ERRO: Nenhuma célula encontrada para {cliente_dest}")
            continue

        # Primeiro: planeja a sala (se houver) para reservar o espaço
        allocated_sala = set()
        room_cells_override = None

        if sala_lugares > 0:
            print(f"   🛠️ Planejando Sala Fechada de {sala_lugares} mesas...")
            allocated_sala, room_cells_override = _gerar_layout_sala_estruturado(
                ws, env_cells, sala_lugares
            )
            print(f"   Sala alocada: {len(allocated_sala)} mesas")

        # Remove o espaço da sala do ambiente disponível
        available_env_cells = env_cells - (set(room_cells_override) if room_cells_override else set())
        print(f"   Células disponíveis para salão: {len(available_env_cells)}")

        # Seleciona mesas para o salão aberto
        allocated_ambiente = _selecionar_mesas_contiguas(available_env_cells, ws, qtd_mesas)
        print(f"   Mesas alocadas para salão: {len(allocated_ambiente)}")

        if not allocated_ambiente:
            print(f"   ERRO: Nenhuma mesa encontrada para {cliente_dest}")
            continue

        room_cells_set = set(room_cells_override) if room_cells_override else set()
        allocated_total = allocated_ambiente | room_cells_set

        # Desenha ambiente com as divisórias
        print(f"   🎨 Desenhando divisórias...")
        separar_ambiente_e_desenhar_divisorias(
            ws=ws,
            env_cells=env_cells,
            allocated_cells=allocated_total,
            reconstruir_sala=False,
            room_cells_override=room_cells_set
        )

        # Desenha sala fechada
        if room_cells_set:
            print(f"   🎨 Desenhando sala fechada...")
            separar_ambiente_e_desenhar_divisorias(
                ws=ws,
                env_cells=env_cells,
                allocated_cells=allocated_sala,
                reconstruir_sala=True,
                room_cells_override=room_cells_set
            )

        # Colore apenas as mesas do salão (não os corredores nem as salas fechadas)
        cor_hex = paleta_cores[amb_idx % len(paleta_cores)]
        fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
        font = Font(color='FFFFFF', bold=True, size=8)

        cells_colored = 0
        coords_coloridas = set()
        for r, c in allocated_ambiente:
            cell = ws.cell(row=r, column=c)
            if cell.value != "CT":
                cell.value = cliente_dest
                cell.fill = fill
                cell.font = font
                cells_colored += 1
                coords_coloridas.add((r, c))

        # Colore sala
        if room_cells_set and allocated_sala:
            fill_sala = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
            font_sala = Font(name="Calibri", size=9)
            for r, c in allocated_sala:
                cell = ws.cell(row=r, column=c)
                if cell.value != "CT":
                    cell.value = "vazio"
                    cell.fill = fill_sala
                    cell.font = font_sala

        resultados[cliente_dest] = {
            "cells_colored": cells_colored,
            "all_coords": coords_coloridas,
            "bloco": bloco_id
        }

        print(f"   ✓ Criado ({cells_colored} células coloridas)")
        wb.save(OUTPUT_FILE)
        print(f"   💾 Salvo em: {OUTPUT_FILE}")

    # Resultado
    print(f"\n{'=' * 80}")
    print(f"RESULTADO: {len(resultados)} ambiente(s) criado(s)")
    for cliente, dados in resultados.items():
        print(f"   - {cliente}: {dados['cells_colored']} células no {dados['bloco']}")
    print(f"{'=' * 80}")
    print(f"\nPlanilha de saída: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERRO: {e}")
        traceback.print_exc()