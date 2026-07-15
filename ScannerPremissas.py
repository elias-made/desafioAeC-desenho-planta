# ScannerPremissas.py

import os
import openpyxl
from collections import defaultdict, deque
from typing import Dict, List, Tuple, Set

# Variável global de cache para evitar múltiplas leituras físicas do Excel
# Chaveado por file_path para suportar múltiplos arquivos (planta.xlsx vs proposta_final.xlsx)
_orange_context_cache = {}


def invalidate_orange_cache(ws=None, file_path=None):
    """Invalida apenas a geometria alterada; sem argumentos, limpa todo o cache."""
    global _orange_context_cache
    if ws is None and file_path is None:
        _orange_context_cache.clear()
        return
    prefixes = set()
    if ws is not None:
        prefixes.add(("worksheet", id(ws)))
    if file_path is not None:
        prefixes.add(("file", os.path.abspath(file_path)))
    for key in list(_orange_context_cache):
        if key[:2] in prefixes:
            _orange_context_cache.pop(key, None)

def manhattan_distance(c1: Tuple[int, int], c2: Tuple[int, int]) -> int:
    return abs(c1[0] - c2[0]) + abs(c1[1] - c2[1])

def normalize_val(v) -> str:
    """Normaliza inteiros, floats e textos para comparação correta de strings."""
    if v is None:
        return ""
    val_str = str(v).strip()
    if not val_str:
        return ""
    try:
        f = float(val_str)
        if f == int(f):
            return str(int(f))
        return str(f)
    except ValueError:
        return val_str.upper()

def is_text_annotation(val) -> bool:
    if val is None:
        return False
    val_str = str(val).strip()
    if not val_str:
        return False
    try:
        float(val_str)
        return False
    except ValueError:
        pass
    if val_str.upper() in ('I', 'Q', 'T', "T'", '##'):
        return False
    return True

def is_color_orange_robust(color) -> bool:
    if not color:
        return False
    if color.rgb and color.rgb != '00000000':
        rgb_str = str(color.rgb).strip().upper()
        clean_rgb = rgb_str[-6:]
        if clean_rgb in ('FF9900', 'FFC000', 'ED7D31', 'F79646', 'FFB347', 'FF8C00', 'E26B0A', 'F2994A', 'FF9F43', 'FF9933', 'FFCC00'):
            return True
        try:
            r = int(clean_rgb[0:2], 16)
            g = int(clean_rgb[2:4], 16)
            b = int(clean_rgb[4:6], 16)
            if r > 200 and 100 <= g <= 180 and b < 80:
                return True
        except:
            pass
    if color.type == 'theme' and color.theme is not None:
        if color.theme == 5:
            return True
    return False

def is_barrier_color(color) -> bool:
    if not color:
        return False
    return is_color_orange_robust(color)

def has_blocking_border_between(all_cells_cache: Dict[Tuple[int, int], any], r1: int, c1: int, r2: int, c2: int, direction: str) -> bool:
    if direction == 'horizontal':
        c_left = min(c1, c2)
        c_right = max(c1, c2)
        cell_left = all_cells_cache.get((r1, c_left))
        cell_right = all_cells_cache.get((r1, c_right))
        if not cell_left or not cell_right:
            return False
        b_left = cell_left.border
        if b_left and b_left.right and b_left.right.border_style and b_left.right.border_style != 'none':
            if is_barrier_color(b_left.right.color):
                return True
        b_right = cell_right.border
        if b_right and b_right.left and b_right.left.border_style and b_right.left.border_style != 'none':
            if is_barrier_color(b_right.left.color):
                return True
    else:
        r_top = min(r1, r2)
        r_bottom = max(r1, r2)
        cell_top = all_cells_cache.get((r_top, c1))
        cell_bottom = all_cells_cache.get((r_bottom, c1))
        if not cell_top or not cell_bottom:
            return False
        b_top = cell_top.border
        if b_top and b_top.bottom and b_top.bottom.border_style and b_top.bottom.border_style != 'none':
            if is_barrier_color(b_top.bottom.color):
                return True
        b_bottom = cell_bottom.border
        if b_bottom and b_bottom.top and b_bottom.top.border_style and b_bottom.top.border_style != 'none':
            if is_barrier_color(b_bottom.top.color):
                return True
    return False

def cell_has_orange_fill(cell) -> bool:
    fill = cell.fill
    if fill and fill.patternType == 'solid' and fill.start_color:
        return is_color_orange_robust(fill.start_color)
    return False

def cell_has_orange_border(cell) -> bool:
    border = cell.border
    if border:
        for side_name in ('top', 'bottom', 'left', 'right'):
            side = getattr(border, side_name, None)
            if side and side.border_style and side.border_style != 'none':
                if is_color_orange_robust(getattr(side, 'color', None)):
                    return True
    return False

def cell_has_orange_border_or_fill(cell) -> bool:
    return cell_has_orange_fill(cell) or cell_has_orange_border(cell)

def is_desk_cell(cell) -> bool:
    """Verifica de forma dinâmica se a célula representa uma mesa de operador."""
    if cell is None or cell.value is None:
        return False
    val_str = str(cell.value).strip()
    val_upper = val_str.upper()
    if not val_upper:
        return False
        
    # Barreiras físicas conhecidas não são mesas
    if val_upper in ('##', 'SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'CATRACA', 'CT', 'ESCANINHOS'):
        return False
        
    # Identificadores numericos tambem podem ter dois ou mais digitos
    # (por exemplo, o cliente 10). O fato de ser maior que 9 nao transforma
    # automaticamente a celula em anotacao de capacidade.
    try:
        float(val_upper)
        return True
    except ValueError:
        pass
        
    # Novos clientes (ex: N_1, N_2) e marcas estáveis são considerados mesas
    if val_upper.startswith("N_") or val_upper in ('VAZIO', 'T', "T'", 'ADM', 'I', 'Q'):
        return True
        
    # Identificadores alfanuméricos curtos adicionais
    if len(val_upper) <= 15:
        return True
        
    return False

def is_barrier_cell(cell) -> bool:
    if cell is None:
        return False
    if is_desk_cell(cell):
        return False
    val_str = str(cell.value).strip() if cell.value is not None else ""
    val_upper = val_str.upper()
    if val_upper in ('##', 'SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'CATRACA', 'CT', 'ESCANINHOS'):
        return True
    fill = cell.fill
    if fill and fill.patternType == 'solid' and fill.start_color:
        rgb_str = str(fill.start_color.rgb).upper()
        if rgb_str[-6:] in ('000000', '1A1A1A', '2C3E50'):
            return True
    return False

def get_interior_cells(border_cells: Set[Tuple[int, int]], max_row: int, max_col: int, max_gap: int = 1) -> Set[Tuple[int, int]]:
    """
    Calcula quais células estão no interior do contorno laranja via BFS a partir das bordas
    da planilha inteira, atravessando tudo que não seja parede.

    Usa a mesma tolerância `max_gap` do agrupamento de macro-blocos (scan_orange_context) para
    dilatar virtualmente a parede: qualquer célula a até `max_gap` de distância Manhattan de uma
    célula de borda real é tratada como bloqueio durante a travessia externa. Isso evita que o
    BFS "vaze" por pequenos buracos de 1 célula na parede física (comuns em plantas desenhadas
    manualmente) e classifique erroneamente uma faixa interior como exterior.

    Importante: a dilatação só afeta a travessia (blocking); o resultado final de `interior`
    continua excluindo apenas as células de borda reais (border_cells), então as células do
    "buraco" tratado como bloqueio acabam corretamente incluídas no interior (nunca alcançadas
    pelo BFS externo, e não fazem parte da parede real).
    """
    blocking = set(border_cells)
    if max_gap > 0:
        for (r, c) in border_cells:
            for dr in range(-max_gap, max_gap + 1):
                for dc in range(-max_gap, max_gap + 1):
                    if abs(dr) + abs(dc) <= max_gap:
                        blocking.add((r + dr, c + dc))

    visited = set()
    queue = deque()
    for r in (1, max_row):
        for c in range(1, max_col + 1):
            if (r, c) not in blocking:
                queue.append((r, c))
                visited.add((r, c))
    for c in (1, max_col):
        for r in range(1, max_row + 1):
            if (r, c) not in blocking and (r, c) not in visited:
                queue.append((r, c))
                visited.add((r, c))
    while queue:
        r, c = queue.popleft()
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if 1 <= nr <= max_row and 1 <= nc <= max_col:
                if (nr, nc) not in blocking and (nr, nc) not in visited:
                    visited.add((nr, nc))
                    queue.append((nr, nc))
    interior = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            coord = (r, c)
            if coord not in border_cells and coord not in visited:
                interior.add(coord)
    return interior

def _anotacoes_proximas(cells, all_cells_cache):
    """Coleta anotacoes em raio Manhattan 2 ao redor de um conjunto de celulas."""
    textos = set()
    for r, c in cells:
        for dr in range(-2, 3):
            for dc in range(-2, 3):
                if abs(dr) + abs(dc) > 2:
                    continue
                vizinha = all_cells_cache.get((r + dr, c + dc))
                if vizinha and is_text_annotation(vizinha.value):
                    textos.add(str(vizinha.value).strip())
    return textos

def _deduplicar_macro_blocos_aninhados(macro_blocks, all_cells_cache):
    '''Mantem o contorno externo quando aneis aninhados cercam as mesmas mesas.'''
    mesas_por_indice = []
    for bloco in macro_blocks:
        mesas_por_indice.append({
            coord for coord in bloco.get('interior_cells', set())
            if is_desk_cell(all_cells_cache.get(coord))
        })

    manter = []
    for indice, bloco in enumerate(macro_blocks):
        r0, r1, c0, c1 = bloco['bounding_box']
        duplicado_interno = False
        for outro_indice, outro in enumerate(macro_blocks):
            if indice == outro_indice or not mesas_por_indice[indice]:
                continue
            or0, or1, oc0, oc1 = outro['bounding_box']
            contem = or0 <= r0 and or1 >= r1 and oc0 <= c0 and oc1 >= c1
            estrito = (or0, or1, oc0, oc1) != (r0, r1, c0, c1)
            mesmas_mesas = mesas_por_indice[outro_indice] == mesas_por_indice[indice]
            if contem and estrito and mesmas_mesas:
                duplicado_interno = True
                break
        if not duplicado_interno:
            manter.append(bloco)

    for indice, bloco in enumerate(manter, start=1):
        bloco['id'] = f'Macro_Bloco_{indice}'
    return manter

def scan_orange_context(
    file_path: str = 'planta.xlsx', sheet_name: str = 'JPIII', max_gap: int = 1,
    ws=None,
) -> List[Dict]:
    """Mapeia a geometria laranja reutilizando uma worksheet ja aberta quando fornecida."""
    global _orange_context_cache
    workbook_carregado = None
    if ws is not None:
        cache_key = ("worksheet", id(ws), sheet_name, max_gap)
    else:
        cache_key = ("file", os.path.abspath(file_path), sheet_name, max_gap)
    if cache_key in _orange_context_cache:
        return _orange_context_cache[cache_key]

    if ws is None:
        workbook_carregado = openpyxl.load_workbook(file_path, data_only=True)
        ws = workbook_carregado[sheet_name]
    orange_cells = set()
    cell_values = {}
    all_cells_cache = {}
    
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            all_cells_cache[(r, c)] = cell
            if cell_has_orange_border_or_fill(cell):
                orange_cells.add((r, c))
            if cell.value is not None:
                val_str = str(cell.value).strip()
                if val_str and len(val_str) > 3 and val_str.upper() not in ('VAZIO', 'SALA', 'CATRACA', 'COWORKING'):
                    cell_values[(r, c)] = val_str

    if not orange_cells:
        _orange_context_cache[cache_key] = []
        if workbook_carregado is not None:
            workbook_carregado.close()
        return _orange_context_cache[cache_key]

    orange_list = list(orange_cells)
    n = len(orange_list)
    parent = list(range(n))
    
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[parent[x]]]
            x = parent[x]
        return x
        
    def union(i, j):
        root_i = find(i)
        root_j = find(j)
        if root_i != root_j:
            parent[root_i] = root_j
            
    coord_to_idx = {coord: idx for idx, coord in enumerate(orange_list)}
    for idx, (r, c) in enumerate(orange_list):
        for dr in range(-max_gap, max_gap + 1):
            for dc in range(-max_gap, max_gap + 1):
                if abs(dr) + abs(dc) <= max_gap:
                    neighbor = (r + dr, c + dc)
                    if neighbor in coord_to_idx:
                        union(idx, coord_to_idx[neighbor])
                
    groups = defaultdict(list)
    for i, coord in enumerate(orange_list):
        groups[find(i)].append(coord)
        
    macro_blocks = []
    valid_groups = [coords for coords in groups.values() if len(coords) >= 10]
    
    valid_groups_meta = []
    for coords in valid_groups:
        min_r = min(r for r, c in coords)
        min_c = min(c for r, c in coords)
        valid_groups_meta.append((min_r, min_c, coords))
        
    valid_groups_meta.sort(key=lambda x: x[0])
    
    bands = []
    current_band = []
    row_tolerance = 15
    
    for item in valid_groups_meta:
        if not current_band:
            current_band.append(item)
        else:
            if item[0] - current_band[0][0] <= row_tolerance:
                current_band.append(item)
            else:
                bands.append(current_band)
                current_band = [item]
    if current_band:
        bands.append(current_band)
        
    valid_groups = []
    for band in bands:
        band.sort(key=lambda x: x[1])
        for item in band:
            valid_groups.append(item[2])
    
    idx_counter = 1
    for coords in valid_groups:
        coords_set = set(coords)
        interior_raw = get_interior_cells(coords_set, ws.max_row, ws.max_column, max_gap=max_gap)
        
        boundary_seats = set()
        for r, c in coords_set:
            cell = all_cells_cache.get((r, c))
            if cell and not cell_has_orange_fill(cell) and is_desk_cell(cell):
                boundary_seats.add((r, c))
                
        interior_cells = interior_raw | boundary_seats
        walkable_cells = set()
        for (r, c) in interior_cells:
            cell = all_cells_cache.get((r, c))
            if cell and not is_barrier_cell(cell):
                walkable_cells.add((r, c))
                
        sub_environments = []
        visited_walkable = set()
        for seed in sorted(list(walkable_cells)):
            if seed in visited_walkable:
                continue
            comp = []
            queue = deque([seed])
            visited_walkable.add(seed)
            while queue:
                r, c = queue.popleft()
                comp.append((r, c))
                for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                    nr, nc = r + dr, c + dc
                    neighbor = (nr, nc)
                    if neighbor in walkable_cells and neighbor not in visited_walkable:
                        direction = 'horizontal' if r == nr else 'vertical'
                        if has_blocking_border_between(all_cells_cache, r, c, nr, nc, direction):
                            continue
                        visited_walkable.add(neighbor)
                        queue.append(neighbor)
            if len(comp) >= 3:
                sub_environments.append(set(comp))
                
        valid_sub_environments = []
        for env_cells in sub_environments:
            desk_count = sum(1 for (r, c) in env_cells if is_desk_cell(all_cells_cache.get((r, c))))
            if desk_count >= 3:
                valid_sub_environments.append(env_cells)
        
        all_desks_in_block = set()
        for (r, c) in interior_cells:
            cell = all_cells_cache.get((r, c))
            if cell and is_desk_cell(cell):
                all_desks_in_block.add((r, c))
                
        unassigned_desks = set()
        for seat in all_desks_in_block:
            if not any(seat in env for env in valid_sub_environments):
                unassigned_desks.add(seat)
        
        for seat in unassigned_desks:
            seat_cell = all_cells_cache.get(seat)
            seat_val = normalize_val(seat_cell.value) if seat_cell else ""
            best_env = None
            highest_score = -1
            for env_cells in valid_sub_environments:
                min_dist = min(manhattan_distance(seat, ec) for ec in env_cells)
                if min_dist <= 5:
                    env_r_min = min(ec[0] for ec in env_cells)
                    env_r_max = max(ec[0] for ec in env_cells)
                    env_c_min = min(ec[1] for ec in env_cells)
                    env_c_max = max(ec[1] for ec in env_cells)
                    has_containment = (env_r_min <= seat[0] <= env_r_max) or (env_c_min <= seat[1] <= env_c_max)
                    env_vals = {normalize_val(all_cells_cache[ec].value) for ec in env_cells if all_cells_cache.get(ec)}
                    env_vals.discard('')
                    env_vals.discard('VAZIO')
                    has_affinity = seat_val and (seat_val in env_vals)
                    score = 0
                    if has_affinity:
                        score += 15
                    score += max(0, 6 - min_dist)
                    if has_containment:
                        score += 2
                    if score > highest_score:
                        highest_score = score
                        best_env = env_cells
            if best_env is not None:
                best_env.add(seat)
        
        r_min_macro = min(r for r, c in coords)
        r_max_macro = max(r for r, c in coords)
        c_min_macro = min(c for r, c in coords)
        c_max_macro = max(c for r, c in coords)

        block_texts = set()
        pad = 3
        for (tr, tc), txt in cell_values.items():
            if (r_min_macro - pad) <= tr <= (r_max_macro + pad) and (c_min_macro - pad) <= tc <= (c_max_macro + pad):
                if is_text_annotation(txt):
                    block_texts.add(txt)

        ambientes = []
        if valid_sub_environments:
            valid_sub_environments.sort(key=lambda s: min(s))
            for sub_idx, env_cells in enumerate(valid_sub_environments, start=1):
                env_r_min = min(r for r, c in env_cells)
                env_r_max = max(r for r, c in env_cells)
                env_c_min = min(c for r, c in env_cells)
                env_c_max = max(c for r, c in env_cells)
                env_texts = _anotacoes_proximas(env_cells, all_cells_cache)
                ambientes.append({
                    'id': chr(64 + sub_idx),
                    'bounding_box': (env_r_min, env_r_max, env_c_min, env_c_max),
                    'cells': env_cells,
                    'texts': sorted(list(env_texts))
                })
        else:
            env_texts = _anotacoes_proximas(interior_cells, all_cells_cache)
            ambientes.append({
                'id': "A",
                'bounding_box': (r_min_macro, r_max_macro, c_min_macro, c_max_macro),
                'cells': interior_cells,
                'texts': sorted(list(env_texts))
            })

        macro_blocks.append({
            'id': f"Macro_Bloco_{idx_counter}",
            'texts': sorted(list(block_texts)),
            'bounding_box': (r_min_macro, r_max_macro, c_min_macro, c_max_macro),
            'interior_cells': interior_cells,
            'ambientes': ambientes
        })
        idx_counter += 1
            
    macro_blocks = _deduplicar_macro_blocos_aninhados(macro_blocks, all_cells_cache)
    _orange_context_cache[cache_key] = macro_blocks
    if workbook_carregado is not None:
        workbook_carregado.close()
    return macro_blocks
