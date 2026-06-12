"""
mapa_reduzido.py
================
Gera um mapa ASCII 2D da planta separando:
  1. GRID DE CÉLULAS — conteúdo de cada célula (cliente, vazio, catraca, sala)
  2. PAREDES VERTICAIS — bordas thick vermelhas na esquerda/direita de células
  3. PAREDES HORIZONTAIS — bordas thick vermelhas no topo/base de células

Esse formato permite que a IA:
  - Entenda o layout atual
  - Decida onde alocar novos espaços
  - Devolva paredes novas como coordenadas (coluna, linhas, lado)

Resultado: mapa_reduzido.txt
"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict

SHEET_NAME  = 'JPIII'
OUTPUT_FILE = 'mapa_reduzido.txt'
CLIENTES_ALVO = None   # None = detecta top N automaticamente
TOP_N   = 3
PADDING = 2

# Única definição de "parede de ambiente"
def is_wall(border_side) -> bool:
    """Parede de ambiente = thick vermelho OU medium preto (divisória escura)."""
    if border_side is None or not border_side.border_style:
        return False
    try:
        color = border_side.color.rgb
    except Exception:
        color = None
    style = border_side.border_style
    if style == 'thick'  and color == 'FFFF0000': return True   # parede vermelha
    if style == 'medium' and color == 'FF000000': return True   # divisória escura
    return False


def is_catraca_border(border_side) -> bool:
    """Borda de catraca = medium amarelo OU dotted preto."""
    if border_side is None or not border_side.border_style:
        return False
    try:
        color = border_side.color.rgb
    except Exception:
        color = None
    style = border_side.border_style
    if style == 'medium' and color == 'FFFFFF00': return True   # amarelo
    if style == 'dotted' and color == 'FF000000': return True   # pontilhado
    return False


def cell_symbol(value) -> str:
    """Retorna símbolo de 2 chars para o conteúdo da célula."""
    if value is None:
        return '  '
    # Normaliza float → int quando aplicável
    if isinstance(value, float) and value == int(value):
        value = int(value)
    v = str(value).strip()
    if not v:
        return '  '
    v_up = v.upper()
    if 'CATRACA' in v_up:
        return 'CT'
    if 'COWORKING' in v_up:
        return 'CW'
    if 'SALA' in v_up:
        return 'SA'
    if 'SEM POSSIB' in v_up:
        return '##'
    return v[:2].upper()


def scan(ws):
    """
    Varre o worksheet uma vez.
    Retorna:
      cells[r][c] = {'sym': str, 'wt': bool, 'wb': bool, 'wl': bool, 'wr': bool}
      client_cells: {valor_str: [(r,c),...]}
    """
    cells = defaultdict(dict)
    client_cells = defaultdict(list)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            b    = cell.border
            sym  = cell_symbol(cell.value)
            cells[r][c] = {
                'sym': sym,
                'wt':  is_wall(b.top),
                'wb':  is_wall(b.bottom),
                'wl':  is_wall(b.left),
                'wr':  is_wall(b.right),
                'ct':  is_catraca_border(b.top)  or is_catraca_border(b.bottom)
                       or is_catraca_border(b.left) or is_catraca_border(b.right),
            }
            if sym.strip():
                val_key = sym  # usa o símbolo normalizado como chave
                client_cells[val_key].append((r, c))

    return cells, client_cells


def find_region(cells_list, padding, max_row, max_col):
    rows = [r for r, c in cells_list]
    cols = [c for r, c in cells_list]
    return (
        max(1, min(rows) - padding),
        min(max_row, max(rows) + padding),
        max(1, min(cols) - padding),
        min(max_col, max(cols) + padding),
    )


def render(cells, r_min, r_max, c_min, c_max):
    """
    Retorna (grid_lines, v_walls_desc, h_walls_desc).

    grid_lines: lista de strings do mapa ASCII
    v_walls_desc: descrição textual das paredes verticais
    h_walls_desc: descrição textual das paredes horizontais
    """
    from openpyxl.utils import get_column_letter as gcl

    # ── Grid de células ─────────────────────────────────────────────
    grid_lines = []

    # Cabeçalho de colunas (a cada 5)
    header = '      '
    for c in range(c_min, c_max + 1):
        label = gcl(c)
        header += label[:4].center(4) if (c - c_min) % 5 == 0 else '    '
    grid_lines.append(header)
    grid_lines.append('      ' + '─' * ((c_max - c_min + 1) * 4))

    for r in range(r_min, r_max + 1):
        row = f'{r:>5} '
        for c in range(c_min, c_max + 1):
            cd  = cells[r].get(c, {})
            sym = cd.get('sym', '  ')
            # Célula com borda de catraca e sem valor = marca como entrada de catraca
            if cd.get('ct') and sym == '  ':
                sym = '>>'
            row += f'[{sym}]'
        grid_lines.append(row)

    # ── Paredes verticais ────────────────────────────────────────────
    # Agrupa segmentos contíguos de paredes na mesma coluna e mesmo lado
    # Estrutura: {(col, side): sorted list of rows}
    vwall_map = defaultdict(list)
    for r in range(r_min, r_max + 1):
        for c in range(c_min, c_max + 1):
            cd = cells[r].get(c, {})
            if cd.get('wl'):
                vwall_map[(c, 'left')].append(r)
            if cd.get('wr'):
                vwall_map[(c, 'right')].append(r)

    # Comprime em intervalos contíguos
    def compress_runs(rows):
        if not rows:
            return []
        rows = sorted(rows)
        runs = []
        start = rows[0]
        prev  = rows[0]
        for r in rows[1:]:
            if r == prev + 1:
                prev = r
            else:
                runs.append((start, prev))
                start = prev = r
        runs.append((start, prev))
        return runs

    v_walls_desc = []
    for (col, side), rows in sorted(vwall_map.items()):
        runs = compress_runs(rows)
        for r_s, r_e in runs:
            r_desc = f'linha {r_s}' if r_s == r_e else f'linhas {r_s}-{r_e}'
            v_walls_desc.append(
                f'  Coluna {gcl(col)}, {r_desc}: borda {side} (parede vermelha)'
            )

    # ── Paredes horizontais ──────────────────────────────────────────
    hwall_map = defaultdict(list)
    for r in range(r_min, r_max + 1):
        for c in range(c_min, c_max + 1):
            cd = cells[r].get(c, {})
            if cd.get('wt'):
                hwall_map[(r, 'top')].append(c)
            if cd.get('wb'):
                hwall_map[(r, 'bottom')].append(c)

    h_walls_desc = []
    for (row, side), cols in sorted(hwall_map.items()):
        runs = compress_runs(cols)
        for c_s, c_e in runs:
            c_desc = f'coluna {gcl(c_s)}' if c_s == c_e else f'colunas {gcl(c_s)}-{gcl(c_e)}'
            h_walls_desc.append(
                f'  Linha {row}, {c_desc}: borda {side} (parede vermelha)'
            )

    # ── Catracas (borda amarela/pontilhada) ─────────────────────────
    catraca_cells_desc = []
    for r in range(r_min, r_max + 1):
        for c in range(c_min, c_max + 1):
            cd = cells[r].get(c, {})
            if cd.get('ct'):
                sym = cd.get('sym', '  ')
                catraca_cells_desc.append(
                    f'  {gcl(c)}{r}: [{sym}] — borda de catraca (amarela/pontilhada)'
                )

    return grid_lines, v_walls_desc, h_walls_desc, catraca_cells_desc


def main():
    print(f'Carregando planta.xlsx — aba {SHEET_NAME}...')
    wb  = openpyxl.load_workbook('planta.xlsx', data_only=True)
    ws  = wb[SHEET_NAME]
    print(f'  {ws.max_row} linhas × {ws.max_column} colunas')

    print('  Lendo células e bordas...')
    cells, client_cells = scan(ws)

    # Decide quais clientes mapear
    if CLIENTES_ALVO:
        targets = CLIENTES_ALVO
    else:
        skip = {'CT', 'CW', 'SA', '##', '  '}
        ranked = sorted(
            [(sym, lst) for sym, lst in client_cells.items()
             if sym not in skip and len(lst) > 5],
            key=lambda x: len(x[1]),
            reverse=True,
        )
        targets = [sym for sym, _ in ranked[:TOP_N]]

    print(f'  Mapeando: {targets}')

    out = [
        '=' * 72,
        'MAPA REDUZIDO DA PLANTA — JPIII',
        '=' * 72,
        '',
        'LEGENDA DO GRID:',
        '  [XX] — célula com conteúdo (2 chars)',
        '  [  ] — célula vazia (corredor/espaço livre)',
        '  [CT] — catraca existente (valor na célula)',
        '  [>>] — entrada de catraca (célula vazia com borda de catraca)',
        '  [SA] — sala nomeada',
        '  [CW] — coworking',
        '  [##] — parede/divisória fixa',
        '',
        'PAREDES DE AMBIENTE:',
        '  Listadas separadamente como bordas de células específicas.',
        '  Formato: Coluna X, linhas A-B: borda left/right',
        '           Linha Y, colunas C-D: borda top/bottom',
        '  Para criar novo ambiente: aplicar bordas thick vermelhas',
        '  nos lados indicados das células de fronteira.',
        '',
    ]

    for target in targets:
        cell_list = client_cells.get(target, [])
        if not cell_list:
            continue

        r_min, r_max, c_min, c_max = find_region(
            cell_list, PADDING, ws.max_row, ws.max_column
        )
        n_rows = r_max - r_min + 1
        n_cols = c_max - c_min + 1

        grid_lines, v_walls, h_walls, catraca_desc = render(cells, r_min, r_max, c_min, c_max)

        out += [
            '─' * 72,
            f"CLIENTE '{target}': {len(cell_list)} células",
            f"Região: linhas {r_min}-{r_max} ({n_rows}L) × "
            f"colunas {get_column_letter(c_min)}-{get_column_letter(c_max)} ({n_cols}C)",
            '',
            'GRID:',
        ]
        out.extend(grid_lines)

        out += ['', 'PAREDES VERTICAIS (thick vermelha ou medium preta left/right):']
        out.extend(v_walls if v_walls else ['  (nenhuma nesta região)'])

        out += ['', 'PAREDES HORIZONTAIS (thick vermelha ou medium preta top/bottom):']
        out.extend(h_walls if h_walls else ['  (nenhuma nesta região)'])

        out += ['', 'CATRACAS — borda amarela (medium FFFFFF00) ou pontilhada (dotted):']
        out.extend(catraca_desc if catraca_desc else ['  (nenhuma nesta região)'])
        out.append('')

    out += [
        '=' * 72,
        'COMO A IA DEVE DEVOLVER NOVOS AMBIENTES:',
        '  Para cada novo espaço, retornar:',
        '    células: lista de (linha, coluna) das PAs',
        '    paredes_verticais: [{coluna, linhas_inicio, linhas_fim, lado}]',
        '    paredes_horizontais: [{linha, colunas_inicio, colunas_fim, lado}]',
        '  O código aplicará as bordas thick vermelhas nas células indicadas.',
        '=' * 72,
    ]

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))

    print(f'\n✓ {OUTPUT_FILE} ({len(out)} linhas)')


if __name__ == '__main__':
    main()
