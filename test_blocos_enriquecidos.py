#!/usr/bin/env python
"""
Teste da função build_blocos_info com anotações laranjas
"""

import openpyxl
from BlockMapper import scan_plant
from main import build_blocos_info

# Carrega planta
wb = openpyxl.load_workbook('planta.xlsx', data_only=True)
ws = wb['JPIII']

# Dados da planta
FORBIDDEN_PATTERNS = {'SALA 1', 'SALA 2', 'SALA 3', 'SALA 4', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4'}
plant_data = scan_plant(ws, FORBIDDEN_PATTERNS)

# Gera info enriquecida
print("🔄 Gerando informações enriquecidas dos blocos...")
blocos_info = build_blocos_info(plant_data, ws.max_row, ws.max_column, ws)

print("\n" + "="*80)
print("BLOCOS INFO ENRIQUECIDOS:")
print("="*80)
print(blocos_info)
print("\n✓ Teste concluído!")
