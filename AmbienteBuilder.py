import os
import re
from typing import List, Set, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell

from ScannerPremissas import scan_orange_context


def _eh_celula_mesclada(ws, r: int, c: int) -> bool:
    """Retorna True se a célula (r, c) faz parte de um range mesclado."""
    if r < 1 or r > ws.max_row or c < 1 or c > ws.max_column:
        return False
    return isinstance(ws.cell(row=r, column=c), MergedCell)

# ══════════════════════════════════════════════════════════════════════════
# Funções de Fluxo, Contornos e Divisórias Dinâmicas (Sem Hardcode)
# ══════════════════════════════════════════════════════════════════════════

def _eh_celula_de_mesa_local(cell) -> bool:
    """Verifica de forma dinâmica se a célula representa uma mesa de operador."""
    if cell is None or cell.value is None:
        return False
    from ScannerPremissas import is_desk_cell, is_barrier_cell, cell_has_orange_fill
    
    if is_desk_cell(cell):
        return True
        
    val = str(cell.value).strip().upper()
    if val:
        if is_barrier_cell(cell) or cell_has_orange_fill(cell):
            return False
        if val in ('SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'CATRACA', 'CT', 'ESCANINHOS', '##'):
            return False
        try:
            val_num = float(val)
            if val_num > 9:
                return False
        except ValueError:
            pass
        return True
    return False

def _tem_parede_laranja_entre(ws, r1, c1, r2, c2) -> bool:
    """Verifica se existe uma parede física laranjada (borda ou preenchimento) entre as duas células."""
    if r1 < 1 or r1 > ws.max_row or c1 < 1 or c1 > ws.max_column:
        return True
    if r2 < 1 or r2 > ws.max_row or c2 < 1 or c2 > ws.max_column:
        return True
        
    cell1 = ws.cell(row=r1, column=c1)
    cell2 = ws.cell(row=r2, column=c2)
    
    from ScannerPremissas import is_color_orange_robust, cell_has_orange_fill
    
    # CORREÇÃO: Células com preenchimento sólido de parede laranja funcionam como barreira sólida
    if cell_has_orange_fill(cell1) or cell_has_orange_fill(cell2):
        return True
        
    def is_orange(side) -> bool:
        if not side or not side.style or side.style == 'none':
            return False
        return is_color_orange_robust(getattr(side, 'color', None))

    if r1 == r2:
        if c2 < c1:
            return is_orange(cell1.border.left) or is_orange(cell2.border.right)
        return is_orange(cell1.border.right) or is_orange(cell2.border.left)
    elif c1 == c2:
        if r2 < r1:
            return is_orange(cell1.border.top) or is_orange(cell2.border.bottom)
        return is_orange(cell1.border.bottom) or is_orange(cell2.border.top)
                
    return False

def _eh_catraca_do_nosso_ambiente(ws, env_cells: Set[Tuple[int, int]], r: int, c: int) -> bool:
    """Verifica se a célula é uma catraca que está fisicamente encostada no nosso ambiente."""
    if (r, c) in env_cells:
        return True
    for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
        nr, nc = r + dr, c + dc
        if (nr, nc) in env_cells:
            if not _tem_parede_laranja_entre(ws, r, c, nr, nc):
                return True
    return False

def _eh_pilar_ou_coluna(ws, r, c) -> bool:
    """Verifica se a célula especificada é um pilar ou coluna física estrutural."""
    if r < 1 or r > ws.max_row or c < 1 or c > ws.max_column:
        return False
    cell = ws.cell(row=r, column=c)
    from ScannerPremissas import is_barrier_cell
    if is_barrier_cell(cell):
        val = str(cell.value).strip().upper() if cell.value is not None else ""
        if val in ('SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'CATRACA', 'CT', 'ESCANINHOS'):
            return False
        return True
    return False

def _eh_faixa_livre(ws, env_cells: Set[Tuple[int, int]], r0, r1, c0, c1) -> bool:
    """Verifica se todas as células no retângulo são corredores livres."""
    if r0 < 1 or r1 > ws.max_row or c0 < 1 or c1 > ws.max_column:
        return False
    from ScannerPremissas import is_barrier_cell, cell_has_orange_fill
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value).strip().upper() if cell.value is not None else ""
            
            # CORREÇÃO: Células de outros clientes (N_*) ou salas (vazio) não são livres
            if _eh_celula_de_mesa_local(cell) or val.startswith("N_") or val == "VAZIO":
                return False
            
            eh_nossa_catraca = False
            if val in ("CT", "CATRACA"):
                if _eh_catraca_do_nosso_ambiente(ws, env_cells, r, c):
                    eh_nossa_catraca = True
            
            if not eh_nossa_catraca:
                if is_barrier_cell(cell):
                    if val in ('SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'ESCANINHOS'):
                        if (r, c) not in env_cells:
                            return False
                    else:
                        return False
                        
                if cell_has_orange_fill(cell):
                    return False
                    
                if val in ("CT", "CATRACA"):
                    return False
    return True

def _celulas_contorno_do_ambiente(ws, env_cells: Set[Tuple[int, int]]) -> Set[Tuple[int, int]]:
    """Retorna as células do contorno laranjas que são piso utilizável."""
    from ScannerPremissas import cell_has_orange_fill, cell_has_orange_border_or_fill

    extra: Set[Tuple[int, int]] = set()
    for (r, c) in env_cells:
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if (nr, nc) in env_cells:
                continue
            if nr < 1 or nr > ws.max_row or nc < 1 or nc > ws.max_column:
                continue
                
            cell = ws.cell(row=nr, column=nc)
            if isinstance(cell, MergedCell):
                continue
            if cell_has_orange_fill(cell):
                continue
            
            if not cell_has_orange_border_or_fill(cell):
                continue
                
            val = str(cell.value).strip().upper() if cell.value is not None else ""
            
            if _eh_celula_de_mesa_local(cell):
                continue
            if val in ('SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'ESCANINHOS'):
                continue
            
            if not _tem_parede_laranja_entre(ws, r, c, nr, nc):
                extra.add((nr, nc))
                
    return extra

def _verificar_acessibilidade_mesas(
    ws, 
    env_cells: Set[Tuple[int, int]], 
    room_cells: Set[Tuple[int, int]], 
    desk_cells: Set[Tuple[int, int]]
) -> Tuple[bool, Set[Tuple[int, int]]]:
    """Verifica se todas as mesas têm ao menos um caminho de corredor até a saída do ambiente."""
    from collections import deque
    
    corridor_cells = room_cells - desk_cells
    
    exit_cells: Set[Tuple[int, int]] = set()
    for (r, c) in corridor_cells:
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if (nr, nc) not in room_cells:
                if not _tem_parede_laranja_entre(ws, r, c, nr, nc):
                    exit_cells.add((r, c))
                    break
    
    if not exit_cells:
        return False, set(desk_cells)
    
    alcancaveis: Set[Tuple[int, int]] = set()
    fila = deque(exit_cells)
    alcancaveis.update(exit_cells)
    
    while fila:
        r, c = fila.popleft()
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if (nr, nc) in alcancaveis:
                continue
            if (nr, nc) not in corridor_cells:
                continue
            if _tem_parede_laranja_entre(ws, r, c, nr, nc):
                continue
            alcancaveis.add((nr, nc))
            fila.append((nr, nc))
    
    mesas_sem_saida: Set[Tuple[int, int]] = set()
    for (r, c) in desk_cells:
        tem_acesso = False
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if (nr, nc) in alcancaveis and not _tem_parede_laranja_entre(ws, r, c, nr, nc):
                tem_acesso = True
                break
        if not tem_acesso:
            mesas_sem_saida.add((r, c))
    
    return len(mesas_sem_saida) == 0, mesas_sem_saida


def _expandir_corredor_para_acessibilidade(
    ws,
    env_cells: Set[Tuple[int, int]],
    room_cells: Set[Tuple[int, int]],
    desk_cells: Set[Tuple[int, int]],
    mesas_sem_saida: Set[Tuple[int, int]]
) -> Set[Tuple[int, int]]:
    """Tenta expandir o room_cells adicionando corredores para conectar mesas isoladas.
    
    IMPORTANTE: Respeita divisórias desenhadas anteriormente (células com valor "N_" ou bordas laranjas).
    """
    from collections import deque
    
    corridor_cells = room_cells - desk_cells
    
    exit_cells: Set[Tuple[int, int]] = set()
    for (r, c) in corridor_cells:
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if (nr, nc) not in room_cells:
                if not _tem_parede_laranja_entre(ws, r, c, nr, nc):
                    exit_cells.add((r, c))
                    break
    
    alcancaveis: Set[Tuple[int, int]] = set()
    fila = deque(exit_cells)
    alcancaveis.update(exit_cells)
    while fila:
        r, c = fila.popleft()
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if (nr, nc) in alcancaveis:
                continue
            if (nr, nc) not in corridor_cells:
                continue
            if _tem_parede_laranja_entre(ws, r, c, nr, nc):
                continue
            alcancaveis.add((nr, nc))
            fila.append((nr, nc))
    
    novos_corredores: Set[Tuple[int, int]] = set()
    
    for (mr, mc) in mesas_sem_saida:
        visitado: Set[Tuple[int, int]] = set()
        predecessores: dict = {}
        fila_bfs = deque()
        
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = mr + dr, mc + dc
            if (nr, nc) in desk_cells:
                continue
            if (nr, nc) not in env_cells:
                continue
            # CORREÇÃO: Não cruza paredes laranjas de ambientes anteriores
            if _tem_parede_laranja_entre(ws, mr, mc, nr, nc):
                continue
            # CORREÇÃO: Não cruza células de outros clientes (N_1, N_2, etc)
            cell_adj = ws.cell(row=nr, column=nc)
            val_adj = str(cell_adj.value).strip().upper() if cell_adj.value else ""
            if val_adj.startswith("N_") or val_adj in ("CT", "vazio"):
                continue
            if _eh_pilar_ou_coluna(ws, nr, nc):
                continue
            fila_bfs.append((nr, nc))
            visitado.add((nr, nc))
            predecessores[(nr, nc)] = None
        
        destino = None
        while fila_bfs and destino is None:
            r, c = fila_bfs.popleft()
            
            if (r, c) in alcancaveis:
                destino = (r, c)
                break
            
            for dr2, dc2 in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                nr2, nc2 = r + dr2, c + dc2
                if (nr2, nc2) not in room_cells and (nr2, nc2) not in env_cells:
                    if not _tem_parede_laranja_entre(ws, r, c, nr2, nc2):
                        destino = (r, c)
                        break
            if destino:
                break
            
            for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                nr, nc = r + dr, c + dc
                if (nr, nc) in visitado:
                    continue
                if (nr, nc) in desk_cells:
                    continue
                if (nr, nc) not in env_cells:
                    continue
                # CORREÇÃO: Não cruza paredes laranjas
                if _tem_parede_laranja_entre(ws, r, c, nr, nc):
                    continue
                # CORREÇÃO: Não cruza células de outros clientes
                cell_next = ws.cell(row=nr, column=nc)
                val_next = str(cell_next.value).strip().upper() if cell_next.value else ""
                if val_next.startswith("N_") or val_next in ("CT", "vazio"):
                    continue
                if _eh_pilar_ou_coluna(ws, nr, nc):
                    continue
                visitado.add((nr, nc))
                predecessores[(nr, nc)] = (r, c)
                fila_bfs.append((nr, nc))
        
        if destino is not None:
            atual = destino
            while atual is not None:
                novos_corredores.add(atual)
                atual = predecessores.get(atual)
    
    return room_cells | novos_corredores


def _absorver_gaps_estreitos(ws, env_cells: Set[Tuple[int, int]], room_cells: Set[Tuple[int, int]]) -> Set[Tuple[int, int]]:
    """Absorve frestas de 1 a 2 células de largura contra as paredes para evitar novos contornos flutuantes.
    
    IMPORTANTE: Respeita células de outros clientes (N_1, N_2, etc).
    """
    room_cells = set(room_cells)
    changed = True
    while changed:
        changed = False
        remaining = env_cells - room_cells
        to_absorb = set()
        for r, c in remaining:
            # CORREÇÃO: Não absorve células de outros clientes
            cell = ws.cell(row=r, column=c)
            val = str(cell.value).strip().upper() if cell.value else ""
            if val.startswith("N_") or val in ("CT", "vazio"):
                continue
                
            adj_to_room = False
            for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                if (r + dr, c + dc) in room_cells:
                    adj_to_room = True
                    break
            if not adj_to_room:
                continue
            
            for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                if (r - dr, c - dc) in room_cells:
                    gap_size = 0
                    curr_r, curr_c = r, c
                    is_blocked = False
                    while gap_size < 3:
                        curr_r += dr
                        curr_c += dc
                        if (curr_r, curr_c) not in env_cells or _tem_parede_laranja_entre(ws, curr_r - dr, curr_c - dc, curr_r, curr_c):
                            is_blocked = True
                            break
                        # CORREÇÃO: Verifica se célula atual é de outro cliente
                        cell_curr = ws.cell(row=curr_r, column=curr_c)
                        val_curr = str(cell_curr.value).strip().upper() if cell_curr.value else ""
                        if val_curr.startswith("N_") or val_curr in ("CT", "vazio"):
                            is_blocked = True
                            break
                        if (curr_r, curr_c) in room_cells:
                            break
                        gap_size += 1
                    
                    if is_blocked and gap_size <= 1:
                        to_absorb.add((r, c))
                        break
        
        if to_absorb:
            room_cells.update(to_absorb)
            changed = True
            
    return room_cells

def get_env_cells(block_id_str: str, env_letter: str, macro_blocks: List[dict]) -> List[Tuple[int, int]]:
    if not block_id_str or not env_letter:
        return []
    block_match = re.search(r'\d+', str(block_id_str))
    if not block_match:
        return []
    block_idx = int(block_match.group())
    if block_idx <= len(macro_blocks):
        block = macro_blocks[block_idx - 1]
        block_envs = {e['id'].upper(): e for e in block.get('ambientes', [])}
        
        target_env = env_letter.upper()
        if target_env not in block_envs:
            # Se não encontrar o ambiente específico (pois é novo), pega o de maior capacidade do bloco (área aberta principal)
            if block_envs:
                sorted_envs = sorted(block_envs.values(), key=lambda e: len(e['cells']), reverse=True)
                target_env = sorted_envs[0]['id'].upper()
            else:
                return []
            
        parts = re.split(r'[-_\s+&,|/]+', target_env)
        matched_envs = []
        for p in parts:
            if p in block_envs:
                matched_envs.append(block_envs[p])
        if matched_envs:
            cells = []
            for env in matched_envs:
                cells.extend(env['cells'])
            return cells
    return []

# ══════════════════════════════════════════════════════════════════════════
# Geração de Layout de Sala Estruturada com Circulação Interna Própria
# ══════════════════════════════════════════════════════════════════════════

def _calcular_altura_sala(W: int, N: int) -> int:
    desks_per_row = 4 if W == 5 else 2
    h_desks = (N + desks_per_row - 1) // desks_per_row
    h_corridors = (h_desks - 1) // 3 if h_desks > 3 else 0
    return h_desks + h_corridors

def _eh_celula_valida_para_sala(ws, r: int, c: int, env_cells: Set[Tuple[int, int]]) -> bool:
    """Verifica se a célula está dentro do bloco, evitando pilares e outras salas."""
    if (r, c) not in env_cells:
        return False
    if _eh_pilar_ou_coluna(ws, r, c):
        return False
        
    from ScannerPremissas import cell_has_orange_fill
    cell = ws.cell(row=r, column=c)
    if cell_has_orange_fill(cell):
        return False
        
    val_str = str(cell.value).strip().upper() if cell.value is not None else ""
    if val_str in ('SALA', 'COWORKING', 'SALA CLIENTE', 'SALA1', 'SALA2', 'SALA3', 'SALA4', 'CATRACA', 'CT', 'ESCANINHOS'):
        return False
        
    return True

def _remocao_preserva_conectividade(env_cells: Set[Tuple[int, int]], rect_cells: Set[Tuple[int, int]]) -> bool:
    from BlockMapper import flood_fill
    restante = env_cells - rect_cells
    if not restante:
        return True
    componentes = flood_fill(restante)
    return len(componentes) <= 1

def _celula_eh_corredor_circulacao(ws, env_cells: Set[Tuple[int, int]], r: int, c: int) -> bool:
    if (r, c) not in env_cells:
        return False
    from ScannerPremissas import cell_has_orange_border, cell_has_orange_fill
    cell = ws.cell(row=r, column=c)
    if cell_has_orange_border(cell) or cell_has_orange_fill(cell):
        return False
    return _eh_faixa_livre(ws, env_cells, r, r, c, c)

def _classificar_celulas_sala(ws, env_cells: Set[Tuple[int, int]]):
    valid_cells: Set[Tuple[int, int]] = set()
    desk_cells: Set[Tuple[int, int]] = set()
    corridor_cells: Set[Tuple[int, int]] = set()
    for (r, c) in env_cells:
        if _eh_celula_valida_para_sala(ws, r, c, env_cells):
            valid_cells.add((r, c))
        cell = ws.cell(row=r, column=c)
        if _eh_celula_de_mesa_local(cell):
            desk_cells.add((r, c))
        if _celula_eh_corredor_circulacao(ws, env_cells, r, c):
            corridor_cells.add((r, c))
    return valid_cells, desk_cells, corridor_cells

def _encontrar_melulo_retangulo_sala(ws, env_cells: Set[Tuple[int, int]], W: int, H: int, precomp=None) -> Tuple[int, int, Set[Tuple[int, int]]]:
    """Encontra o melhor retângulo W x H para a sala fechada, priorizando o corredor de ligação."""
    if not env_cells:
        return None, None, set()
        
    r_max = max(r for r, c in env_cells)
    c_max = max(c for r, c in env_cells)
    
    rows = sorted(list({r for r, c in env_cells}))
    cols = sorted(list({c for r, c in env_cells}))

    if precomp is not None:
        valid_cells, desk_cells, corridor_cells = precomp
    else:
        valid_cells, desk_cells, corridor_cells = _classificar_celulas_sala(ws, env_cells)

    def _gap(r0, r1, c0, c1, direcao):
        g, limite = 0, 3
        if direcao in ('left', 'right'):
            c = c0 if direcao == 'left' else c1
            step = -1 if direcao == 'left' else 1
            while g < limite:
                c += step
                if not all((r, c) in valid_cells for r in range(r0, r1 + 1)):
                    break
                g += 1
        else:
            r = r0 if direcao == 'top' else r1
            step = -1 if direcao == 'top' else 1
            while g < limite:
                r += step
                if not all((r, c) in valid_cells for c in range(c0, c1 + 1)):
                    break
                g += 1
        return g

    def _conecta(r0, r1, c0, c1):
        for c in range(c0, c1 + 1):
            if (r0 - 1, c) in corridor_cells and not _tem_parede_laranja_entre(ws, r0, c, r0 - 1, c):
                return True
            if (r1 + 1, c) in corridor_cells and not _tem_parede_laranja_entre(ws, r1, c, r1 + 1, c):
                return True
        for r in range(r0, r1 + 1):
            if (r, c0 - 1) in corridor_cells and not _tem_parede_laranja_entre(ws, r, c0, r, c0 - 1):
                return True
            if (r, c1 + 1) in corridor_cells and not _tem_parede_laranja_entre(ws, r, c1, r, c1 + 1):
                return True
        return False

    candidates = []
    for r_start in rows:
        if r_start + H - 1 > r_max:
            break
        r_end = r_start + H - 1
        for c_start in cols:
            if c_start + W - 1 > c_max:
                break
            c_end = c_start + W - 1

            if 0 < _gap(r_start, r_end, c_start, c_end, 'left') < 3:
                continue
            if 0 < _gap(r_start, r_end, c_start, c_end, 'right') < 3:
                continue
            if 0 < _gap(r_start, r_end, c_start, c_end, 'top') < 3:
                continue
            if 0 < _gap(r_start, r_end, c_start, c_end, 'bottom') < 3:
                continue

            rect_cells = set()
            desks_destroyed = 0
            valido = True
            for r in range(r_start, r_end + 1):
                for c in range(c_start, c_end + 1):
                    if (r, c) not in valid_cells:
                        valido = False
                        break
                    rect_cells.add((r, c))
                    if (r, c) in desk_cells:
                        desks_destroyed += 1
                if not valido:
                    break
            if not valido:
                continue

            touching_sides = 0
            if any((r_start - 1, c) not in valid_cells for c in range(c_start, c_end + 1)):
                touching_sides += 1
            if any((r_end + 1, c) not in valid_cells for c in range(c_start, c_end + 1)):
                touching_sides += 1
            if any((r, c_start - 1) not in valid_cells for r in range(r_start, r_end + 1)):
                touching_sides += 1
            if any((r, c_end + 1) not in valid_cells for r in range(r_start, r_end + 1)):
                touching_sides += 1

            conecta = 1 if _conecta(r_start, r_end, c_start, c_end) else 0
            candidates.append((conecta, touching_sides, desks_destroyed, r_start, c_start, rect_cells))

    if not candidates:
        return None, None, set()

    candidates.sort(key=lambda x: (-x[0], -x[1], x[2], x[3], x[4]))

    for _conecta_flag, _touch, _destroyed, r_start, c_start, rect_cells in candidates:
        if _remocao_preserva_conectividade(env_cells, rect_cells):
            return r_start, c_start, rect_cells

    return None, None, set()

def _gerar_bancadas_dinamicas(ws, env_cells: Set[Tuple[int, int]], N: int) -> Set[Tuple[int, int]]:
    usable = {coord for coord in env_cells if not _eh_pilar_ou_coluna(ws, coord[0], coord[1])}
    if len(usable) < N:
        return set(list(usable)[:N])
        
    rows = sorted(list({r for r, c in usable}))
    cols = sorted(list({c for r, c in usable}))
    
    col_pairs = []
    i = 0
    while i < len(cols) - 1:
        if cols[i+1] == cols[i] + 1:
            col_pairs.append((cols[i], cols[i+1]))
            i += 3
        else:
            i += 1
            
    row_pairs = []
    i = 0
    while i < len(rows) - 1:
        if rows[i+1] == rows[i] + 1:
            row_pairs.append((rows[i], rows[i+1]))
            i += 3
        else:
            i += 1
            
    desks = set()
    
    if len(col_pairs) >= len(row_pairs) and col_pairs:
        for c1, c2 in col_pairs:
            if len(desks) >= N:
                break
            common_rows = sorted([r for r in rows if (r, c1) in usable and (r, c2) in usable])
            for r in common_rows:
                if len(desks) >= N:
                    break
                desks.add((r, c1))
                if len(desks) >= N:
                    break
                desks.add((r, c2))
    elif row_pairs:
        for r1, r2 in row_pairs:
            if len(desks) >= N:
                break
            common_cols = sorted([c for c in cols if (r1, c) in usable and (r2, c) in usable])
            for c in common_cols:
                if len(desks) >= N:
                    break
                desks.add((r1, c))
                if len(desks) >= N:
                    break
                desks.add((r2, c))
    else:
        sorted_usable = sorted(list(usable), key=lambda x: (x[1], x[0]))
        desks = set(sorted_usable[:N])
        
    # REGRA DE PREENCHIMENTO RESIDUAL
    if len(desks) < N:
        remaining_usable = sorted(list(usable - desks), key=lambda x: (x[1], x[0]))
        for cell in remaining_usable:
            if len(desks) >= N:
                break
            desks.add(cell)
        
    return desks

def _gerar_layout_sala_estruturado(ws, env_cells: Set[Tuple[int, int]], N: int) -> Tuple[Set[Tuple[int, int]], Set[Tuple[int, int]]]:
    env_cells = set(env_cells) | _celulas_contorno_do_ambiente(ws, env_cells)

    if N <= 0:
        return set(), set()

    import math

    precomp = _classificar_celulas_sala(ws, env_cells)

    def _dims(orient, L):
        prof = (N + 2 * L - 1) // (2 * L)
        if orient == 'v':
            return 3 * L + 1, prof + 1, prof
        return prof + 1, 3 * L + 1, prof

    L0 = max(1, int(round(math.sqrt(N / 6.0))))
    Ls = sorted({l for l in range(max(1, L0 - 2), L0 + 3)} | {1})

    melhor = None
    for orient in ('v', 'h'):
        for L in Ls:
            W, H, prof = _dims(orient, L)
            if 2 * L * prof < N:
                continue
            r, c, _rect = _encontrar_melulo_retangulo_sala(ws, env_cells, W, H, precomp=precomp)
            if r is None:
                continue
            chave = (max(W, H), W * H)
            cand = (chave, orient, L, r, c, W, H, prof)
            if melhor is None or cand[0] < melhor[0]:
                melhor = cand

    if melhor is not None:
        _chave, orient, L, r_start, c_start, W, H, prof = melhor
        room_all_cells = {(r, c)
                          for r in range(r_start, r_start + H)
                          for c in range(c_start, c_start + W)}
        desks_set = set()
        placed = 0
        if orient == 'v':
            for i in range(L):
                dc1, dc2 = c_start + 3 * i + 1, c_start + 3 * i + 2
                for r in range(r_start, r_start + prof):
                    for dc in (dc1, dc2):
                        if placed < N:
                            desks_set.add((r, dc)); placed += 1
        else:
            for i in range(L):
                dr1, dr2 = r_start + 3 * i + 1, r_start + 3 * i + 2
                for c in range(c_start, c_start + prof):
                    for dr in (dr1, dr2):
                        if placed < N:
                            desks_set.add((dr, c)); placed += 1

        if placed >= N:
            return desks_set, room_all_cells

    allocated = _gerar_bancadas_dinamicas(ws, env_cells, N)
    return allocated, allocated

# ══════════════════════════════════════════════════════════════════════════
# Funções de Desenho, Limpeza e Formatação
# ══════════════════════════════════════════════════════════════════════════

def separar_ambiente_e_desenhar_divisorias(
    ws, 
    env_cells: Set[Tuple[int, int]], 
    allocated_cells: Set[Tuple[int, int]], 
    border_style: str = "medium", 
    border_color: str = "FF9900",
    reconstruir_sala: bool = False,
    room_cells_override: Set[Tuple[int, int]] = None
) -> dict:
    """
    Desenha as divisórias ao redor das bancadas e suporta a criação de salas estruturadas.
    
    Returns:
        dict com valido, motivo, room_cells, corridor_cells, ct_cell, mesas_sem_saida, relocated_cells
    """
    resultado = {
        "valido": False,
        "motivo": "",
        "room_cells": set(),
        "corridor_cells": set(),
        "ct_cell": None,
        "mesas_sem_saida": set(),
        "relocated_cells": set()
    }
    
    if not allocated_cells or not env_cells:
        resultado["motivo"] = "sem células alocadas ou ambiente"
        return resultado

    from BlockMapper import flood_fill

    env_cells = set(env_cells) | _celulas_contorno_do_ambiente(ws, env_cells)
    
    side_style = Side(border_style=border_style, color=border_color)
    
    # CORREÇÃO 1: Isola as mesas reais do ambiente principal das células da sala estruturada
    if room_cells_override is not None:
        mesas_reais = set(allocated_cells) - set(room_cells_override)
        if not mesas_reais:
            mesas_reais = set(allocated_cells)
        allocated = mesas_reais & env_cells
    else:
        allocated = set(allocated_cells) & env_cells
        if not allocated:
            allocated = set(allocated_cells)

    if room_cells_override is not None and reconstruir_sala:
        room_cells = set(room_cells_override)
        target_bench_cells = set(allocated_cells) & room_cells
        if not target_bench_cells:
            target_bench_cells = set(allocated_cells)
        
        # LIMPEZA CRÍTICA: Apaga todas as bordas internas antigas de todas as células da nova sala
        for r, c in room_cells:
            if not _eh_celula_mesclada(ws, r, c):
                cell = ws.cell(row=r, column=c)
                cell.border = Border()  # Reseta qualquer borda preexistente interna

        fill_desk = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
        font_desk = Font(name="Calibri", size=9)
        for r, c in target_bench_cells:
            if _eh_celula_mesclada(ws, r, c):
                continue
            cell = ws.cell(row=r, column=c)
            cell.value = "vazio"
            cell.fill = fill_desk
            cell.font = font_desk
    else:
        all_desks_in_env = set()
        for r, c in env_cells:
            # CORREÇÃO 2: Não misturar o espaço da sala estruturada com a varredura de mesas comuns
            if room_cells_override is not None and (r, c) in room_cells_override:
                continue
            cell = ws.cell(row=r, column=c)
            val_str = str(cell.value).strip().upper() if cell.value is not None else ""
            # CORREÇÃO: Só inclui célula se pertence ao cliente atual OU é mesa livre
            # NÃO inclui células de OUTROS clientes (N_1, N_2, etc)
            if (r, c) in allocated:
                all_desks_in_env.add((r, c))
            elif _eh_celula_de_mesa_local(cell) and not val_str.startswith("N_"):
                all_desks_in_env.add((r, c))

        benches = flood_fill(all_desks_in_env)
        target_bench_cells = set()
        
        for bench in benches:
            bench_set = set(bench)
            allocated_in_bench = bench_set & allocated
            
            if allocated_in_bench:
                b_rows = {r for r, c in bench_set}
                b_cols = {c for r, c in bench_set}
                physical_rows_count = len(b_rows)
                physical_cols_count = len(b_cols)
                
                allocated_rows_count = len({r for r, c in allocated_in_bench})
                allocated_cols_count = len({c for r, c in allocated_in_bench})
                
                corte_colunas_intencional = allocated_cols_count < physical_cols_count
                corte_linhas_intencional = allocated_rows_count < physical_rows_count
                
                total_capacity = len(bench_set)
                allocated_count = len(allocated_in_bench)
                missing_to_complete = total_capacity - allocated_count
                
                excesso_potencial_global = (len(target_bench_cells) + missing_to_complete) - len(allocated)
                max_excesso_permitido = 0 if len(allocated) >= len(allocated_cells) else 2
                
                if (missing_to_complete <= 2 
                        and not corte_colunas_intencional 
                        and not corte_linhas_intencional 
                        and excesso_potencial_global <= max_excesso_permitido):
                    target_bench_cells.update(bench_set)
                else:
                    target_bench_cells.update(allocated_in_bench)

        if not target_bench_cells:
            target_bench_cells = allocated

        room_cells = set(target_bench_cells)
        
        for bench in benches:
            bench_set = set(bench)
            bench_allocated = bench_set & target_bench_cells
            if not bench_allocated:
                continue
                
            r0_b = min(r for r, c in bench_allocated)
            r1_b = max(r for r, c in bench_allocated)
            c0_b = min(c for r, c in bench_allocated)
            c1_b = max(c for r, c in bench_allocated)
            
            b_rows = {r for r, c in bench_set}
            b_cols = {c for r, c in bench_set}
            is_vertical = len(b_rows) >= len(b_cols)
            
            if is_vertical:
                for r in range(r0_b, r1_b + 1):
                    if (r, c0_b) in target_bench_cells:
                        target_cell = (r, c0_b - 1)
                        if not _tem_parede_laranja_entre(ws, r, c0_b, r, c0_b - 1):
                            if _eh_faixa_livre(ws, env_cells, r, r, c0_b - 1, c0_b - 1):
                                room_cells.add(target_cell)
                for r in range(r0_b, r1_b + 1):
                    if (r, c1_b) in target_bench_cells:
                        target_cell = (r, c1_b + 1)
                        if not _tem_parede_laranja_entre(ws, r, c1_b, r, c1_b + 1):
                            if _eh_faixa_livre(ws, env_cells, r, r, c1_b + 1, c1_b + 1):
                                room_cells.add(target_cell)
                left_col = c0_b - 1 if (r1_b, c0_b - 1) in room_cells else c0_b
                right_col = c1_b + 1 if (r1_b, c1_b + 1) in room_cells else c1_b
                
                miolo_livre_inferior = True
                for c in range(c0_b, c1_b + 1):
                    if not _eh_faixa_livre(ws, env_cells, r1_b + 1, r1_b + 1, c, c) or _tem_parede_laranja_entre(ws, r1_b, c, r1_b + 1, c):
                        miolo_livre_inferior = False
                        break
                        
                if miolo_livre_inferior:
                    for c in range(left_col, right_col + 1):
                        target_cell = (r1_b + 1, c)
                        if not _tem_parede_laranja_entre(ws, r1_b, c, r1_b + 1, c):
                            if _eh_faixa_livre(ws, env_cells, r1_b + 1, r1_b + 1, c, c):
                                room_cells.add(target_cell)
            else:
                for c in range(c0_b, c1_b + 1):
                    if (r0_b, c) in target_bench_cells:
                        target_cell = (r0_b - 1, c)
                        if not _tem_parede_laranja_entre(ws, r0_b, c, r0_b - 1, c):
                            if _eh_faixa_livre(ws, env_cells, r0_b - 1, r0_b - 1, c, c):
                                room_cells.add(target_cell)
                for c in range(c0_b, c1_b + 1):
                    if (r1_b, c) in target_bench_cells:
                        target_cell = (r1_b + 1, c)
                        if not _tem_parede_laranja_entre(ws, r1_b, c, r1_b + 1, c):
                            if _eh_faixa_livre(ws, env_cells, r1_b + 1, r1_b + 1, c, c):
                                room_cells.add(target_cell)
                top_row = r0_b - 1 if (r0_b - 1, c0_b) in room_cells else r0_b
                bottom_row = r1_b + 1 if (r1_b + 1, c0_b) in room_cells else r1_b
                
                miolo_livre_esquerdo = True
                for r in range(r0_b, r1_b + 1):
                    if not _eh_faixa_livre(ws, env_cells, r, r, c0_b - 1, c0_b - 1) or _tem_parede_laranja_entre(ws, r, c0_b, r, c0_b - 1):
                        miolo_livre_esquerdo = False
                        break
                        
                if miolo_livre_esquerdo:
                    for r in range(top_row, bottom_row + 1):
                        target_cell = (r, c0_b - 1)
                        if not _tem_parede_laranja_entre(ws, r, c0_b, r, c0_b - 1):
                            if _eh_faixa_livre(ws, env_cells, r, r, c0_b - 1, c0_b - 1):
                                room_cells.add(target_cell)
                                
                top_row = r0_b - 1 if (r0_b - 1, c1_b) in room_cells else r0_b
                bottom_row = r1_b + 1 if (r1_b + 1, c1_b) in room_cells else r1_b
                
                miolo_livre_direito = True
                for r in range(r0_b, r1_b + 1):
                    if not _eh_faixa_livre(ws, env_cells, r, r, c1_b + 1, c1_b + 1) or _tem_parede_laranja_entre(ws, r, c1_b, r, c1_b + 1):
                        miolo_livre_direito = False
                        break
                        
                if miolo_livre_direito:
                    for r in range(top_row, bottom_row + 1):
                        target_cell = (r, c1_b + 1)
                        if not _tem_parede_laranja_entre(ws, r, c1_b, r, c1_b + 1):
                            if _eh_faixa_livre(ws, env_cells, r, r, c1_b + 1, c1_b + 1):
                                room_cells.add(target_cell)

        corridor_columns = {c for r, c in (room_cells - target_bench_cells)}
        for col in sorted(list(corridor_columns)):
            rows_in_col = sorted(list({r for r, c in room_cells if c == col}))
            if len(rows_in_col) > 1:
                for i in range(len(rows_in_col) - 1):
                    r_start = rows_in_col[i]
                    r_end = rows_in_col[i + 1]
                    if r_end - r_start > 1:
                        gap_rows = range(r_start + 1, r_end)
                        all_gap_free = True
                        for gr in gap_rows:
                            target_cell = (gr, col)
                            if _tem_parede_laranja_entre(ws, gr - 1, col, gr, col) or not _eh_faixa_livre(ws, env_cells, gr, gr, col, col):
                                all_gap_free = False
                                break
                        if all_gap_free:
                            for gr in gap_rows:
                                room_cells.add((gr, col))

        room_cells = _absorver_gaps_estreitos(ws, env_cells, room_cells)

        # Verificação de acessibilidade: todas as mesas devem ter caminho até a saída
        acessivel, mesas_isoladas = _verificar_acessibilidade_mesas(ws, env_cells, room_cells, target_bench_cells)
        resultado["mesas_sem_saida"] = mesas_isoladas
        
        if not acessivel and mesas_isoladas:
            room_cells_antes = set(room_cells)
            room_cells = _expandir_corredor_para_acessibilidade(
                ws, env_cells, room_cells, target_bench_cells, mesas_isoladas
            )
            resultado["relocated_cells"] = room_cells - room_cells_antes

        # CORREÇÃO 3: Unifica o contorno geométrico das salas apenas no final para evitar 'dentes na parede'
        if room_cells_override is not None:
            room_cells = room_cells | room_cells_override

    corridor_cells = room_cells - target_bench_cells
    resultado["room_cells"] = room_cells
    resultado["corridor_cells"] = corridor_cells
    
    cell_ct_coord = None
    if corridor_cells and not reconstruir_sala:
        max_r = max(r for r, c in room_cells)
        bottom_corridors = [cell for cell in corridor_cells if cell[0] == max_r]
        if not bottom_corridors:
            max_r_corridors = max(cell[0] for cell in corridor_cells)
            bottom_corridors = [cell for cell in corridor_cells if cell[0] == max_r_corridors]
        
        min_c_tables = min(c for r, c in target_bench_cells)
        min_c_room = min(c for r, c in room_cells)
        if min_c_room < min_c_tables:
            bottom_corridors.sort(key=lambda x: x[1])
        else:
            bottom_corridors.sort(key=lambda x: x[1], reverse=True)
            
        bottom_corridors = [cell for cell in bottom_corridors if not _eh_celula_mesclada(ws, cell[0], cell[1])]
        if bottom_corridors:
            cell_ct_coord = bottom_corridors[0]
            cell_ct = ws.cell(row=cell_ct_coord[0], column=cell_ct_coord[1])
            
            fill_ct = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
            cell_ct.value = "CT"
            cell_ct.fill = fill_ct
            cell_ct.font = Font(color="FFFFFF", bold=True, size=8)
    
    resultado["ct_cell"] = cell_ct_coord

    if reconstruir_sala:
        for r, c in room_cells:
            # Se a célula está dentro da sala física, mas não é uma das mesas planejadas para ela:
            if (r, c) not in target_bench_cells and (r, c) != cell_ct_coord:
                if not _eh_pilar_ou_coluna(ws, r, c) and not _eh_celula_mesclada(ws, r, c):
                    cell = ws.cell(row=r, column=c)
                    cell.value = ""                         # <--- Esvazia o texto/nome completamente
                    cell.fill = PatternFill(fill_type=None) # <--- Remove cores residuais
                    cell.border = Border()                  # <--- Limpa as bordas internas

    for r, c in room_cells:
        if (r - 1, c) not in room_cells:
            _aplicar_borda_espelhada(ws, r, c, 'top', side_style)
        if (r + 1, c) not in room_cells:
            _aplicar_borda_espelhada(ws, r, c, 'bottom', side_style)
        if (r, c - 1) not in room_cells:
            _aplicar_borda_espelhada(ws, r, c, 'left', side_style)
        if (r, c + 1) not in room_cells:
            _aplicar_borda_espelhada(ws, r, c, 'right', side_style)

    # Validação final
    if not reconstruir_sala:
        acessivel_final, mesas_isoladas_final = _verificar_acessibilidade_mesas(
            ws, env_cells, room_cells, target_bench_cells
        )
        resultado["mesas_sem_saida"] = mesas_isoladas_final
        
        if mesas_isoladas_final:
            resultado["motivo"] = f"{len(mesas_isoladas_final)} mesas sem acesso à saída"
        elif cell_ct_coord is None:
            resultado["motivo"] = "não foi possível posicionar CT"
        else:
            resultado["valido"] = True
    else:
        resultado["valido"] = True
    
    return resultado

def _copiar_lado(side_obj):
    if not side_obj or not side_obj.style or side_obj.style == 'none':
        return None
    return Side(border_style=side_obj.style, color=side_obj.color)

def _aplicar_borda_celula_unica(ws, r, c, side, side_style):
    if _eh_celula_mesclada(ws, r, c):
        return
    cell = ws.cell(row=r, column=c)
    b = cell.border
    top = _copiar_lado(b.top)
    bottom = _copiar_lado(b.bottom)
    left = _copiar_lado(b.left)
    right = _copiar_lado(b.right)
    
    if side == 'top': top = side_style
    elif side == 'bottom': bottom = side_style
    elif side == 'left': left = side_style
    elif side == 'right': right = side_style
    
    cell.border = Border(top=top, bottom=bottom, left=left, right=right)

def _aplicar_borda_espelhada(ws, r, c, side, side_style):
    _aplicar_borda_celula_unica(ws, r, c, side, side_style)
    if side == 'bottom' and r + 1 <= ws.max_row:
        _aplicar_borda_celula_unica(ws, r + 1, c, 'top', side_style)
    elif side == 'top' and r - 1 >= 1:
        _aplicar_borda_celula_unica(ws, r - 1, c, 'bottom', side_style)
    elif side == 'right' and c + 1 <= ws.max_column:
        _aplicar_borda_celula_unica(ws, r, c + 1, 'left', side_style)
    elif side == 'left' and c - 1 >= 1:
        _aplicar_borda_celula_unica(ws, r, c - 1, 'right', side_style)

# ══════════════════════════════════════════════════════════════════════════
# Função de Seleção e Execução de Teste Manual
# ══════════════════════════════════════════════════════════════════════════

def _selecionar_mesas_contiguas(env_cells: Set[Tuple[int, int]], ws, target_qty: int, cliente_atual: str = None) -> Set[Tuple[int, int]]:
    """
    Seleciona mesas contíguas para formar um ambiente.
    
    O AmbienteBuilder é AGNÓSTICO quanto ao conteúdo das células.
    Ele seleciona qualquer célula que seja fisicamente uma mesa,
    independente de qual cliente está nela atualmente.
    """
    all_desks = set()
    for r, c in env_cells:
        cell = ws.cell(row=r, column=c)
        if _eh_celula_de_mesa_local(cell):
            all_desks.add((r, c))
            
    if not all_desks:
        return set()
        
    from BlockMapper import flood_fill
    benches = flood_fill(all_desks)
    n_benches = len(benches)
    if n_benches == 0:
        return set()
        
    parent = list(range(n_benches))
    
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
        
    def union(i, j):
        root_i = find(i)
        root_j = find(j)
        if root_i != root_j:
            parent[root_i] = root_j
            
    bboxes = []
    for b in benches:
        rs = [r for r, c in b]
        cs = [c for r, c in b]
        bboxes.append((min(rs), max(rs), min(cs), max(cs)))

    for i in range(n_benches):
        ri0, ri1, ci0, ci1 = bboxes[i]
        for j in range(i + 1, n_benches):
            rj0, rj1, cj0, cj1 = bboxes[j]
            dr = max(0, ri0 - rj1, rj0 - ri1)
            dc = max(0, ci0 - cj1, cj0 - ci1)
            if dr + dc > 4:
                continue
            min_dist = min(
                abs(r1 - r2) + abs(c1 - c2)
                for r1, c1 in benches[i]
                for r2, c2 in benches[j]
            )
            if min_dist <= 4:
                union(i, j)
                
    from collections import defaultdict
    clusters_map = defaultdict(list)
    for i in range(n_benches):
        clusters_map[find(i)].append(benches[i])
        
    clusters = list(clusters_map.values())
    
    adjusted_target_qty = target_qty
    possui_bancada_dupla = False
    for cl in clusters:
        for b in cl:
            b_cols = {c for r, c in b}
            b_rows = {r for r, c in b}
            if len(b_cols) == 2 or len(b_rows) == 2:
                possui_bancada_dupla = True
                break
        if possui_bancada_dupla:
            break
            
    if possui_bancada_dupla and adjusted_target_qty % 2 != 0:
        adjusted_target_qty += 1

    valid_clusters = []
    for cl in clusters:
        cap = sum(len(b) for b in cl)
        if cap >= adjusted_target_qty:
            valid_clusters.append((cl, cap))
            
    if valid_clusters:
        valid_clusters.sort(key=lambda x: x[1])
        selected_benches = valid_clusters[0][0]
    else:
        clusters_with_cap = []
        for cl in clusters:
            cap = sum(len(b) for b in cl)
            clusters_with_cap.append((cl, cap))
        clusters_with_cap.sort(key=lambda x: -x[1])
        
        selected_benches = []
        for cl, cap in clusters_with_cap:
            selected_benches.extend(cl)
            
    selected_set = set()
    selected_list = []
    
    max_bench_cap = max(len(b) for b in selected_benches) if selected_benches else 0
    
    if adjusted_target_qty > max_bench_cap:
        selected_benches_sorted = sorted(selected_benches, key=lambda b: (min(r for r, c in b), min(c for r, c in b)))
        
        remaining_qty = adjusted_target_qty
        for b in selected_benches_sorted:
            if remaining_qty <= 0:
                break
                
            b_cols = {c for r, c in b}
            b_rows = {r for r, c in b}
            if len(b_rows) >= len(b_cols):
                sorted_cells = sorted(list(b), key=lambda x: (x[0], x[1]))
            else:
                sorted_cells = sorted(list(b), key=lambda x: (x[1], x[0]))
                
            take_qty = min(len(sorted_cells), remaining_qty)
            for cell in sorted_cells[:take_qty]:
                if cell not in selected_set:
                    selected_set.add(cell)
                    selected_list.append(cell)
            remaining_qty -= take_qty
            
    else:
        vertical_benches = []
        horizontal_benches = []
        for b in selected_benches:
            b_rows = {r for r, c in b}
            b_cols = {c for r, c in b}
            if len(b_rows) >= len(b_cols):
                vertical_benches.append(b)
            else:
                horizontal_benches.append(b)
                
        is_mostly_vertical = len(vertical_benches) >= len(horizontal_benches)
        
        if is_mostly_vertical and vertical_benches:
            vertical_benches.sort(key=lambda b: (min(r for r, c in b), min(c for r, c in b)))
            groups = []
            for b in vertical_benches:
                br_y_0 = min(r for r, c in b)
                placed = False
                for grp in groups:
                    if abs(min(r for r, c in grp[0]) - br_y_0) <= 2:
                        grp.append(b)
                        placed = True
                        break
                if not placed:
                    groups.append([b])
            
            valid_groups = [grp for grp in groups if sum(len(b) for b in grp) >= adjusted_target_qty]
            if valid_groups:
                valid_groups.sort(key=lambda grp: min(min(r for r, c in b) for b in grp))
                benches_to_use = valid_groups[0]
            else:
                benches_to_use = vertical_benches
                
            max_rows = max(len({r for r, c in b}) for b in benches_to_use)
            
            best_N, best_H = None, None
            candidates = []
            
            for N in range(1, len(benches_to_use) + 1):
                for H in range(1, max_rows + 1):
                    cap = 0
                    for b in benches_to_use[:N]:
                        b_rows_sorted = sorted(list({r for r, c in b}))
                        allowed_rows = set(b_rows_sorted[:H])
                        cap += sum(1 for r, c in b if r in allowed_rows)
                    
                    if cap >= adjusted_target_qty:
                        excess = cap - adjusted_target_qty
                        corridor_height = max_rows - H
                        candidates.append((cap, N, H, excess, corridor_height))
            
            if candidates:
                candidates.sort(key=lambda x: (x[3], x[1], -x[4]))
                best_N = candidates[0][1]
                best_H = candidates[0][2]
            else:
                deficit_candidates = []
                for N in range(1, len(benches_to_use) + 1):
                    for H in range(1, max_rows + 1):
                        cap = 0
                        for b in benches_to_use[:N]:
                            b_rows_sorted = sorted(list({r for r, c in b}))
                            allowed_rows = set(b_rows_sorted[:H])
                            cap += sum(1 for r, c in b if r in allowed_rows)
                        corridor_height = max_rows - H
                        deficit_candidates.append((cap, N, H, corridor_height))
                
                deficit_candidates.sort(key=lambda x: (-x[0], x[1], -x[3]))
                best_N = deficit_candidates[0][1]
                best_H = deficit_candidates[0][2]
            
            if best_N is not None and best_H is not None:
                remaining_qty = adjusted_target_qty
                for b in benches_to_use[:best_N]:
                    b_cols = {c for r, c in b}
                    physical_cols_count = len(b_cols)
                    
                    if physical_cols_count == 2 and remaining_qty % 2 != 0:
                        remaining_qty += 1
                        
                    b_rows_sorted = sorted(list({r for r, c in b}))
                    allowed_rows = set(b_rows_sorted[:best_H])
                    allowed_cells = [cell for cell in b if cell[0] in allowed_rows]
                    allowed_cells.sort(key=lambda x: (x[0], x[1]))
                    
                    take_qty = min(len(allowed_cells), remaining_qty)
                    for cell in allowed_cells[:take_qty]:
                        if cell not in selected_set:
                            selected_set.add(cell)
                            selected_list.append(cell)
                    remaining_qty -= take_qty
                    if remaining_qty <= 0:
                        break
                        
        elif not is_mostly_vertical and horizontal_benches:
            horizontal_benches.sort(key=lambda b: (min(c for r, c in b), min(r for r, c in b)))
            groups = []
            for b in horizontal_benches:
                bc_x_0 = min(c for r, c in b)
                placed = False
                for grp in groups:
                    if abs(min(c for grp_b in grp for r, c in grp_b) - bc_x_0) <= 2:
                        grp.append(b)
                        placed = True
                        break
                if not placed:
                    groups.append([b])
            
            valid_groups = [grp for grp in groups if sum(len(b) for b in grp) >= adjusted_target_qty]
            if valid_groups:
                valid_groups.sort(key=lambda grp: min(min(c for r, c in b) for b in grp))
                benches_to_use = valid_groups[0]
            else:
                benches_to_use = horizontal_benches
                
            max_cols = max(len({c for r, c in b}) for b in benches_to_use)
            
            best_N, best_W = None, None
            candidates = []
            
            for N in range(1, len(benches_to_use) + 1):
                for W in range(1, max_cols + 1):
                    cap = 0
                    for b in benches_to_use[:N]:
                        b_cols_sorted = sorted(list({c for r, c in b}))
                        allowed_cols = set(b_cols_sorted[:W])
                        cap += sum(1 for r, c in b if c in allowed_cols)
                    
                    if cap >= adjusted_target_qty:
                        excess = cap - adjusted_target_qty
                        corridor_width = max_cols - W
                        candidates.append((cap, N, W, excess, corridor_width))
                        
            if candidates:
                candidates.sort(key=lambda x: (x[3], x[1], -x[4]))
                best_N = candidates[0][1]
                best_W = candidates[0][2]
            else:
                deficit_candidates = []
                for N in range(1, len(benches_to_use) + 1):
                    for W in range(1, max_cols + 1):
                        cap = 0
                        for b in benches_to_use[:N]:
                            b_cols_sorted = sorted(list({c for r, c in b}))
                            allowed_cols = set(b_cols_sorted[:W])
                            cap += sum(1 for r, c in b if c in allowed_cols)
                        corridor_width = max_cols - W
                        deficit_candidates.append((cap, N, W, corridor_width))
                
                deficit_candidates.sort(key=lambda x: (-x[0], x[1], -x[3]))
                best_N = deficit_candidates[0][1]
                best_W = deficit_candidates[0][2]
                            
            if best_N is not None and best_W is not None:
                remaining_qty = adjusted_target_qty
                for b in benches_to_use[:best_N]:
                    b_rows = {r for r, c in b}
                    physical_rows_count = len(b_rows)
                    
                    if physical_rows_count == 2 and remaining_qty % 2 != 0:
                        remaining_qty += 1
                        
                    b_cols_sorted = sorted(list({c for r, c in b}))
                    allowed_cols = set(b_cols_sorted[:best_W])
                    allowed_cells = [cell for cell in b if cell[1] in allowed_cols]
                    allowed_cells.sort(key=lambda x: (x[1], x[0]))
                    
                    take_qty = min(len(allowed_cells), remaining_qty)
                    for cell in allowed_cells[:take_qty]:
                        if cell not in selected_set:
                            selected_set.add(cell)
                            selected_list.append(cell)
                    remaining_qty -= take_qty
                    if remaining_qty <= 0:
                        break

    if not selected_list:
        benches.sort(key=lambda b: min(b))
        remaining_qty = adjusted_target_qty
        for bench in benches:
            if remaining_qty <= 0:
                break
            bench_set = set(bench)
            cols = {c for r, c in bench_set}
            rows = {r for r, c in bench_set}
            if len(rows) >= len(cols):
                sorted_bench = sorted(list(bench_set), key=lambda x: (x[0], x[1]))
            else:
                sorted_bench = sorted(list(bench_set), key=lambda x: (x[1], x[0]))
                
            if len(bench_set) <= remaining_qty:
                for cell in sorted_bench:
                    if cell not in selected_set:
                        selected_set.add(cell)
                        selected_list.append(cell)
                remaining_qty -= len(bench_set)
            else:
                for cell in sorted_bench[:remaining_qty]:
                    if cell not in selected_set:
                        selected_set.add(cell)
                        selected_list.append(cell)
                remaining_qty = 0
                break

    if len(selected_list) > target_qty:
        selected_list = selected_list[:target_qty]
    
    # RETANGULARIZAÇÃO: Tenta completar linhas/colunas parciais para evitar "dentes"
    selected_set = set(selected_list)
    if selected_set:
        # Identifica bancadas que têm células selecionadas
        from BlockMapper import flood_fill
        selected_benches = flood_fill(selected_set)
        
        for bench in selected_benches:
            bench_set = set(bench)
            # Pega todas as células dessa bancada que estão em all_desks
            full_bench = bench_set & all_desks
            
            # Se a bancada tem 2 colunas (bancada dupla vertical)
            bench_cols = {c for r, c in full_bench}
            bench_rows = {r for r, c in full_bench}
            
            if len(bench_cols) == 2:
                # Verifica se há linhas "incompletas" (só 1 célula de 2)
                for row in bench_rows:
                    cells_in_row = [(r, c) for r, c in full_bench if r == row]
                    selected_in_row = [(r, c) for r, c in cells_in_row if (r, c) in selected_set]
                    
                    # Se tem 1 selecionada mas a bancada tem 2 nessa linha
                    if len(selected_in_row) == 1 and len(cells_in_row) == 2:
                        # Adiciona a célula faltante para completar a linha
                        missing = [c for c in cells_in_row if c not in selected_set]
                        if missing:
                            selected_set.add(missing[0])
            
            elif len(bench_rows) == 2:
                # Bancada dupla horizontal - completa colunas
                for col in bench_cols:
                    cells_in_col = [(r, c) for r, c in full_bench if c == col]
                    selected_in_col = [(r, c) for r, c in cells_in_col if (r, c) in selected_set]
                    
                    if len(selected_in_col) == 1 and len(cells_in_col) == 2:
                        missing = [c for c in cells_in_col if c not in selected_set]
                        if missing:
                            selected_set.add(missing[0])
        
        selected_list = list(selected_set)
                
    return set(selected_list)


def gerar_alternativas_mesas(
    env_cells: Set[Tuple[int, int]], 
    ws, 
    target_qty: int, 
    max_alternativas: int = 5
) -> List[Set[Tuple[int, int]]]:
    """
    Gera múltiplas alternativas de seleção de mesas, priorizando regiões compactas e retangulares.
    
    REGRA CRÍTICA: Só seleciona mesas de bancadas que são FISICAMENTE CONTÍGUAS (adjacentes),
    sem células de outros clientes entre elas.
    """
    from BlockMapper import flood_fill
    
    # Coleta todas as mesas válidas (excluindo células de outros clientes)
    all_desks = set()
    for r, c in env_cells:
        cell = ws.cell(row=r, column=c)
        val = str(cell.value).strip().upper() if cell.value else ""
        if val.startswith("N_"):
            continue
        if _eh_celula_de_mesa_local(cell):
            all_desks.add((r, c))
    
    if not all_desks:
        return []
    
    # Agrupa mesas em bancadas contíguas
    benches = flood_fill(all_desks)
    if not benches:
        return []
    
    alternativas = []
    
    # Ordena bancadas por tamanho (maior primeiro)
    benches_sorted = sorted(benches, key=lambda b: len(b), reverse=True)
    
    # ESTRATÉGIA 1: Tenta usar UMA ÚNICA bancada grande
    for bench in benches_sorted:
        bench_set = set(bench)
        if len(bench_set) >= target_qty:
            # Gera alternativa pegando do topo da bancada
            b_rows = sorted(list({r for r, c in bench_set}))
            b_cols = sorted(list({c for r, c in bench_set}))
            is_vertical = len(b_rows) >= len(b_cols)
            
            selected = set()
            if is_vertical:
                # Pega linhas do topo para baixo
                for row in b_rows:
                    cells_in_row = sorted([(r, c) for r, c in bench_set if r == row], key=lambda x: x[1])
                    for cell in cells_in_row:
                        if len(selected) >= target_qty:
                            break
                        selected.add(cell)
                    if len(selected) >= target_qty:
                        break
            else:
                # Pega colunas da esquerda para direita
                for col in b_cols:
                    cells_in_col = sorted([(r, c) for r, c in bench_set if c == col], key=lambda x: x[0])
                    for cell in cells_in_col:
                        if len(selected) >= target_qty:
                            break
                        selected.add(cell)
                    if len(selected) >= target_qty:
                        break
            
            if len(selected) >= target_qty:
                rs = [r for r, c in selected]
                cs = [c for r, c in selected]
                bbox_area = (max(rs) - min(rs) + 1) * (max(cs) - min(cs) + 1)
                alternativas.append((bbox_area, len(selected), selected))
    
    # ESTRATÉGIA 2: Combina APENAS bancadas adjacentes (distância 1, sem obstáculos entre elas)
    if not alternativas or all(len(alt[2]) < target_qty for alt in alternativas):
        n_benches = len(benches)
        
        # Para cada bancada, encontra bancadas DIRETAMENTE adjacentes
        adjacency = {i: set() for i in range(n_benches)}
        
        for i in range(n_benches):
            for j in range(i + 1, n_benches):
                # Verifica se há célula de bench_i adjacente a célula de bench_j
                adjacent = False
                for r1, c1 in benches[i]:
                    if adjacent:
                        break
                    for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                        nr, nc = r1 + dr, c1 + dc
                        if (nr, nc) in set(benches[j]):
                            adjacent = True
                            break
                        # Também aceita se há apenas 1 célula de corredor livre entre elas
                        if (nr, nc) in env_cells and (nr, nc) not in all_desks:
                            cell_between = ws.cell(row=nr, column=nc)
                            val_between = str(cell_between.value).strip().upper() if cell_between.value else ""
                            # Célula entre deve ser corredor vazio, não outro cliente
                            if not val_between.startswith("N_") and not _eh_celula_de_mesa_local(cell_between):
                                for dr2, dc2 in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                                    if (nr + dr2, nc + dc2) in set(benches[j]):
                                        adjacent = True
                                        break
                        if adjacent:
                            break
                
                if adjacent:
                    adjacency[i].add(j)
                    adjacency[j].add(i)
        
        # BFS para encontrar grupos de bancadas conectadas
        visited = set()
        grupos_conectados = []
        
        for start in range(n_benches):
            if start in visited:
                continue
            grupo = []
            fila = [start]
            while fila:
                curr = fila.pop(0)
                if curr in visited:
                    continue
                visited.add(curr)
                grupo.append(curr)
                for adj in adjacency[curr]:
                    if adj not in visited:
                        fila.append(adj)
            if grupo:
                grupos_conectados.append(grupo)
        
        # Para cada grupo conectado, gera alternativa
        for grupo in grupos_conectados:
            grupo_benches = [benches[i] for i in grupo]
            total_cap = sum(len(b) for b in grupo_benches)
            
            if total_cap < target_qty:
                continue
            
            # Ordena bancadas do grupo por posição (top-left first)
            grupo_benches_sorted = sorted(grupo_benches, key=lambda b: (min(r for r, c in b), min(c for r, c in b)))
            
            selected = set()
            for bench in grupo_benches_sorted:
                bench_list = sorted(list(bench), key=lambda x: (x[0], x[1]))
                for cell in bench_list:
                    if len(selected) >= target_qty:
                        break
                    selected.add(cell)
                if len(selected) >= target_qty:
                    break
            
            if len(selected) >= target_qty:
                rs = [r for r, c in selected]
                cs = [c for r, c in selected]
                bbox_area = (max(rs) - min(rs) + 1) * (max(cs) - min(cs) + 1)
                alternativas.append((bbox_area, len(selected), selected))
    
    # Fallback: usa a seleção original
    if not alternativas:
        original = _selecionar_mesas_contiguas(env_cells, ws, target_qty)
        if original:
            rs = [r for r, c in original]
            cs = [c for r, c in original]
            bbox_area = (max(rs) - min(rs) + 1) * (max(cs) - min(cs) + 1) if original else float('inf')
            alternativas.append((bbox_area, len(original), original))
    
    # Ordena por compacidade (menor bbox_area primeiro)
    alternativas.sort(key=lambda x: (x[0], -x[1]))
    
    # Remove duplicatas
    seen = set()
    resultado = []
    for _, _, selected in alternativas:
        key = frozenset(selected)
        if key not in seen:
            seen.add(key)
            resultado.append(selected)
            if len(resultado) >= max_alternativas:
                break
    
    return resultado


def testar_criacao_sala_manual(
    file_path: str,
    sheet_name: str,
    bloco_id: str,
    ambiente_letra: str,
    quantidade_mesas: int,
    quantidade_mesas_sala: int = None,
    output_path: str = "planta_teste_sala.xlsx",
    border_style: str = "medium",
    border_color: str = "FF9900"
):
    print(f"\n[INICIANDO TESTE MANUAL] Bloco: {bloco_id} | Ambiente: {ambiente_letra} | PAs Solicitadas: {quantidade_mesas}")
    
    if not os.path.exists(file_path):
        print(f"Erro: Arquivo '{file_path}' não encontrado.")
        return

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    macro_blocks = scan_orange_context(file_path, sheet_name)
    raw_env_cells = get_env_cells(bloco_id, ambiente_letra, macro_blocks)

    if not raw_env_cells:
        print(f"Erro: Não foi possível localizar células para o ambiente '{bloco_id}-{ambiente_letra}'.")
        return

    env_cells = set(raw_env_cells) | _celulas_contorno_do_ambiente(ws, set(raw_env_cells))

    allocated_sala = set()
    room_cells_override = None
    
    if quantidade_mesas_sala is not None:
        print(f"🛠️ Planejando Sala Fechada de {quantidade_mesas_sala} mesas...")
        allocated_sala, room_cells_override = _gerar_layout_sala_estruturado(ws, env_cells, quantidade_mesas_sala)

    available_env_cells = env_cells - (room_cells_override if room_cells_override else set())
    allocated_ambiente = _selecionar_mesas_contiguas(available_env_cells, ws, quantidade_mesas)
    
    if not allocated_ambiente:
        print("Erro: Nenhuma mesa utilizável encontrada nesse ambiente para a equipe principal.")
        return

    allocated_total = allocated_ambiente | (room_cells_override if room_cells_override else set())
    separar_ambiente_e_desenhar_divisorias(
        ws=ws, 
        env_cells=env_cells, 
        allocated_cells=allocated_total,
        border_style=border_style, 
        border_color=border_color,
        reconstruir_sala=False,
        room_cells_override=room_cells_override
    )

    if quantidade_mesas_sala is not None:
        separar_ambiente_e_desenhar_divisorias(
            ws=ws, 
            env_cells=env_cells, 
            allocated_cells=allocated_sala,
            border_style=border_style, 
            border_color=border_color,
            reconstruir_sala=True,
            room_cells_override=room_cells_override
        )

    wb.save(output_path)
    print(f"✓ Teste concluído com sucesso! Resultado salvo em: '{output_path}'")

if __name__ == "__main__":
    testar_criacao_sala_manual(
        file_path="planta.xlsx",
        sheet_name="JPIII",
        bloco_id="Bloco_7",
        ambiente_letra="A",         
        quantidade_mesas=124, 
        quantidade_mesas_sala=4,
        output_path="planta_teste_sala.xlsx"
    )