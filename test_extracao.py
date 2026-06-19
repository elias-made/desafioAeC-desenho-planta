import openpyxl
from ScannerPremissas import scan_orange_context
from main import build_blocos_info, SHEET_NAME, FORBIDDEN_PATTERNS
from BlockMapper import scan_plant

def testar_extracao():
    print("🔍 Iniciando extração de blocos e sub-ambientes...")
    
    # 1. Carrega dados brutos da planilha
    wb = openpyxl.load_workbook('planta.xlsx', data_only=True)
    ws = wb[SHEET_NAME]
    plant_data = scan_plant(ws, FORBIDDEN_PATTERNS)

    # 2. Executa a varredura das molduras e divisões internas
    macro_blocks = scan_orange_context('planta.xlsx', SHEET_NAME)

    print("\n" + "="*60)
    print("📊 ESTRUTURA DOS BLOCOS MAPEADOS NO VEÍCULO (RAW DATA):")
    print("="*60)
    for block in macro_blocks:
        print(f"ID do Bloco: {block['id']}")
        print(f"  Caixa Delimitadora: {block['bounding_box']}")
        print(f"  Células no Interior Caminhável: {len(block['interior_cells'])} PAs")
        print(f"  Anotações Próximas: {block['texts']}")
        print("-" * 60)

    # 3. Gera o formato final enviado para a IA
    print("\n" + "="*60)
    print("📝 TEXTO FORMATADO ENVIADO PARA A IA (BLOCOS INFO):")
    print("="*60)
    blocos_info = build_blocos_info(plant_data, ws.max_row, ws.max_column, ws)
    print(blocos_info)

if __name__ == '__main__':
    testar_extracao()