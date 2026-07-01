"""
BlockMapper — funções puras de geometria para plantas de call center.
"""

from collections import deque, defaultdict
from typing import Dict, List, Set, Tuple

Cell = Tuple[int, int]


# ── Primitivas de geometria ─────────────────────────────────────────────

def flood_fill(cells: Set[Cell]) -> List[List[Cell]]:
    """Retorna blocos contíguos (4-vizinhos), ordenados do maior para o menor."""
    visited: Set[Cell] = set()
    blocks: List[List[Cell]] = []
    for seed in sorted(cells):
        if seed in visited:
            continue
        block: List[Cell] = []
        q = deque([seed])
        while q:
            cur = q.popleft()
            if cur in visited or cur not in cells:
                continue
            visited.add(cur)
            block.append(cur)
            r, c = cur
            for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                nb = (r + dr, c + dc)
                if nb not in visited and nb in cells:
                    q.append(nb)
        blocks.append(sorted(block))
    return sorted(blocks, key=len, reverse=True)


def _block_gap(b1: List[Cell], b2: List[Cell]) -> int:
    """Distância mínima de Manhattan entre as bordas de dois blocos."""
    def border(blk):
        s = set(blk)
        return [(r, c) for r, c in blk
                if any((r + dr, c + dc) not in s
                       for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)))]
    best = float('inf')
    for r1, c1 in border(b1):
        for r2, c2 in border(b2):
            d = abs(r1 - r2) + abs(c1 - c2)
            if d < best:
                best = d
            if best <= 1:
                return 1
    return int(best)


def group_zones(blocks: List[List[Cell]], gap: int) -> List[List[Cell]]:
    """Agrupa blocos cujo gap ≤ `gap` em zonas (Union-Find)."""
    n = len(blocks)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for i in range(n):
        for j in range(i + 1, n):
            if find(i) != find(j) and _block_gap(blocks[i], blocks[j]) <= gap:
                parent[find(i)] = find(j)

    groups: Dict[int, List[Cell]] = defaultdict(list)
    for i, blk in enumerate(blocks):
        groups[find(i)].extend(blk)

    return sorted(groups.values(), key=len, reverse=True)


# ── Leitura da planta ────────────────────────────────────────────────────

def scan_plant(ws, forbidden_patterns: Set[str]) -> Dict:
    client_cells: Dict[str, Set[Cell]] = {}
    forbidden: Set[Cell] = set()

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                v_str = ""
            elif isinstance(v, float) and v == int(v):
                v_str = str(int(v))
            else:
                v_str = str(v).strip()
            v_up = v_str.upper()

            if any(pat in v_up for pat in forbidden_patterns):
                forbidden.add((r, c))
            else:
                client_cells.setdefault(v_str, set()).add((r, c))

    return {'client_cells': client_cells, 'forbidden': forbidden}


# ── Descrição para o LLM ─────────────────────────────────────────────────

def describe_for_llm(client_value: str, cells: Set[Cell], corridor_gap: int) -> str:
    from openpyxl.utils import get_column_letter as gcl
    blocks = flood_fill(cells)
    zones = group_zones(blocks, gap=corridor_gap)
    lines = [f"Cliente '{client_value}': {len(cells)} PAs em {len(zones)} zona(s)"]
    for i, zone in enumerate(zones):
        r_min = min(r for r, c in zone)
        r_max = max(r for r, c in zone)
        c_min = min(c for r, c in zone)
        c_max = max(c for r, c in zone)
        lines.append(
            f"  Zona {i}: {len(zone):>4} PAs"
            f" — linhas {r_min}-{r_max}, colunas {gcl(c_min)}-{gcl(c_max)}"
        )
    return "\n".join(lines)


# ── Alocação ─────────────────────────────────────────────────────────────

def pick_zone(available: Set[Cell], needed: int,
              forbidden: Set[Cell], corridor_gap: int) -> Tuple[List[Cell], List[Cell]]:
    usable = available - forbidden
    if not usable or needed <= 0:
        return [], list(available)

    zones = group_zones(flood_fill(usable), gap=corridor_gap)
    candidates = [z for z in zones if len(z) >= needed]
    zone = min(candidates, key=len) if candidates else zones[0]

    allocated = sorted(zone)[:needed]
    return allocated, list(available - set(allocated))


def pick_sala(pa_cells: List[Cell], available: Set[Cell],
              sala_size: int, forbidden: Set[Cell],
              corridor_gap: int) -> Tuple[List[Cell], List[Cell]]:
    """Seleciona células contíguas adjacentes ao bloco para criar a sala."""
    if sala_size <= 0:
        return [], list(available)

    pa_set = set(pa_cells)
    usable = available - forbidden
    seen: Set[Cell] = set()
    candidates: List[Cell] = []

    for r, c in pa_cells:
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nb = (r + dr, c + dc)
            if nb not in pa_set and nb in usable and nb not in seen:
                candidates.append(nb)
                seen.add(nb)

    if len(candidates) < sala_size:
        for r, c in pa_cells:
            for nb in usable:
                if nb in seen or nb in pa_set:
                    continue
                if abs(nb[0] - r) + abs(nb[1] - c) <= corridor_gap:
                    candidates.append(nb)
                    seen.add(nb)

    if len(candidates) < sala_size:
        border = sorted(
            p for p in pa_cells
            if any((p[0] + dr, p[1] + dc) not in pa_set
                   for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)))
        )
        for p in border:
            if p not in seen:
                candidates.append(p)
                seen.add(p)

    sala = sorted(candidates)[:sala_size]
    return sala, list(available - set(sala))


# ── Detecção de blocos vazios e catracas ─────────────────────────────────

def find_empty_blocks(client_cells: Dict[str, Set[Cell]], ws_max_row: int, ws_max_col: int, 
                      corridor_gap: int = 3) -> Dict[str, Dict]:
    """Identifica blocos vazios na planilha utilizando o raio de 1 célula de segurança vertical."""
    empty_cells = set()
    for key in ('VAZIO', 'vazio', ''):  # '0' removido e não é mais detectado como espaço vazio
        empty_cells.update(client_cells.get(key, set()))
        
    if not empty_cells:
        return {}
    
    active_cells = set()
    for client, cells in client_cells.items():
        client_up = str(client).upper().strip()
        if client_up not in ('VAZIO', 'SEM POSSIB', 'CT', 'SA', 'SA', 'CW', '##', ''):
            active_cells.update(cells)
    
    clean_empty_cells = set()
    for r, c in empty_cells:
        is_fringe = False
        for offset in (-1, 1):
            if (r + offset, c) in active_cells:
                is_fringe = True
                break
        if not is_fringe:
            clean_empty_cells.add((r, c))
    
    if not clean_empty_cells:
        return {}
    
    blocks = flood_fill(clean_empty_cells)
    zones = group_zones(blocks, gap=corridor_gap)
    
    empty_blocks = {}
    for i, zone in enumerate(zones):
        r_min = min(r for r, c in zone)
        r_max = max(r for r, c in zone)
        c_min = min(c for r, c in zone)
        c_max = max(c for r, c in zone)
        
        empty_blocks[f'Bloco_{i+1}'] = {  # <--- CORRIGIDO: Modificado de "vazio-" para "Bloco_"
            'cells': set(zone),
            'size': len(zone),
            'position': (r_min, r_max, c_min, c_max),
        }
    
    return empty_blocks


def count_catracas_in_zone(client_cells: Dict[str, Set[Cell]], client: str) -> int:
    client_zone = client_cells.get(client, set())
    catraca_count = 0
    
    for catraca_key in ['CT', 'CATRACA']:
        catraca_cells = client_cells.get(catraca_key, set())
        if client_zone and catraca_cells:
            for r, c in catraca_cells:
                for pr, pc in client_zone:
                    if abs(r - pr) <= 2 and abs(c - pc) <= 2:
                        catraca_count += 1
                        break
    return catraca_count


def calculate_catracas_needed(n_pas: int) -> int:
    import math
    return max(1, math.ceil(n_pas / 250))