# main.py

import asyncio
import os
import shutil
import openpyxl
import traceback
import re
from copy import copy
from typing import List, Tuple
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.cell import MergedCell

# Importando dependencias do LayoutEngine
from LayoutEngine import (
    SHEET_NAME, FORBIDDEN_PATTERNS,
    Acao, PropostaMock, clean_json_string, extrair_e_carregar_json,
    load_plant, clone_ws, build_plant_info, build_blocos_info,
    execute_alocacao, write_report, salvar_auditoria,
    normalizar_acoes, validar_inventario, FILL_LIBERADO, FONT_SMALL,
    FILL_NEW_CLIENT, FONT_WHITE
)

# Importando dependências do AmbienteBuilder
from AmbienteBuilder import (
    separar_ambiente_e_desenhar_divisorias,
    _selecionar_mesas_contiguas,
    _gerar_layout_sala_estruturado,
    _celulas_contorno_do_ambiente,
    _eh_celula_de_mesa_local,
    _tem_parede_laranja_entre,
)

# Importando dependências e agentes do Agents.py (Fluxo de 2 Agentes)
from Agents import PosicionadorDeps, OrganizadorDeps, posicionador, organizador
from ScannerPremissas import scan_orange_context, normalize_val
from BlockMapper import scan_plant

load_dotenv()

# ══════════════════════════════════════════════════════════════════════════
# Fluxo de Execução Principal (Unificado de 2 Etapas com Auto-Correção)
# ══════════════════════════════════════════════════════════════════════════

async def main():
    # 1. Carrega as premissas brutas do arquivo de texto
    with open('premissas.txt', encoding='utf-8') as f: 
        premissas_txt = f.read().strip()

    # Diretrizes de nomenclatura inseridas diretamente via prompt
    regras_nomes_sistema = (
        "\n=== DIRETRIZES DE NOMENCLATURA SISTÊMICA (OBRIGATÓRIO) ===\n"
        "Ao criar novas operações sugeridas nas premissas, você DEVE nomeá-las de forma sequencial:\n"
        "  - A primeira operação criada deve ser nomeada rigorosamente como: 'N_1'\n"
        "  - A segunda operação criada deve ser nomeada rigorosamente como: 'N_2'\n"
        "  - A terceira operação criada deve ser nomeada rigorosamente como: 'N_3'\n"
        "E assim por diante. Use sempre este padrão e mantenha a consistência nominal.\n"
    )

    # Extrai o contexto visual das bordas laranjas (necessário para o mapeamento de células permitidas)
    dados_laranjas = scan_orange_context('planta.xlsx', SHEET_NAME)
    
    # Consolida as diretrizes do sistema com o texto de entrada (as anotações visuais vão direto no bloco)
    premissas_completas = f"{premissas_txt}\n{regras_nomes_sistema}"

    # Carrega a planta original e escaneia os dados iniciais
    wb, ws = load_plant()
    
    # --- SNAPSHOT DA PLANILHA ORIGINAL (Restauração Perfeita de Inventário) ---
    original_snapshot = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            original_snapshot[(r, c)] = {
                'value': cell.value,
                'fill': copy(cell.fill) if cell.has_style else None,
                'font': copy(cell.font) if cell.has_style else None
            }

    plant_data = scan_plant(ws, FORBIDDEN_PATTERNS)
    plant_info_str = build_plant_info(plant_data)
    blocos_info_str = build_blocos_info(plant_data, ws.max_row, ws.max_column, ws, file_path='planta.xlsx')
    
    # ── PROTEÇÃO CONTRA USO DE CORREDORES ──
    allowed_cells = set()
    boundary_cells = set()
    for block in dados_laranjas:
        for env in block.get('ambientes', []):
            boundary_cells.update(env['cells'])
            
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (r, c) not in boundary_cells:
                continue  
                
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                v_str = str(v).strip()
                if v_str != "" and v_str.upper() not in ('CT', 'CATRACA', 'SA', 'CW', 'COWORKING', '##'):
                    allowed_cells.add((r, c))
    # ─────────────────────────────────────────────────────────────────────────────

    # ══════════════════════════════════════════════════════════════════════════
    # ETAPA 1: Execução Direta do Posicionador (Sem Loops Desnecessários)
    # ══════════════════════════════════════════════════════════════════════════
    print("1. Posicionador gerando ações primárias (alocação bruta)...")
    
    dest_file = "propostas/proposta_final.xlsx"
    safe_name = "final"
    os.makedirs('propostas', exist_ok=True)
    shutil.copy("planta.xlsx", dest_file)
    new_wb = openpyxl.load_workbook(dest_file)
    new_ws = new_wb[SHEET_NAME]

    pos_deps = PosicionadorDeps(
        plant_info=plant_info_str,
        blocos_info=blocos_info_str,
        premissas=premissas_completas
    )
    
    # Execução única e direta do Posicionador para gerar o rascunho inicial
    res_posicionador = await posicionador.run(
        "Gere a proposta de alocação inicial baseada nas premissas de negócio.", 
        deps=pos_deps
    )
    rascunho_layout_json = clean_json_string(res_posicionador.output)
    pos_data = extrair_e_carregar_json(rascunho_layout_json)
    
    # Executa a alocação bruta na planilha master
    acoes_totais = []
    for ac in pos_data.get("acoes_primarias", []):
        tipo = str(ac.get("tipo", "")).strip().lower()
        if tipo not in ("liberar", "realocar", "posicionar"):
            raise ValueError(
                f"Acao invalida do posicionador: '{tipo}'. "
                "O posicionador so pode liberar ou realocar/posicionar."
            )
        acoes_totais.append(Acao(**ac))
        
    gabarito_dados = pos_data.get("gabarito", {})
    
    # FALLBACK: Se o bloco "gabarito" for omitido, reconstrói a partir das ações planejadas
    if not gabarito_dados and "acoes_primarias" in pos_data:
        reducoes_fallback = {}
        novos_fallback = []
        for ac in pos_data.get("acoes_primarias", []):
            tipo = ac.get("tipo", "").strip().lower()
            cli = ac.get("cliente", "").strip()
            qtd = ac.get("quantidade", 0)
            if tipo == "liberar":
                reducoes_fallback[cli] = reducoes_fallback.get(cli, 0) + qtd
            elif tipo in ("realocar", "posicionar") and cli.upper().startswith("N_"):
                existente = next((nc for nc in novos_fallback if nc["nome"] == cli.upper()), None)
                if existente:
                    existente["PAs"] += qtd
                else:
                    novos_fallback.append({"nome": cli.upper(), "PAs": qtd})
        gabarito_dados = {"reducoes": reducoes_fallback, "novos_clientes": novos_fallback}
        
    novos_clientes_brutos = gabarito_dados.get("novos_clientes", [])
    novos_clientes_extraidos = []
    for nc in novos_clientes_brutos:
        if isinstance(nc, dict):
            nome_orig = nc.get("nome", "")
            pas = nc.get("PAs", 0)
        else:
            nome_orig = str(nc)
            pas = 0
        nome_norm = nome_orig.upper().strip().replace(" ", "_")
        novos_clientes_extraidos.append({
            "nome": nome_norm,
            "PAs": pas
        })
        
    parametros_premissas = {
        "reducoes": {normalize_val(k): v for k, v in gabarito_dados.get("reducoes", {}).items()},
        "novos_clientes": novos_clientes_extraidos
    }
        
    # Normaliza as ações primárias antes de executar
    acoes_totais = normalizar_acoes(acoes_totais, parametros_premissas["novos_clientes"])
        
    criar_ambientes_solicitados = pos_data.get("criar_ambientes", [])
    clientes_criados_por_parede = {
        normalize_val(amb.get("cliente_destinado"))
        for amb in criar_ambientes_solicitados
        if amb.get("cliente_destinado")
    }
    acoes_pre_ambiente_builder = [
        acao for acao in acoes_totais
        if not (
            acao.tipo.lower().strip() in ("realocar", "posicionar")
            and normalize_val(acao.cliente) in clientes_criados_por_parede
        )
    ]
    proposta_inicial = PropostaMock(
        nome=pos_data.get("nome", "Alocacao Bruta"),
        custo_obras="Baixo",
        acoes=acoes_pre_ambiente_builder
    )
    
    # Executa primeiro apenas o que precisa existir antes das obras fisicas.
    # Os novos clientes com ambiente fechado sao posicionados pelo AmbienteBuilder.
    log_acumulado, _ = execute_alocacao(new_ws, proposta_inicial, plant_data, allowed_cells, file_path='planta.xlsx')
    
    # SALVA A PLANILHA APÓS execute_alocacao - CRÍTICO!
    # Isso garante que dest_file reflita as modificações em memória
    new_wb.save(dest_file)
    
    # === INTEGRAÇÃO DO AMBIENTEBUILDER ===
    ambientes_criados_info = []
    salas_internas_cells = set()
    ambientes_fisicos_falhos = set()
    ambientes_ocupados_cells = set()
    origem_disponivel_por_chave = {}
    if criar_ambientes_solicitados:
        print("[AMBIENTE BUILDER] Planejando criacoes fisicas em lote por origem.")

        def _resolver_origem_cells(bloco_id, ambiente_letra, macro_blocks):
            origem_cells = set()
            ambiente_resolvido = ambiente_letra
            block_match = re.search(r'\d+', str(bloco_id))
            if block_match:
                block_idx = int(block_match.group())
                if block_idx <= len(macro_blocks):
                    block = macro_blocks[block_idx - 1]
                    envs = block.get('ambientes', [])
                    target_letter = str(ambiente_letra).upper() if ambiente_letra else ""
                    env_match = next((env for env in envs if env.get('id', '').upper() == target_letter), None)
                    if env_match is None and envs:
                        env_match = max(envs, key=lambda env: len(env.get('cells', [])))
                    if env_match is not None:
                        ambiente_resolvido = env_match.get('id', ambiente_letra)
                        origem_cells = set(env_match.get('cells', []))
            return origem_cells, ambiente_resolvido

        def _componentes_sem_cruzar_paredes(ws_ref, cells):
            restantes = set(cells)
            componentes = []
            while restantes:
                inicio = min(restantes)
                pilha = [inicio]
                componente = {inicio}
                restantes.remove(inicio)
                while pilha:
                    r, c = pilha.pop()
                    for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                        vizinha = (r + dr, c + dc)
                        if vizinha not in restantes:
                            continue
                        if _tem_parede_laranja_entre(ws_ref, r, c, vizinha[0], vizinha[1]):
                            continue
                        restantes.remove(vizinha)
                        componente.add(vizinha)
                        pilha.append(vizinha)
                componentes.append(componente)
            return componentes

        def _estilo_cliente(cliente_dest):
            fill_to_apply = None
            font_to_apply = None
            for r_search in range(1, new_ws.max_row + 1):
                for c_search in range(1, new_ws.max_column + 1):
                    cell_search = new_ws.cell(row=r_search, column=c_search)
                    if cell_search.value is not None and normalize_val(cell_search.value) == normalize_val(cliente_dest):
                        fill_to_apply = copy(cell_search.fill) if cell_search.has_style else None
                        font_to_apply = copy(cell_search.font) if cell_search.has_style else None
                        break
                if fill_to_apply:
                    break
            if not fill_to_apply:
                paleta_cores = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']
                idx = 0
                for i, nc in enumerate(parametros_premissas["novos_clientes"]):
                    if nc["nome"] == cliente_dest:
                        idx = i
                        break
                cor_hex = paleta_cores[idx % len(paleta_cores)]
                fill_to_apply = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
                font_to_apply = Font(color='FFFFFF', bold=True, size=8)
            return fill_to_apply, font_to_apply

        import ScannerPremissas
        ScannerPremissas._orange_context_cache = {}
        macro_blocks_atual = scan_orange_context(dest_file, SHEET_NAME)

        grupos_por_origem = {}
        for amb in criar_ambientes_solicitados:
            bloco_solicitado = amb.get("bloco")
            letra_solicitada = amb.get("ambiente")
            origem_resolvida, _ = _resolver_origem_cells(
                bloco_solicitado,
                letra_solicitada,
                macro_blocks_atual,
            )

            # Letras de ambientes mudam depois que novas divisorias sao criadas.
            # A identidade estavel da origem e o conjunto de celulas fisicas do
            # escaneamento anterior as obras, nao a letra sugerida pelo agente.
            if origem_resolvida:
                origem_key = (str(bloco_solicitado), frozenset(origem_resolvida))
            else:
                origem_key = (
                    str(bloco_solicitado),
                    "origem_nao_localizada",
                    str(letra_solicitada).upper() if letra_solicitada else "",
                )
            grupos_por_origem.setdefault(origem_key, []).append(amb)

        for origem_key, grupo in grupos_por_origem.items():
            bloco_id = grupo[0].get("bloco")
            ambiente_letra = grupo[0].get("ambiente")
            origem_cells, ambiente_resolvido = _resolver_origem_cells(bloco_id, ambiente_letra, macro_blocks_atual)
            origem_disponivel_por_chave[origem_key] = set(origem_cells)

            if not origem_cells:
                for amb in grupo:
                    cliente_dest = amb.get("cliente_destinado")
                    print(f"   Nao foi possivel localizar origem fisica para {cliente_dest} em {bloco_id}.")
                    ambientes_fisicos_falhos.add(normalize_val(cliente_dest))
                    ambientes_criados_info.append(
                        f"- FALHA ao criar ambiente para '{cliente_dest}' no bloco '{bloco_id}': origem fisica nao localizada."
                    )
                continue

            reservas = []
            reservado_grupo = set()
            ws_planejamento = new_wb.copy_worksheet(new_ws)
            ws_planejamento.title = "__planejamento_ambientes__"

            for amb in grupo:
                qtd_mesas = amb.get("quantidade_mesas")
                cliente_dest = amb.get("cliente_destinado")
                sala_lugares = amb.get("sala_lugares", 0)
                env_cells_base = set(origem_cells) - ambientes_ocupados_cells - reservado_grupo

                if reservado_grupo and env_cells_base:
                    componentes = _componentes_sem_cruzar_paredes(ws_planejamento, env_cells_base)
                    componentes_viaveis = []
                    for componente in componentes:
                        mesas_componente = sum(
                            1
                            for r, c in componente
                            if _eh_celula_de_mesa_local(ws_planejamento.cell(row=r, column=c))
                        )
                        if mesas_componente >= qtd_mesas:
                            componentes_viaveis.append((componente, mesas_componente))

                    if componentes_viaveis:
                        componentes_viaveis.sort(
                            key=lambda item: (
                                item[1],
                                min(c for r, c in item[0]),
                                min(r for r, c in item[0]),
                            )
                        )
                        env_cells_base = set(componentes_viaveis[0][0])
                    else:
                        env_cells_base = set()

                env_cells = env_cells_base | _celulas_contorno_do_ambiente(ws_planejamento, env_cells_base)

                if not env_cells_base:
                    print(f"   Sem area restante para criar {cliente_dest} em {bloco_id}.")
                    ambientes_fisicos_falhos.add(normalize_val(cliente_dest))
                    ambientes_criados_info.append(
                        f"- FALHA ao criar ambiente para '{cliente_dest}' no bloco '{bloco_id}': sem area restante na origem."
                    )
                    continue

                allocated_sala = set()
                room_cells_override = None
                if sala_lugares and sala_lugares > 0:
                    ancora_ambiente = _selecionar_mesas_contiguas(
                        env_cells, ws_planejamento, qtd_mesas,
                        priorizar_sobras=bool(reservado_grupo),
                    )
                    allocated_sala, room_cells_override = _gerar_layout_sala_estruturado(
                        ws_planejamento, env_cells, sala_lugares,
                        anchor_cells=ancora_ambiente,
                    )
                    if room_cells_override:
                        salas_internas_cells.update(room_cells_override)

                room_cells_set = set(room_cells_override) if room_cells_override else set()
                available_env_cells = env_cells - room_cells_set
                allocated_ambiente = _selecionar_mesas_contiguas(
                    available_env_cells,
                    ws_planejamento,
                    qtd_mesas,
                    priorizar_sobras=bool(reservado_grupo),
                )

                if room_cells_set and allocated_ambiente:
                    # A sala pode se ligar ao salao por um pequeno corredor livre.
                    # Exigir contato direto rejeitava layouts validos, como o N_1,
                    # que possuia exatamente uma coluna de circulacao entre ambos.
                    fronteira = {(coord, 0) for coord in room_cells_set}
                    visitadas = set(room_cells_set)
                    sala_encostada = False
                    while fronteira and not sala_encostada:
                        (r_atual, c_atual), distancia = fronteira.pop()
                        if distancia >= 3:
                            continue
                        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
                            vizinha = (r_atual + dr, c_atual + dc)
                            if vizinha not in env_cells or vizinha in visitadas:
                                continue
                            if _tem_parede_laranja_entre(
                                ws_planejamento, r_atual, c_atual, vizinha[0], vizinha[1]
                            ):
                                continue
                            if vizinha in allocated_ambiente:
                                sala_encostada = True
                                break
                            # Somente celulas sem mesa podem servir de corredor.
                            if not _eh_celula_de_mesa_local(
                                ws_planejamento.cell(row=vizinha[0], column=vizinha[1])
                            ):
                                visitadas.add(vizinha)
                                fronteira.add((vizinha, distancia + 1))

                    if not sala_encostada:
                        print(f"   Sala interna de {cliente_dest} ficou desconectada do novo ambiente.")
                        ambientes_fisicos_falhos.add(normalize_val(cliente_dest))
                        ambientes_criados_info.append(
                            f"- FALHA ao criar ambiente para '{cliente_dest}' no bloco '{bloco_id}': "
                            "a sala interna nao ficou conectada ao salao principal."
                        )
                        continue

                if allocated_ambiente and len(allocated_ambiente) < qtd_mesas:
                    msg = (
                        f"   Capacidade insuficiente para {cliente_dest} em {bloco_id}: "
                        f"solicitado {qtd_mesas}, alocavel {len(allocated_ambiente)}. Ambiente nao criado."
                    )
                    print(msg)
                    ambientes_fisicos_falhos.add(normalize_val(cliente_dest))
                    ambientes_criados_info.append(
                        f"- FALHA ao criar ambiente para '{cliente_dest}' no bloco '{bloco_id}': "
                        f"solicitado {qtd_mesas} PAs, mas apenas {len(allocated_ambiente)} cabiam na area restante."
                    )
                    continue

                if not allocated_ambiente:
                    print(f"   Nenhuma mesa encontrada para criar {cliente_dest} em {bloco_id}.")
                    ambientes_fisicos_falhos.add(normalize_val(cliente_dest))
                    ambientes_criados_info.append(
                        f"- FALHA ao criar ambiente para '{cliente_dest}' no bloco '{bloco_id}': nenhuma mesa utilizavel encontrada."
                    )
                    continue

                allocated_total = allocated_ambiente | room_cells_set
                resultado_planejado = separar_ambiente_e_desenhar_divisorias(
                    ws=ws_planejamento,
                    env_cells=env_cells,
                    allocated_cells=allocated_total,
                    reconstruir_sala=False,
                    room_cells_override=room_cells_set
                )
                ambiente_room_cells = resultado_planejado.get("room_cells", set()) or allocated_total

                if room_cells_set:
                    resultado_sala_planejada = separar_ambiente_e_desenhar_divisorias(
                        ws=ws_planejamento,
                        env_cells=env_cells,
                        allocated_cells=allocated_sala,
                        reconstruir_sala=True,
                        room_cells_override=room_cells_set
                    )
                    ambiente_room_cells |= (resultado_sala_planejada.get("room_cells", set()) or room_cells_set)

                # O proximo ambiente pode reutilizar mesas que ficaram fora da
                # alocacao efetiva, mesmo que estejam proximas ao contorno do
                # ambiente anterior. Reservar todo `room_cells` fazia bancadas
                # parcialmente usadas desaparecerem da selecao seguinte.
                reservado_grupo.update(allocated_total)
                reservas.append({
                    "amb": amb,
                    "ambiente_letra": ambiente_resolvido,
                    "env_cells": env_cells,
                    "allocated_ambiente": allocated_ambiente,
                    "allocated_sala": allocated_sala,
                    "room_cells_set": room_cells_set,
                    "reserved_cells": ambiente_room_cells,
                })

            if ws_planejamento in new_wb.worksheets:
                new_wb.remove(ws_planejamento)

            for reserva in reservas:
                amb = reserva["amb"]
                bloco_id = amb.get("bloco")
                cliente_dest = amb.get("cliente_destinado")
                allocated_ambiente = reserva["allocated_ambiente"]
                allocated_sala = reserva["allocated_sala"]
                room_cells_set = reserva["room_cells_set"]
                env_cells = reserva["env_cells"]
                allocated_total = allocated_ambiente | room_cells_set

                resultado_ambiente = separar_ambiente_e_desenhar_divisorias(
                    ws=new_ws,
                    env_cells=env_cells,
                    allocated_cells=allocated_total,
                    reconstruir_sala=False,
                    room_cells_override=room_cells_set
                )
                ambiente_room_cells = resultado_ambiente.get("room_cells", set()) or reserva["reserved_cells"]
                ambientes_ocupados_cells.update(ambiente_room_cells)
                origem_disponivel_por_chave[origem_key].difference_update(ambiente_room_cells)

                if room_cells_set:
                    resultado_sala = separar_ambiente_e_desenhar_divisorias(
                        ws=new_ws,
                        env_cells=env_cells,
                        allocated_cells=allocated_sala,
                        reconstruir_sala=True,
                        room_cells_override=room_cells_set
                    )
                    sala_room_cells = resultado_sala.get("room_cells", set()) or room_cells_set
                    ambientes_ocupados_cells.update(sala_room_cells)
                    origem_disponivel_por_chave[origem_key].difference_update(sala_room_cells)

                fill_to_apply, font_to_apply = _estilo_cliente(cliente_dest)
                for r, c in allocated_ambiente:
                    cell = new_ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell) and cell.value != "CT":
                        cell.value = cliente_dest
                        cell.fill = fill_to_apply
                        cell.font = font_to_apply

                all_allocated_cells = allocated_ambiente | room_cells_set
                for r, c in env_cells:
                    if (r, c) not in all_allocated_cells:
                        cell = new_ws.cell(row=r, column=c)
                        if not isinstance(cell, MergedCell) and cell.value != "CT":
                            snap = original_snapshot.get((r, c))
                            if snap:
                                cell.value = snap['value']
                                if snap['fill']:
                                    cell.fill = snap['fill']
                                if snap['font']:
                                    cell.font = snap['font']

                ambientes_criados_info.append(
                    f"- Criado Novo Ambiente Fechado no Bloco '{bloco_id}', Ambiente '{reserva['ambiente_letra']}' "
                    f"com {len(allocated_ambiente)} PAs operacionais"
                    + (f" + sala fechada de {len(allocated_sala)} lugares" if allocated_sala else "")
                    + f" para o cliente '{cliente_dest}'."
                )
                print(f"   Ambiente fisico em lote criado para {cliente_dest}: {len(allocated_ambiente)} PAs")

            new_wb.save(dest_file)

    # Salva a planilha intermediária (para garantir as alterações físicas das paredes)
    new_wb.save(dest_file)

    # === RECONCILIAÇÃO MATEMÁTICA DE INVENTÁRIO (GARANTE SOMA ZERO) ===
    print("⚖️ Reconciliando inventário global para garantir consistência física...")
    
    # 1. Popular estilos de preenchimento e fontes dos clientes (Busca unificada e robusta)
    client_fills = {}
    client_fonts = {}
    
    # Primeiro busca na original para ter os estilos dos estáveis/reduzidos
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                norm_v = normalize_val(v)
                if norm_v != "" and norm_v not in ('VAZIO', 'CT', 'SA', 'SALA', 'CW', '##'):
                    if cell.fill and cell.fill.patternType == 'solid':
                        client_fills[norm_v] = copy(cell.fill)
                    if cell.font:
                        client_fonts[norm_v] = copy(cell.font)
                        
    # Depois busca na modificada para capturar os estilos gerados pelo LayoutEngine para novos clientes
    for r in range(1, new_ws.max_row + 1):
        for c in range(1, new_ws.max_column + 1):
            cell = new_ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                norm_v = normalize_val(v)
                if norm_v != "" and norm_v not in ('VAZIO', 'CT', 'SA', 'SALA', 'CW', '##'):
                    if cell.fill and cell.fill.patternType == 'solid':
                        client_fills[norm_v] = copy(cell.fill)
                    if cell.font:
                        client_fonts[norm_v] = copy(cell.font)

    paleta_cores = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']
    default_new_fills = {}
    default_new_fonts = {}
    for idx, nc in enumerate(parametros_premissas["novos_clientes"]):
        nome = nc["nome"]
        if nome in ambientes_fisicos_falhos:
            continue
        cor_hex = paleta_cores[idx % len(paleta_cores)]
        default_new_fills[nome] = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
        default_new_fonts[nome] = Font(color='FFFFFF', bold=True, size=8)

    # 2. Determinar as metas esperadas (targets) de cada cliente
    expected_targets = {}
    orig_counts = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (r, c) not in allowed_cells:
                continue
            val = ws.cell(row=r, column=c).value
            if val is not None:
                norm_val = normalize_val(val)
                if norm_val != "" and norm_val not in ('VAZIO', 'CT', 'CATRACA', 'SA', 'SALA', 'CW', 'COWORKING', '##'):
                    orig_counts[norm_val] = orig_counts.get(norm_val, 0) + 1
                    
    for client, qty in orig_counts.items():
        if client in parametros_premissas["reducoes"]:
            expected_targets[client] = qty - parametros_premissas["reducoes"][client]
        else:
            expected_targets[client] = qty
        
    # MODIFICAÇÃO: Deduz as mesas em branco das salas de reunião da contagem de mesas ativas exigidas
    for nc in parametros_premissas["novos_clientes"]:
        nome = nc["nome"]
        if nome in ambientes_fisicos_falhos:
            continue
        target_pas = nc["PAs"]
        for amb in criar_ambientes_solicitados:
            if amb.get("cliente_destinado") == nome:
                target_pas -= amb.get("sala_lugares", 0)
        expected_targets[nome] = target_pas
        
    # 3. Mapear as coordenadas das novas salas com base na presença física do cliente destino
    room_cells_by_client = {}
    if criar_ambientes_solicitados:
        import ScannerPremissas
        ScannerPremissas._orange_context_cache = {}
        macro_blocks_reconcile = scan_orange_context(dest_file, SHEET_NAME)
        for amb in criar_ambientes_solicitados:
            bloco_id = amb.get("bloco")
            cliente_dest = amb.get("cliente_destinado")
            
            block_match = re.search(r'\d+', str(bloco_id))
            if block_match:
                block_idx = int(block_match.group())
                if block_idx <= len(macro_blocks_reconcile):
                    block = macro_blocks_reconcile[block_idx - 1]
                    for env in block.get('ambientes', []):
                        client_presence = sum(
                            1 for coord in env['cells'] 
                            if new_ws.cell(row=coord[0], column=coord[1]).value == cliente_dest
                        )
                        if client_presence > 0:
                            room_cells_by_client.setdefault(cliente_dest, set()).update(env['cells'])

    # 4. Obter contagem atual na planilha modificada
    def get_actual_counts(sheet):
        counts = {}
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                if (r, c) not in allowed_cells:
                    continue
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    norm_val = normalize_val(val)
                    if norm_val != "" and norm_val not in ('VAZIO', 'CT', 'CATRACA', 'SA', 'SALA', 'CW', 'COWORKING', '##'):
                        counts[norm_val] = counts.get(norm_val, 0) + 1
        return counts

    actual_counts = get_actual_counts(new_ws)
    
    # 5. Ajustar excessos (limpar duplicatas criadas fora das salas físicas oficiais)
    for client, target in expected_targets.items():
        actual = actual_counts.get(client, 0)
        if actual > target:
            diff = actual - target
            client_cells = []
            for r in range(1, new_ws.max_row + 1):
                for c in range(1, new_ws.max_column + 1):
                    if (r, c) not in allowed_cells:
                        continue
                    val = new_ws.cell(row=r, column=c).value
                    if val is not None and normalize_val(val) == client:
                        dedicated_cells = room_cells_by_client.get(client, set())
                        if (r, c) not in dedicated_cells:
                            client_cells.append((r, c))
            
            for r, c in sorted(client_cells, key=lambda x: (x[1], x[0]))[:diff]:
                new_ws.cell(row=r, column=c).value = 'vazio'
                new_ws.cell(row=r, column=c).fill = FILL_LIBERADO
                new_ws.cell(row=r, column=c).font = FONT_SMALL
            
    # Re-computa contagens antes de suprir déficits
    actual_counts = get_actual_counts(new_ws)
    
    # 6. Ajustar déficits (alocar assentos em vazios aplicando os estilos corretos preservados)
    for client, target in expected_targets.items():
        actual = actual_counts.get(client, 0)
        if actual < target:
            diff = target - actual
            vazio_cells = []
            for r in range(1, new_ws.max_row + 1):
                for c in range(1, new_ws.max_column + 1):
                    if (r, c) not in allowed_cells:
                        continue
                    val = new_ws.cell(row=r, column=c).value
                    if (r, c) in salas_internas_cells:
                        continue
                    if val is None or normalize_val(val) in ('VAZIO', ''):
                        vazio_cells.append((r, c))
            
            vazio_cells.sort(key=lambda x: (x[1], x[0]))
            
            fill_to_apply = client_fills.get(client) or default_new_fills.get(client, FILL_NEW_CLIENT)
            font_to_apply = client_fonts.get(client) or default_new_fonts.get(client) or (FONT_WHITE if client.startswith("N_") else FONT_SMALL)
            
            for r, c in vazio_cells[:diff]:
                new_ws.cell(row=r, column=c).value = client
                new_ws.cell(row=r, column=c).fill = fill_to_apply
                new_ws.cell(row=r, column=c).font = font_to_apply
                
    # Salva a planilha reconciliada final
    new_wb.save(dest_file)
    print("⚖️ Inventário reconciliado com sucesso! Iniciando validação física...")

    # Re-scan dos blocos e paredes laranjas (ScannerPremissas) para forçar o remapeamento dinâmico
    import ScannerPremissas
    ScannerPremissas._orange_context_cache = {}
    
    # Ajusta os parâmetros de validação para que desconsidere as mesas em branco das salas de reunião
    adjusted_novos_clientes = []
    for nc in parametros_premissas["novos_clientes"]:
        nome = nc["nome"]
        if nome in ambientes_fisicos_falhos:
            continue
        pas_originais = nc["PAs"]
        # Deduz os assentos em branco das salas de reunião correspondentes
        for amb in criar_ambientes_solicitados:
            if amb.get("cliente_destinado") == nome:
                pas_originais -= amb.get("sala_lugares", 0)
        adjusted_novos_clientes.append({
            "nome": nome,
            "PAs": pas_originais
        })
        
    parametros_validacao = {
        "reducoes": parametros_premissas["reducoes"],
        "novos_clientes": adjusted_novos_clientes
    }
    
    ambientes_criados_str = "\n".join(ambientes_criados_info) if ambientes_criados_info else "Nenhum ambiente físico criado nesta rodada."
    erros_validacao = validar_inventario(ws, new_ws, allowed_cells, parametros_validacao)

    # Salva auditoria do posicionador
    salvar_auditoria("1_posicionador", (
        f"[PREMISSAS DO TXT]\n{pos_deps.premissas}\n\n"
        f"[BLOCOS INFO]\n{pos_deps.blocos_info}\n\n"
        f"[PLANT INFO]\n{pos_deps.plant_info}\n\n"
        f"[ACOES PLANEJADAS]\n{rascunho_layout_json}"
    ), res_posicionador.output, safe_name)

    # ══════════════════════════════════════════════════════════════════════════
    # ETAPA 2: Loop de Auto-Correção e Swaps via Organizador
    # ══════════════════════════════════════════════════════════════════════════
    print("2. Organizador iniciando loop de alinhamento e swaps...")

    for iteracao in range(1, 4):
        print(f"   -> Tentativa do Organizador {iteracao}/3...")
        
        try:
            # O scan de blocos físicos é refeito de forma dinâmica na planilha modificada pelo AmbienteBuilder
            current_wb = openpyxl.load_workbook(dest_file)
            current_ws = current_wb[SHEET_NAME]
            
            current_plant_data = scan_plant(current_ws, FORBIDDEN_PATTERNS)
            current_plant_info = build_plant_info(current_plant_data)
            
            # Chamada passando o caminho dinâmico do arquivo modificado
            current_blocos_info = build_blocos_info(
                current_plant_data, 
                current_ws.max_row, 
                current_ws.max_column, 
                current_ws, 
                file_path=dest_file
            )
        except Exception as e:
            print(f"\n❌ ERRO DETECTADO NA PREPARAÇÃO DA TENTATIVA {iteracao}:")
            traceback.print_exc()
            return
        
        org_deps = OrganizadorDeps(
            plant_info=current_plant_info,
            blocos_info=current_blocos_info,
            premissas=premissas_completas,
            rascunho_layout=rascunho_layout_json,
            ambientes_criados=ambientes_criados_str
        )
        
        feedback_erros = ""
        if erros_validacao:
            feedback_erros = (
                "\n\n⚠️ INCONSISTÊNCIAS DE SOMA OPERACIONAL NO LAYOUT ATUAL:\n"
                + "\n".join(f"- {err}" for err in erros_validacao)
                + "\n\nAplique ações de 'transferir' (permutas/swaps) para reequilibrar as quantidades exatas de volta!"
            )
        
        res_organizador = await organizador.run(
            f"Avalie a planta física atual. Se houver qualquer violação de premissas (como unificação, exclusividade, novo cliente desalinhado de sua sala física), utilize a função 'transferir' para corrigir. Se estiver tudo certo, retorne a lista vazia.{feedback_erros}",
            deps=org_deps
        )
        reorganizacao_json = clean_json_string(res_organizador.output)
        
        salvar_auditoria(f"2_organizador_iter{iteracao}", (
            f"[RASCUNHO DO LAYOUT]\n{org_deps.rascunho_layout}\n\n"
            f"[AMBIENTES RECENTEMENTE CRIADOS]\n{org_deps.ambientes_criados}\n\n"
            f"[PREMISSAS DO TXT]\n{org_deps.premissas}\n\n"
            f"[BLOCOS INFO ATUALIZADOS NA PLANILHA FISICA]\n{org_deps.blocos_info}\n\n"
            f"[PLANT INFO ATUALIZADA]\n{org_deps.plant_info}"
        ), res_organizador.output, safe_name)
        
        org_data = extrair_e_carregar_json(reorganizacao_json)
        acoes_org = org_data.get("acoes_organizacao", [])
        
        if not acoes_org and not erros_validacao:
            print(f"✓ Sistema auditado e validado. Todas as premissas e inventários estão plenamente satisfatórios!")
            break
            
        print(f"   ⚠ Validando ações corretivas em sandbox...")
        
        acoes_iteracao = []
        for ac in acoes_org:
            tipo = str(ac.get("tipo", "")).strip().lower()
            if tipo != "transferir":
                raise ValueError(
                    f"Acao invalida do organizador: '{tipo}'. "
                    "O organizador so pode transferir."
                )
            acoes_iteracao.append(Acao(**ac))
            
        acoes_iteracao = normalizar_acoes(acoes_iteracao, parametros_premissas["novos_clientes"])
            
        proposta_correcao = PropostaMock(f"Swaps Iteracao {iteracao}", "Baixo", acoes_iteracao)
        
        sandbox_wb, sandbox_ws = clone_ws(new_ws)
        log_iter, _ = execute_alocacao(sandbox_ws, proposta_correcao, current_plant_data, allowed_cells, file_path=dest_file)
        
        erros_validacao_temp = validar_inventario(ws, sandbox_ws, allowed_cells, parametros_validacao)
        
        if not erros_validacao_temp:
            print(f"   ✓ Correção da Iteração {iteracao} validada com sucesso! Consolidando alterações físicas...")
            # Consolida usando o caminho físico dinâmico
            log_iter, _ = execute_alocacao(new_ws, proposta_correcao, current_plant_data, allowed_cells, file_path=dest_file)
            
            for k, v in log_iter.get('realocadas', {}).items():
                log_acumulado.setdefault('realocadas', {}).setdefault(k, []).extend(v)
            for k, v in log_iter.get('liberadas', {}).items():
                log_acumulado.setdefault('liberadas', {})[k] = v
            log_acumulado.setdefault('avisos', []).extend(log_iter.get('avisos', []))
            acoes_totais.extend(acoes_iteracao)
            erros_validacao = []  
            
            # Salva o arquivo intermediário após a rodada bem sucedida
            new_wb.save(dest_file)
        else:
            print(f"   ❌ Correção inválida na Iteração {iteracao}. Executando rollback...")
            print("\n      ⚠️ DETALHES DAS INCONSISTÊNCIAS NO LAYOUT DO ORGANIZADOR:")
            for err in erros_validacao_temp:
                print(f"      -> {err}")
            print()
            erros_validacao = erros_validacao_temp  

    # ══════════════════════════════════════════════════════════════════════════
    # ETAPA 3: Finalização e Escrita de Relatórios
    # ══════════════════════════════════════════════════════════════════════════
    print("\n3. Gravando resultados e relatórios finais...")
    new_wb.save(dest_file)
    
    final_plant_data = scan_plant(new_ws, FORBIDDEN_PATTERNS)
    final_blocos_info = build_blocos_info(final_plant_data, new_ws.max_row, new_ws.max_column, new_ws, file_path=dest_file)
    
    print("\n================================================================================")
    print("ESTADO FINAL DOS BLOCOS APÓS TODAS AS MOVIMENTAÇÕES E CORREÇÕES:")
    print("================================================================================")
    print(final_blocos_info)
    
    proposta_final = PropostaMock(
        nome=pos_data.get("nome", "Reorganizacao Consolidada"),
        custo_obras="Baixo",
        acoes=acoes_totais
    )
    
    write_report(ws, new_ws, proposta_final, log_acumulado, "propostas/proposta_final_mudancas.txt")
    
    caminho_relatorio_final = "propostas/proposta_final_blocos_finais.txt"
    with open(caminho_relatorio_final, "w", encoding="utf-8") as f:
        f.write("================================================================================\n")
        f.write("RELATÓRIO DO ESTADO FINAL DOS BLOCOS APÓS AS MOVIMENTAÇÕES\n")
        f.write("================================================================================\n\n")
        f.write(final_blocos_info)
        
    print(f"✓ Processo concluído com sucesso. Arquivos gravados em 'propostas/'.")

if __name__ == '__main__':
    asyncio.run(main())
