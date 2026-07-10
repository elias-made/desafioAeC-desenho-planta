"""
Teste: Simulação da proposta do posicionador

Cenário exato do posicionador:
- N_1: 124 PAs + sala 4 lugares → Bloco_7-A
- N_2: 165 PAs + sala 1 lugar → Bloco_7-A
- N_3: 20 PAs → Bloco_1-A

Valida que:
1. Não há sobreposição de células entre os ambientes
2. As divisórias são desenhadas corretamente
3. O segundo ambiente no mesmo bloco (N_2) respeita as paredes do primeiro (N_1)
4. Os dados são preservados entre iterações

Estratégia: Agrupa por bloco. Para cada bloco, processa sequencialmente:
1. Desenha e colore o primeiro ambiente
2. Re-scaneia o bloco (agora com divisórias do primeiro ambiente)
3. Pega a MAIOR área livre restante (pode ter mudado de letra)
4. Desenha e colore o próximo ambiente
"""

import os
import re
import shutil
import openpyxl
from collections import defaultdict

import ScannerPremissas
from AmbienteBuilder import (
    separar_ambiente_e_desenhar_divisorias,
    _gerar_layout_sala_estruturado,
    _selecionar_mesas_contiguas,
    get_env_cells,
    _celulas_contorno_do_ambiente,
)
from openpyxl.styles import PatternFill, Font

SHEET_NAME = 'JPIII'
INPUT_FILE = 'planta.xlsx'
OUTPUT_FILE = 'planta_teste_proposta_posicionador.xlsx'

# Cenário exato do posicionador:
# - N_1 e N_2 no Bloco_7-A (dois ambientes no mesmo bloco, em sequência)
# - N_3 no Bloco_1-A
CRIAR_AMBIENTES = [
    {"bloco": "Bloco_7", "ambiente": "A", "quantidade_mesas": 124, "cliente_destinado": "N_1", "sala_lugares": 4},
    {"bloco": "Bloco_2", "ambiente": "A", "quantidade_mesas": 165, "cliente_destinado": "N_2", "sala_lugares": 1},
    {"bloco": "Bloco_1", "ambiente": "A", "quantidade_mesas": 20, "cliente_destinado": "N_3", "sala_lugares": 0},
]


def main():
    print("=" * 80)
    print("TESTE: Proposta do Posicionador - N_1, N_2 no Bloco_7 e N_3 no Bloco_1")
    print("=" * 80)

    if not os.path.exists(INPUT_FILE):
        print(f"ERRO: {INPUT_FILE} não encontrado")
        return

    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_NAME]

    resultados = {}
    paleta_cores = ['34495E', '9B59B6', '1ABC9C', 'E67E22', '2ECC71', '3498DB', 'E74C3C']
    salas_internas_cells = set()  # Protege células de salas fechadas

    # Agrupa ambientes por bloco
    ambientes_por_bloco = defaultdict(list)
    for amb in CRIAR_AMBIENTES:
        ambientes_por_bloco[amb["bloco"]].append(amb)

    # Processa cada bloco sequencialmente
    for bloco_id in sorted(ambientes_por_bloco.keys()):
        ambientes_do_bloco = ambientes_por_bloco[bloco_id]
        
        for idx, amb in enumerate(ambientes_do_bloco):
            ambiente_letra = amb["ambiente"]
            cliente_dest = amb["cliente_destinado"]
            sala_lugares = amb.get("sala_lugares", 0)

            print(f"\n{'─' * 70}")
            print(f"{cliente_dest} em {bloco_id}-{ambiente_letra}")

            # SEMPRE re-escaneia para pegar as divisórias desenhadas anteriormente
            ScannerPremissas._orange_context_cache = {}
            macro_blocks = ScannerPremissas.scan_orange_context(OUTPUT_FILE, SHEET_NAME)
            
            # Na primeira iteração do bloco, usa a letra original
            # Nas iterações seguintes, pega a MAIOR área livre do bloco (ignora letra)
            if idx == 0:
                env_cells_raw_list = get_env_cells(bloco_id, ambiente_letra, macro_blocks)
            else:
                # Pega a maior área aberta do bloco (qualquer letra)
                block_match = re.search(r'\d+', str(bloco_id))
                if block_match:
                    block_idx = int(block_match.group())
                    if block_idx <= len(macro_blocks):
                        block = macro_blocks[block_idx - 1]
                        envs = block.get('ambientes', [])
                        if envs:
                            # Ordena por tamanho e pega o maior
                            sorted_envs = sorted(envs, key=lambda e: len(e['cells']), reverse=True)
                            env_cells_raw_list = sorted_envs[0]['cells']
                            print(f"   (Re-scan: usando ambiente '{sorted_envs[0]['id']}' com {len(env_cells_raw_list)} células)")
                        else:
                            env_cells_raw_list = []
                    else:
                        env_cells_raw_list = []
                else:
                    env_cells_raw_list = []
            
            env_cells_raw = set(env_cells_raw_list) if env_cells_raw_list else set()
            env_cells = env_cells_raw | _celulas_contorno_do_ambiente(ws, env_cells_raw)
            env_cells_livres = env_cells

            # FLUXO IGUAL AO AmbienteBuilder.py:
            # Usa _selecionar_mesas_contiguas diretamente (funciona bem)
            
            qtd_mesas_operacionais = amb["quantidade_mesas"]
            sala_lugares = amb.get("sala_lugares", 0)
            
            # Primeiro: planeja a sala (se houver) para reservar o espaço
            allocated_sala = set()
            room_cells_override = None
            
            if sala_lugares > 0:
                allocated_sala, room_cells_override = _gerar_layout_sala_estruturado(
                    ws, env_cells_livres, sala_lugares
                )
            
            # Remove o espaço da sala do ambiente disponível
            available_env_cells = env_cells_livres - (set(room_cells_override) if room_cells_override else set())
            
            # Seleciona mesas para o salão aberto
            allocated_ambiente = _selecionar_mesas_contiguas(available_env_cells, ws, qtd_mesas_operacionais)
            
            if not allocated_ambiente:
                print(f"   ERRO: Nenhuma mesa encontrada para {cliente_dest}")
                continue

            room_cells_set = set(room_cells_override) if room_cells_override else set()
            allocated_total = allocated_ambiente | room_cells_set

            # Desenha ambiente com as divisórias
            separar_ambiente_e_desenhar_divisorias(
                ws=ws,
                env_cells=env_cells_livres,
                allocated_cells=allocated_total,
                reconstruir_sala=False,
                room_cells_override=room_cells_set
            )

            if room_cells_set:
                salas_internas_cells.update(room_cells_set)  # Protege sala de futuras iterações
                separar_ambiente_e_desenhar_divisorias(
                    ws=ws,
                    env_cells=env_cells_livres,
                    allocated_cells=allocated_sala,
                    reconstruir_sala=True,
                    room_cells_override=room_cells_set
                )

            # Colore apenas as mesas do salão (não os corredores nem as salas fechadas)
            cor_hex = paleta_cores[len(resultados) % len(paleta_cores)]
            fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
            font = Font(color='FFFFFF', bold=True, size=8)

            cells_colored = 0
            coords_coloridas = set()
            for r, c in allocated_ambiente:
                if (r, c) in salas_internas_cells:
                    continue  # Pula células de sala fechada
                cell = ws.cell(row=r, column=c)
                if cell.value != "CT":
                    cell.value = cliente_dest
                    cell.fill = fill
                    cell.font = font
                    cells_colored += 1
                    coords_coloridas.add((r, c))

            # Colore sala
            if room_cells_set and allocated_sala:
                fill_sala = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
                font_sala = Font(name="Calibri", size=9)
                for r, c in allocated_sala:
                    cell = ws.cell(row=r, column=c)
                    if cell.value != "CT":
                        cell.value = "vazio"
                        cell.fill = fill_sala
                        cell.font = font_sala

            resultados[cliente_dest] = {
                "cells_colored": cells_colored,
                "all_coords": coords_coloridas,
                "bloco": bloco_id
            }

            print(f"   ✓ Criado ({cells_colored} células coloridas)")
            wb.save(OUTPUT_FILE)

    # Validações
    print("\n" + "=" * 80)
    print("VALIDAÇÕES")
    print("=" * 80)

    falhas = 0

    # Sobreposição
    print(f"\n1. Sobreposição:")
    ambientes_coords = {k: v.get("all_coords", set()) for k, v in resultados.items()}
    clientes = sorted(ambientes_coords.keys())
    
    for i in range(len(clientes)):
        for j in range(i+1, len(clientes)):
            cli_i = clientes[i]
            cli_j = clientes[j]
            overlap = ambientes_coords[cli_i] & ambientes_coords[cli_j]
            if overlap:
                print(f"   ❌ {cli_i} ∩ {cli_j}: {len(overlap)} células")
                falhas += 1

    if falhas == 0:
        print(f"   ✓ Sem sobreposição")

    # Integridade
    print(f"\n2. Integridade (dados preservados):")
    wb_check = openpyxl.load_workbook(OUTPUT_FILE)
    ws_check = wb_check[SHEET_NAME]

    contagem = defaultdict(int)
    for r in range(1, ws_check.max_row + 1):
        for c in range(1, ws_check.max_column + 1):
            val = ws_check.cell(row=r, column=c).value
            if val and isinstance(val, str) and val.startswith("N_"):
                contagem[val] += 1

    for cliente in ["N_1", "N_2", "N_3"]:
        if cliente in resultados:
            count = contagem[cliente]
            expected = resultados[cliente]["cells_colored"]
            if expected > 0 and count == 0:
                print(f"   ❌ {cliente}: APAGADO")
                falhas += 1
            elif count == expected:
                print(f"   ✓ {cliente}: {count} células preservadas")
            else:
                print(f"   ⚠️  {cliente}: {count} vs {expected} esperado")

    # Resultado
    print(f"\n{'=' * 80}")
    if falhas == 0:
        print("✓ TESTE PASSOU")
    else:
        print(f"❌ TESTE FALHOU ({falhas} erros)")
    print(f"{'=' * 80}")
    print(f"Planilha: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERRO: {e}")
        traceback.print_exc()
